import os
import sys
import datetime
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
from selenium.webdriver.common.keys import Keys
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from openpyxl.styles import Font
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class TagUnlink(unittest.TestCase):
    """
    Automates the Tag Unlink flow under Tagging > Tag Unlink
    """
    def __init__(self, driver):
        super().__init__('test_tag_unlink')
        self.driver = driver
        self.wait = WebDriverWait(driver, 15)
        self.fc = Function_Call(driver)

    def _take_screenshot(self, name):
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TagUnlink_{name}_{timestamp}.png"
            os.makedirs("Screenshots/TagUnlink", exist_ok=True)
            path = os.path.join("Screenshots/TagUnlink", filename)
            self.driver.save_screenshot(path)
            print(f"📸 Screen saved: {filename}")
        except Exception as e:
            print(f"⚠️ Screenshot Failed: {e}")

    def test_tag_unlink(self):
        driver, wait = self.driver, self.wait

        print("🧭 Navigating to Tag Unlink...")
        try:
            if "admin_ret_tagging/tagging/tag_unlink" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                self.fc.click(self, "//span[contains(text(), 'Inventory')]")
                sleep(1)
                self.fc.click(self, "(//span[contains(normalize-space(), 'Tag Unlink')])")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            pass
        
        driver.get(BASE_URL + "index.php/admin_ret_tagging/tagging/tag_unlink")
        sleep(2)

        sheet_name = "TagUnlink"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 2} test cases in '{sheet_name}' sheet\n")
        except Exception as e:
            print(f"❌ Failed to read valid rows from '{sheet_name}': {e}")
            return []

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]

            data_map = {
                "TestCaseId":  1, "TestStatus": 2, "ActualStatus":3,
                "Branch":      4, "TagCode":    5, "OldTagCode": 6, 
                "ExpectedMsg": 7, "Remarks":    8
            }

            row_data = {
                key: sheet.cell(row=row_num, column=col).value
                for key, col in data_map.items()
            }
            workbook.close()

            print(f"\n{'='*80}")
            tagcode_log = row_data.get('TagCode') or row_data.get('OldTagCode')
            print(f"🧪 TC: {row_data.get('TestCaseId')}  |  Tag: {tagcode_log}")
            print(f"{'='*80}")

            try:
                result = self._run_tag_unlink(row_data)
                print(f"🏁 Result: {result[0]} — {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
            except Exception as e:
                print(f"❌ TC {row_data.get('TestCaseId')} exception: {e}")
                self._take_screenshot(f"EX_{row_data.get('TestCaseId')}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)

    def _run_tag_unlink(self, row_data):
        driver, wait = self.driver, self.wait
        current_field = "Initialization"
        
        try:
            # 1. Select Branch
            branch = str(row_data.get("Branch", "")).strip()
            if branch:
                current_field = "Branch"
                self.fc.dropdown_select2(
                    '//select[@id="branch_select"]/following-sibling::span',
                    branch,
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 2. Search Tag / Old Tag
            tag_val = row_data.get("TagCode")
            tag_code = str(tag_val).strip() if tag_val is not None else ""
            
            old_tag_val = row_data.get("OldTagCode")
            old_tag_code = str(old_tag_val).strip() if old_tag_val is not None else ""

            if not tag_code and not old_tag_code:
                raise ValueError("Both TagCode and OldTagCode cannot be empty.")
            
            if tag_code:
                current_field = "Search Tag Code"
                search_tc_input = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@id="tag_unlink"]')))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", search_tc_input)
                search_tc_input.clear()
                search_tc_input.send_keys(tag_code)
                sleep(0.5)
                # Press ENTER to forcefully trigger the table load if standard blur isn't enough
                search_tc_input.send_keys(Keys.ENTER)
                print(f"  -> Searched Tag Code: {tag_code}")
                
            elif old_tag_code:
                current_field = "Search Old Tag Code"
                search_old_input = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@id="old_tag_unlink"]')))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", search_old_input)
                search_old_input.clear()
                search_old_input.send_keys(old_tag_code)
                sleep(0.5)
                search_old_input.send_keys(Keys.ENTER)
                print(f"  -> Searched Old Tag Code: {old_tag_code}")
            
            sleep(2)
            # Wait for table
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[contains(@class,"dataTable") or contains(@class,"table")]/tbody/tr')))

            current_field = "Table Row Processing"
            self._take_screenshot(f"TableLoaded_{row_data.get('TestCaseId')}")

            # Make sure main checkbox is selected
            checkboxes = driver.find_elements(By.XPATH, '//table/tbody/tr//input[@type="checkbox"]')
            if checkboxes:
                cb = checkboxes[0]
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
            else:
                print("  -> No checkbox found in the table. The table may be unexpectedly empty.")
            
            # 3. Save All
            current_field = "Save All Button"
            sleep(1)
            save_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "save all")] | //button[@id="save_link"] | //button[@id="save_unlink"]')))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", save_btn)
            sleep(0.5)
            driver.execute_script("arguments[0].click();", save_btn)

            # 4. Verify Success
            current_field = "Success Message Banner"
            expected_msg = str(row_data.get("ExpectedMsg", "")).strip()
            
            try:
                banner = wait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "alert-success")] | //div[@class="toast-message"]')))
                actual_msg = banner.text.strip()
                driver.execute_script("arguments[0].scrollIntoView(true);", banner)
                self._take_screenshot(f"Success_{row_data.get('TestCaseId')}")
                print(f"  -> Validated Banner: {actual_msg}")

                if expected_msg and expected_msg.lower() not in actual_msg.lower():
                    return ("Warning", f"Expected '{expected_msg}', got '{actual_msg}'")
                return ("Pass", actual_msg)
            except TimeoutException:
                self._take_screenshot(f"NoBanner_{row_data.get('TestCaseId')}")
                return ("Fail", "Success banner did not appear after save.")

        except Exception as e:
            print(f"❌ Error at '{current_field}': {e}")
            self._take_screenshot(f"Error_{row_data.get('TestCaseId', 'Unknown')}")
            return ("Fail", f"Error in [{current_field}]: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Write Pass/Fail + actual status back to Excel."""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
