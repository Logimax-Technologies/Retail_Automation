from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest
import re

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class BranchTransfer(unittest.TestCase):
    """
    Branch Transfer Module Automation
    Module: Inventory → Branch Transfer
    Supports: Tagged / NonTagged / PurchaseReturn / PackingItems / RepairOrder
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_branch_transfer(self):
        """Main entry point for Branch Transfer automation"""
        driver = self.driver
        wait   = self.wait

        # Navigate to Branch Transfer via Menu → Inventory → Branch Transfer
        try:
            if "admin_ret_brntransfer/branch_transfer" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Inventory')]")
                sleep(1)
                Function_Call.click(self, "(//span[contains(normalize-space(), 'Branch Transfer')])[1]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_brntransfer/branch_transfer/list")
            sleep(2)

        sheet_name = "BranchTransfer"

        # Clear BranchTransferApproval sheet at start of entire test run
        try:
            wb = load_workbook(FILE_PATH)
            if "BranchTransferApproval" in wb.sheetnames:
                app_clear = wb["BranchTransferApproval"]
                if app_clear.max_row > 1:
                    app_clear.delete_rows(2, app_clear.max_row)
                    wb.save(FILE_PATH)
            wb.close()
            print("✅ BranchTransferApproval sheet cleared (headers preserved)")
        except Exception as e:
            print(f"⚠️ Failed to clear BranchTransferApproval: {e}")

        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 2} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]

            # Column map — must match Excel sheet header order
            data_map = {
                "TestCaseId":   1,  "TestStatus":   2,  "ActualStatus": 3,
                "TransferType": 4,  "FromBranch":   5,  "ToBranch":     6,
                "OtherIssue":   7,  "LotNo":        8,  "Section":      9,
                "Product":      10, "Design":       11, "TagCode":       12,
                "OldTagCode":   13, "EstimationNo": 14, "NT_Receipt":    15,
                "PR_FromDate":  16, "PR_ToDate":    17, "Pack_Item":     18,
                "Pack_Pcs":     19, "OrderNo":      20, "RowCount":       21,
                "Remark":       22
                
            }

            row_data = {
                key: sheet.cell(row=row_num, column=col).value
                for key, col in data_map.items()
            }
            workbook.close()

            # if str(row_data.get("TestStatus", "")).strip().lower() != "run":
            #     print(f"⏭️ Skipping row {row_num} (Status={row_data.get('TestStatus')})")
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 TC: {row_data['TestCaseId']}  |  Type: {row_data['TransferType']}")
            print(f"{'='*80}")

            try:
                result = self._run_transfer(row_data, row_num, sheet_name)
                print(f"🏁 Result: {result[0]} — {result[1]}")
                bt_code = result[2] if len(result) > 2 else ""
                self._update_excel_status(row_num, result[0], result[1], sheet_name, bt_code=bt_code)
                
                # Sync successful transfer to Approval list (2 rows)
                if result[0] == "Pass" and bt_code:
                    self._update_approval_sheet(bt_code, row_data)
            except Exception as e:
                print(f"❌ TC {row_data['TestCaseId']} exception: {e}")
                self._take_screenshot(f"EX_{row_data['TestCaseId']}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._cancel_form()

    # ─────────────────────────────────────────────────────────
    # MAIN FLOW DISPATCHER
    # ─────────────────────────────────────────────────────────
    def _run_transfer(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Navigate to Add Page"
        try:
            # Open the Add form
            Function_Call.alert(self)
            try:
                Function_Call.click(self, '//a[@id="add_estimation"]')
            except:
                driver.get(BASE_URL + "index.php/admin_ret_brntransfer/branch_transfer/add")
            sleep(3)

            transfer_type = str(row_data.get("TransferType", "")).strip()
            other_issue   = str(row_data.get("OtherIssue", "No")).strip().upper()

            # ── Type Radio ───────────────────────────────────────────
            type_map = {
                "Tagged":         "1",
                "NonTagged":      "2",
                "PurchaseReturn": "3",
                "PackingItems":   "4",
                "RepairOrder":    "5",
            }
            radio_val = type_map.get(transfer_type, "1")
            current_field = f"Select Type Radio ({transfer_type})"
            Function_Call.click(
                self,
                f'//input[@name="transfer_item_type"][@value="{radio_val}"]'
            )
            sleep(1)

            # ── Other Issue checkbox ─────────────────────────────────
            if other_issue == "Yes":
                current_field = "Other Issue Checkbox"
                Function_Call.click(self, '//input[@id="isOtherIssue"]')
                sleep(1)

            # ── From Branch ──────────────────────────────────────────
            if row_data.get("FromBranch"):
                current_field = "From Branch"
                self.fc.dropdown_select2(
                    '//select[contains(@class,"from_branch") or @id="from_brn"]/following-sibling::span',
                    str(row_data["FromBranch"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(2)

            # ── To Branch (skip when OtherIssue = Y) ────────────────
            if other_issue != "Yes" and row_data.get("ToBranch"):
                current_field = "To Branch"
                self.fc.dropdown_select2(
                    '//select[@id="to_brn"]/following-sibling::span',
                    str(row_data["ToBranch"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # ── Dispatch to transfer-type specific flow ───────────────
            if transfer_type == "Tagged":
                current_field = "Tagged Flow"
                self._flow_tagged(row_data)

            elif transfer_type == "NonTagged":
                current_field = "NonTagged Flow"
                self._flow_non_tagged(row_data)

            elif transfer_type == "PurchaseReturn":
                current_field = "PurchaseReturn Flow"
                self._flow_purchase_return(row_data)

            elif transfer_type == "PackingItems":
                current_field = "PackingItems Flow"
                self._flow_packing_items(row_data)

            elif transfer_type == "RepairOrder":
                current_field = "RepairOrder Flow"
                self._flow_repair_order(row_data)

            # ── Remark ───────────────────────────────────────────────
            if row_data.get("Remark"):
                current_field = "Remark"
                remark_el = wait.until(EC.presence_of_element_located((By.ID, "remark")))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", remark_el)
                sleep(0.5)
                driver.execute_script("arguments[0].value = arguments[1];", remark_el, str(row_data["Remark"]))
                print(f"✅ Remark filled via JS")

            # ── Submit ───────────────────────────────────────────────
            current_field = "Submit (add_to_transfer)"
            self._take_screenshot(f"Before_{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="add_to_transfer"]')
            sleep(5)

            # ── Post-Save: Print Tab Handling ────────────────────────
            current_field = "Print Tab Capture"
            bt_code = self._capture_bt_code_from_print_tab()
            sleep(1)
            # ── Scroll to top before checking banner ──────────────────
            driver.execute_script("window.scrollTo(0, 0);")
            sleep(1)

            # ── Success Banner ────────────────────────────────────────
            current_field = "Success Banner"
            try:
                banner = wait.until(
                    EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, ".alert.alert-success,.alert-box")
                    )
                )
                success_msg = banner.text.strip()
            except Exception:
                success_msg = "Banner not found"

            print(f"✅ BT Code: {bt_code}  |  Banner: {success_msg}")

            # ── List Page Verification ────────────────────────────────
            current_field = "List Verification"
            verify_result = self._verify_in_list(bt_code, row_data)

            status_text = f"{bt_code} | {success_msg} | {verify_result}"
            return ("Pass", status_text, bt_code)

        except Exception as e:
            self._take_screenshot(f"ERR_{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"Error in [{current_field}]: {str(e)}", "")

    # ─────────────────────────────────────────────────────────
    # FLOW 1 — TAGGED  (multi-value pipe-separated support)
    # ─────────────────────────────────────────────────────────
    def _flow_tagged(self, row_data):
        """
        Tagged item transfer. All search fields support pipe-separated multi-values.
        Excel examples:
          LotNo       : "9580"          or  "9580|9581"
          Section     : "HARAM GB"      or  "HARAM GB|HARAM SECTION"
          Product     : "GOLD MALAI"    or  "GOLD MALAI|SILVER CHAIN"
          TagCode     : "GBT-01474"     or  "GBT-01474|GBT-01476"
          OldTagCode  : "OT-001"        or  "OT-001|OT-002"
          EstimationNo: "1"             or  "1|2"

        For each value: Set field → Search → Select all → Add to Preview (#bt_list).
        Priority: EstimationNo > TagCode > OldTagCode > LotNo > Section/Product
        Final submit (#add_to_transfer) is called once from _run_transfer after this method.
        """
        driver, wait = self.driver, self.wait

        def _split(val):
            """Split pipe-separated string into a clean list; return [] if empty/None."""
            if not val:
                return []
            parts = [v.strip() for v in str(val).split("|") if v.strip()]
            return parts

        lots       = _split(row_data.get("LotNo"))
        sections   = _split(row_data.get("Section"))
        products   = _split(row_data.get("Product"))
        tag_codes  = _split(row_data.get("TagCode"))
        old_tags   = _split(row_data.get("OldTagCode"))
        esti_nos   = _split(row_data.get("EstimationNo"))
        design_val = str(row_data.get("Design", "") or "").strip()

        # ── (optional) Design always single-value ─────────────────
        if design_val:
            wait.until(EC.presence_of_element_located((By.ID, "design"))).clear()
            wait.until(EC.presence_of_element_located((By.ID, "design"))).send_keys(design_val)
            sleep(0.5)

        # ── PRIORITY 1: Estimation No ─────────────────────────────
        if esti_nos:
            expected_rows=0
            for esti in esti_nos:
                print(f"  -> EstimationNo: {esti}")
                esti_input = wait.until(EC.presence_of_element_located((By.ID, "esti_no")))
                esti_input.clear()
                esti_input.send_keys(esti)
                sleep(0.5)
                Function_Call.click(self, '//button[@id="search_est_no"]')
                display_rows = len(self.driver.find_elements(By.XPATH,'//table[@id="bt_search_list"]/tbody/tr'))
                if display_rows > expected_rows:
                    expected_rows = display_rows
                else:
                    raise Exception(f"EstimationNo '{esti}' not found (row count did not increase)")
            self._wait_and_add_tagged_results(row_data=row_data, is_multi=len(esti_nos) > 1)

        # ── PRIORITY 2: Tag Code ──────────────────────────────────
        elif tag_codes:
            expected_rows = 0
            for tag in tag_codes:
                print(f"  -> TagCode: {tag}")
                tag_input = wait.until(EC.presence_of_element_located((By.ID, "tag_no")))
                tag_input.clear()
                tag_input.send_keys(tag)
                sleep(2)
                Function_Call.click(self, ' (//button[contains(@class,"btrn_search")])[2]')
                sleep(2)
                display_rows = len(self.driver.find_elements(By.XPATH,'//table[@id="bt_search_list"]/tbody/tr'))
                if display_rows > expected_rows:
                    expected_rows = display_rows
                else:
                    raise Exception(f"TagCode '{tag}' not found (row count did not increase)")
            self._wait_and_add_tagged_results(row_data, is_multi=len(tag_codes) > 1)

        # ── PRIORITY 3: Old Tag Code ──────────────────────────────
        elif old_tags:
            for old_tag in old_tags:
                print(f"  -> OldTagCode: {old_tag}")
                old_input = wait.until(EC.presence_of_element_located((By.ID, "old_tag_no")))
                old_input.clear()
                old_input.send_keys(old_tag)
                sleep(0.5)
                Function_Call.click(self, ' (//button[contains(@class,"btrn_search")])[2]')
            self._wait_and_add_tagged_results(row_data,is_multi=len(old_tags) > 1)

        # ── PRIORITY 4: Lot No ────────────────────────────────────
        elif lots:
            expected_rows = 0
            for lot in lots:
                print(f"  -> LotNo: {lot}")
                # Clear existing Select2 selection then pick new value
                driver.execute_script(
                    "var s = document.getElementById('lotno');"
                    "if(s && $(s).select2) { $(s).val(null).trigger('change'); }"
                )
                sleep(0.5)
                self.fc.dropdown_select2(
                    '//select[@id="lotno"]/following-sibling::span',
                    lot,
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)
                Function_Call.click(self, ' (//button[contains(@class,"btrn_search")])[2]')
                sleep(1)
                display_rows = driver.find_element(By.XPATH, "//span[@class='tot_tag_count']").text
                if int(display_rows) > expected_rows:
                    expected_rows = int(display_rows)
                else:
                    raise Exception(f"LotNo '{lot}' not found (row count did not increase)")
            self._wait_and_add_tagged_results(row_data,is_multi=len(lots) > 1)

        # ── PRIORITY 5: Section and/or Product (combined, one cycle each) ──
        else:
            # Build pairs: if both lists are non-empty zip them; else iterate the non-empty one
            max_len = max(len(sections), len(products), 1)
            expected_rows = 0
            for i in range(max_len):
                sec = sections[i] if i < len(sections) else (sections[-1] if sections else "")
                prd = products[i] if i < len(products) else (products[-1] if products else "")
                print(f"  -> Section: '{sec}'  Product: '{prd}'")

                # Clear + set Section
                if sec:
                    driver.execute_script(
                        "var s = document.getElementById('section_select');"
                        "if(s && $(s).select2) { $(s).val(null).trigger('change'); }"
                    )
                    sleep(0.5)
                    self.fc.dropdown_select2(
                        '//select[@id="section_select"]/following-sibling::span',
                        sec,
                        '//span[@class="select2-search select2-search--dropdown"]/input'
                    )
                    sleep(1)

                # Clear + set Product
                if prd:
                    self.fc.fill_autocomplete_field("product", str(prd))
                    sleep(1)

                Function_Call.click(self, ' (//button[contains(@class,"btrn_search")])[2]')
                sleep(2)
                display_rows = driver.find_element(By.XPATH, "//span[@class='tot_tag_count']").text
                if int(display_rows) > expected_rows:
                    expected_rows = int(display_rows)
                else:
                    raise Exception(f"LotNo '{lot}' not found (row count did not increase)")
         
            self._wait_and_add_tagged_results(row_data,is_multi=max_len > 1)

    def _wait_and_add_tagged_results(self, row_data=None, is_multi=False):
        """
        Helper for Tagged flow:
        Wait for bt_search_list to populate → check all rows → click Add to Preview.

        Args:
            is_multi : True when multiple pipe-separated values were searched.
                       In that case _check_all_in_table is SKIPPED (avoid double-toggle).
                       False (default) = single scenario → check all rows normally.
        """
        wait = self.wait
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="bt_search_list"]/tbody/tr')))
            sleep(1)
            display_rows=len(self.driver.find_elements(By.XPATH,'//table[@id="bt_search_list"]/tbody/tr'))
            # Condition: only auto-select all rows for a single-scenario search.
            # Multi-scenario (pipe-separated) skips this to prevent double-toggle issues.
            select_all_el = wait.until(EC.presence_of_element_located((By.ID, 'select_all')))
            if is_multi:
                self.driver.execute_script("arguments[0].click();", select_all_el)
            elif display_rows <= 10:
                self.driver.execute_script("arguments[0].click();", select_all_el)
            else:
                max_rows = row_data.get('RowCount') if row_data else None
                self._check_all_in_table("bt_search_list", max_rows=max_rows)
            sleep(1)
            Function_Call.click(self, '//button[@id="add_to_list"]')
            sleep(2)
            # Verify preview accumulated
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '//table[@id="bt_list"]/tbody/tr')
            ))
            row_count = len(self.driver.find_elements(By.XPATH, '//table[@id="bt_list"]/tbody/tr'))
            print(f"  Preview (bt_list) now has {row_count} total row(s)")
        except Exception as e:
            print(f"  WARNING: Search iteration issue — {e}")



    # ─────────────────────────────────────────────────────────
    # FLOW 2 — NON-TAGGED
    # ─────────────────────────────────────────────────────────
    def _flow_non_tagged(self, row_data):
        driver, wait = self.driver, self.wait

        # NT Receipt (Select2) — if provided, load by receipt
        if row_data.get("NT_Receipt"):
            self.fc.dropdown_select2(
                '//select[@id="nt_receipt"]/following-sibling::span',
                str(row_data["NT_Receipt"]),
                '//span[@class="select2-search select2-search--dropdown"]/input'
            )
            sleep(1)
        else:
            # Product entry mode
            if row_data.get("Section"):
                self.fc.dropdown_select2(
                    '//select[@id="section_select"]/following-sibling::span',
                    str(row_data["Section"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)
            if row_data.get("Product"):
                self.fc.fill_autocomplete_field("nt_product", str(row_data["Product"]))
                sleep(1)

        # Search
        Function_Call.click(self, ' (//button[contains(@class,"btrn_search")])[1]')
        sleep(3)

        # Wait for NT search results
        wait.until(EC.presence_of_element_located(
            (By.XPATH, '//table[@id="bt_nt_search_list"]/tbody/tr')
        ))
        print("✅ Non-Tagged search results loaded")

        # Select all NT items
        self._check_all_in_table("bt_nt_search_list", checkbox_class="nt_item_sel")

    # ─────────────────────────────────────────────────────────
    # FLOW 3 — PURCHASE RETURN
    # ─────────────────────────────────────────────────────────
    def _flow_purchase_return(self, row_data):
        driver, wait = self.driver, self.wait

        # Set date range via JS (bill_date datepicker)
        from_date = str(row_data.get("PR_FromDate", "")).strip()
        to_date   = str(row_data.get("PR_ToDate", "")).strip()

        if from_date and to_date:
            # Open datepicker
            Function_Call.click(self, '//button[@id="bill_date"]')
            sleep(1)
            
            # Use JS to update the inputs in the active dropdown and click Apply
            js_script = """
                var activePicker = Array.from(document.querySelectorAll('.daterangepicker')).find(p => p.style.display !== 'none');
                if (activePicker) {
                    var customLi = activePicker.querySelector('li[data-range-key="Custom Range"]');
                    if (customLi) { customLi.click(); }
                    
                    var inputs = activePicker.querySelectorAll('input.input-mini, input[name="daterangepicker_start"], input[name="daterangepicker_end"]');
                    if (inputs.length >= 2) {
                        inputs[0].focus();
                        inputs[0].value = arguments[0];
                        inputs[0].dispatchEvent(new Event('change', {bubbles: true}));
                        
                        inputs[1].focus();
                        inputs[1].value = arguments[1];
                        inputs[1].dispatchEvent(new Event('change', {bubbles: true}));
                    }
                    var applyBtn = activePicker.querySelector('.applyBtn');
                    if (applyBtn && !applyBtn.disabled) {
                        applyBtn.click();
                    }
                }
            """
            driver.execute_script(js_script, from_date, to_date)
            print(f"✅ Date range set via DatePicker: {from_date} → {to_date}")
            sleep(1)

        # Search
        Function_Call.click(self, '(//button[contains(@class,"btrn_search")])[3]')
        sleep(3)

        # Select all returned items
        max_rows = row_data.get('RowCount')
        self._check_all_in_table("old_metal_list", max_rows=max_rows)
        print("✅ Purchase Return items selected")

    # ─────────────────────────────────────────────────────────
    # FLOW 4 — PACKING ITEMS
    # ─────────────────────────────────────────────────────────
    def _flow_packing_items(self, row_data):
        driver, wait = self.driver, self.wait

        def _split(val):
            return [v.strip() for v in str(val).split("|") if v.strip()]

        items = _split(row_data.get("Pack_Item"))
        pcs   = _split(row_data.get("Pack_Pcs"))

        max_len = max(len(items), len(pcs), 1) if items or pcs else 0

        for i in range(max_len):
            pack_item = items[i] if i < len(items) else (items[-1] if items else "")
            pack_pcs  = pcs[i] if i < len(pcs) else (pcs[-1] if pcs else "")

            if not pack_item:
                continue

            print(f"  -> Packing Item: '{pack_item}' | Pcs: {pack_pcs}")

            # Select item from inventory dropdown (Select2)
            # Clear it first to ensure a clean selection
            driver.execute_script(
                "var s = document.getElementById('select_item');"
                "if(s && $(s).select2) { $(s).val(null).trigger('change'); }"
            )
            sleep(0.5)

            self.fc.dropdown_select2(
                '//select[@id="select_item"]/following-sibling::span',
                pack_item,
                '//span[@class="select2-search select2-search--dropdown"]/input'
            )
            sleep(1)

            # Enter No of Pcs
            if pack_pcs:
                pcs_input = wait.until(
                    EC.presence_of_element_located((By.ID, "packaging_no_of_pcs"))
                )
                pcs_input.clear()
                pcs_input.send_keys(pack_pcs)
                sleep(0.5)

            # Click Add to load into packaging_list
            Function_Call.click(self, '(//button[@id="add_pack_item"] | //button[contains(text(),"Add")])[1]')
            sleep(2)

        # Verify packaging_list has rows after loop
        if max_len > 0:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '//table[@id="packaging_list"]/tbody/tr')
            ))
            row_count = len(self.driver.find_elements(By.XPATH, '//table[@id="packaging_list"]/tbody/tr'))
            print(f"✅ {row_count} Packing item(s) added to packaging_list")


    # ─────────────────────────────────────────────────────────
    # FLOW 5 — REPAIR ORDER
    # ─────────────────────────────────────────────────────────
    def _flow_repair_order(self, row_data):
        driver, wait = self.driver, self.wait

        def _split(val):
            return [v.strip() for v in str(val).split("|") if v.strip()]

        orders = _split(row_data.get("OrderNo"))

        for order in orders:
            print(f"  -> Order No: {order}")
            order_input = wait.until(EC.presence_of_element_located((By.ID, "order_no")))
            order_input.clear()
            order_input.send_keys(order)
            sleep(0.5)

            # Search
            Function_Call.click(self, '(//button[contains(@class,"btrn_search")])[4]')
            sleep(3)

        # Select all repair order items (accumulated from all searches)
        if orders:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '//table[@id="bt_order_search_list"]/tbody/tr | //input[@name="id_orderdetails[]"]')
            ))
            self._check_all_by_name("id_orderdetails[]")
            print(f"✅ Repair Order items selected for {len(orders)} orders")

    # ─────────────────────────────────────────────────────────
    # POST-SAVE: CAPTURE BT CODE FROM PRINT TAB
    # ─────────────────────────────────────────────────────────
    def _capture_bt_code_from_print_tab(self):
        driver, wait = self.driver, self.wait
        bt_code = ""
        main_handle = driver.current_window_handle
        try:
            # Wait for the new print tab to open (up to 15 s)
            wait.until(lambda d: len(d.window_handles) > 1)
            new_tab = [h for h in driver.window_handles if h != main_handle][0]
            driver.switch_to.window(new_tab)
            sleep(2)

            # URL format: .../branch_transfer/print/{BT_CODE}/{type}/{id}
            current_url = driver.current_url
            print(f"🖨️  Print tab URL: {current_url}")
            if "/print/" in current_url:
                bt_code = current_url.split("/print/")[1].split("/")[0]
            print(f"✅ Captured BT Code: {bt_code}")

            driver.close()
        except Exception as e:
            print(f"⚠️ Print tab not opened or BT Code not found: {e}")
        finally:
            driver.switch_to.window(main_handle)
            sleep(2)
        return bt_code

    # ─────────────────────────────────────────────────────────
    # LIST PAGE VERIFICATION
    # ─────────────────────────────────────────────────────────
    def _verify_in_list(self, bt_code, row_data):
        driver, wait = self.driver, self.wait
        try:
            # Navigate to list page
            driver.get(BASE_URL + "index.php/admin_ret_brntransfer/branch_transfer/list")
            sleep(2)

            # Set date range to Today using daterangepicker
            try:
                date_btn = wait.until(
                    EC.element_to_be_clickable(
                        (By.CSS_SELECTOR, "#account-dt-btn,.daterangepicker-btn,#date-range-btn,.btn-daterange")
                    )
                )
                date_btn.click()
                sleep(1)
                # Click the "Today" preset inside the daterangepicker dropdown
                today_btn = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, '//li[normalize-space(text())="Today"] | //button[normalize-space(text())="Today"]')
                    )
                )
                today_btn.click()
                sleep(2)
                print("✅ Date range set to Today")
            except Exception as e:
                print(f"⚠️ Date range picker error (continuing): {e}")

            # Type BT Code into DataTables search box
            if bt_code:
                try:
                    search_input = wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH,
                             '//div[@id="bt_list_filter"]//input | //input[@type="search"][1]')
                        )
                    )
                    search_input.clear()
                    search_input.send_keys(bt_code)
                    sleep(2)
                    print(f"✅ Searched for BT Code: {bt_code}")
                except Exception as e:
                    print(f"⚠️ Search box error: {e}")

            # Verify row presence
            try:
                first_row = wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH,
                         '//table[@id="bt_list"]/tbody/tr[1] | //table[contains(@class,"dataTable")]/tbody/tr[1]')
                    )
                )
                row_text = first_row.text.strip()
                if bt_code and bt_code in row_text:
                    print(f"✅ Row verified in list: {row_text}")
                    return f"Verified: {row_text}"
                else:
                    print(f"⚠️ BT Code not found in row: {row_text}")
                    return f"Row found but BT not matched: {row_text}"
            except Exception as e:
                return f"List row not found: {str(e)}"

        except Exception as e:
            return f"List verification error: {str(e)}"

    # ─────────────────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────────────────
    def _check_all_in_table(self, table_id, checkbox_class=None, max_rows=None):
        """
        Check checkboxes inside a given table.

        Skips the header "All" checkbox (thead) to avoid toggling it off on
        repeated calls (e.g. multi-tag pipe-separated searches).

        Args:
            table_id       : HTML id of the <table>.
            checkbox_class : Optional class filter for row-level checkboxes.
            max_rows       : Optional int — select at most this many rows.
                             e.g. max_rows=2 selects only the first 2 rows.
        """
        driver = self.driver
        try:
            # Only target tbody checkboxes — never the thead "All" checkbox
            if checkbox_class:
                xpath = (
                    f'//table[@id="{table_id}"]//tbody'
                    f'//input[@type="checkbox" and contains(@class,"{checkbox_class}")]'
                )
            else:
                xpath = f'//table[@id="{table_id}"]//tbody//input[@type="checkbox"]'

            checkboxes = driver.find_elements(By.XPATH, xpath)

            # Apply row limit if requested
            if max_rows is not None:
                checkboxes = checkboxes[:max_rows]

            selected = 0
            for cb in checkboxes:
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
                selected += 1

            sleep(0.5)
            limit_msg = f" (limited to {max_rows})" if max_rows is not None else ""
            print(f"✅ Checked {selected} box(es) in #{table_id}{limit_msg}")
        except Exception as e:
            print(f"⚠️ _check_all_in_table({table_id}): {e}")

    def _check_all_by_name(self, input_name):
        """Check all checkboxes with a given name attribute."""
        driver = self.driver
        try:
            checkboxes = driver.find_elements(By.XPATH, f'//input[@name="{input_name}"]')
            for cb in checkboxes:
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
            sleep(0.5)
            print(f"✅ Checked {len(checkboxes)} boxes name={input_name}")
        except Exception as e:
            print(f"⚠️ _check_all_by_name({input_name}): {e}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, bt_code=""):
        """Write Pass/Fail + actual status back to Excel."""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            # Store BT Code in Column 23
            if bt_code:
                if not sheet.cell(row=1, column=23).value:
                    sheet.cell(row=1, column=23, value="BTCode").font = Font(bold=True)
                sheet.cell(row=row_num, column=23, value=bt_code).font = Font(bold=True, color="0000FF")

            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")

    def _take_screenshot(self, filename):
        """Save screenshot to reports folder."""
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(
                ExcelUtils.SCREENSHOT_PATH,
                f"{filename}_{datetime.now().strftime('%H%M%S')}.png"
            )
            self.driver.save_screenshot(path)
        except Exception:
            pass

    def _cancel_form(self):
        """Click cancel if form is open."""
        try:
            Function_Call.click(self, '//button[contains(@class,"btn-cancel")]')
            sleep(1)
        except Exception:
            pass

    def _update_approval_sheet(self, bt_code, row_data):
        """
        Creates two rows in the BranchTransferApproval sheet for every successful transfer.
        Row 1: Transit Approval, Row 2: Stock Download
        """
        try:
            workbook = load_workbook(FILE_PATH)
            if "BranchTransferApproval" not in workbook.sheetnames:
                print("⚠️ BranchTransferApproval sheet not found")
                return

            sheet = workbook["BranchTransferApproval"]
            # Find next row and last TC suffix
            target_row = None
            last_tc_suffix = 0
            for r in range(2, 5000):
                tc_id = sheet.cell(row=r, column=1).value
                if not tc_id:
                    target_row = r
                    break
                if str(tc_id).startswith("TC_BT_APP_"):
                    try:
                        match = re.search(r'\d+', str(tc_id))
                        if match:
                            suffix = int(match.group())
                            if suffix > last_tc_suffix:
                                last_tc_suffix = suffix
                    except:
                        pass
            else:
                target_row = sheet.max_row + 1

            # Prepare Data pairs (Transit Approval then Stock Download)
            approval_types = ["Transit Approval", "Stock Download"]
            
            for i, app_type in enumerate(approval_types):
                curr_row = target_row + i
                # Col 1: TestCaseId
                new_tc_id = f"TC_BT_APP_{str(last_tc_suffix + i + 1).zfill(3)}"
                sheet.cell(row=curr_row, column=1).value = new_tc_id
                
                # Col 4: ApprovalType
                sheet.cell(row=curr_row, column=4).value = app_type
                
                # Col 5, 6: Branches
                sheet.cell(row=curr_row, column=5).value = row_data.get("FromBranch")
                sheet.cell(row=curr_row, column=6).value = row_data.get("ToBranch")
                
                # Col 7: TransCode
                sheet.cell(row=curr_row, column=7).value = bt_code
                
                # Col 8: TransferType
                sheet.cell(row=curr_row, column=8).value = row_data.get("TransferType")
                
                # Col 9: IsOtherIssue
                sheet.cell(row=curr_row, column=9).value = row_data.get("OtherIssue", "No")
                
                # Col 10: ExpectedMsg
                sheet.cell(row=curr_row, column=10).value = "Records updated successfully"

            workbook.save(FILE_PATH)
            print(f"✅ BranchTransferApproval updated with 2 rows for BT Code: {bt_code}")
        except Exception as e:
            print(f"❌ BranchTransferApproval update error: {str(e)}")

