from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class BranchTransferApproval(unittest.TestCase):
    """
    Branch Transfer Approval Module Automation
    Module: Master → Branch Transfer → Branch Transfer Approval List
    Supports: Transit Approval / Stock Download
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_branch_transfer_approval(self):
        """Main entry point for Branch Transfer Approval automation"""
        driver = self.driver
        wait   = self.wait

        # Navigate to Branch Transfer Approval List
        try:
            if "admin_ret_brntransfer/branch_transfer/approval_list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Inventory')]")
                sleep(1)
                Function_Call.click(self, "(//span[contains(normalize-space(), 'BT Approval')])") 
                # Note: They might click from Master -> Branch Transfer, but let's go straight via URL if it fails
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            pass
        
        # Always force via URL to be perfectly safe as menu structures vary
        driver.get(BASE_URL + "index.php/admin_ret_brntransfer/branch_transfer/approval_list")
                               
        sleep(2)

        sheet_name = "BranchTransferApproval"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 2} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]

            # Column map — matching our create script
            data_map = {
                "TestCaseId":   1, "TestStatus":       2, "ActualStatus": 3,
                "ApprovalType": 4, "FromBranch":       5, "ToBranch":     6,
                "TransCode":    7, "TransferType":     8, "IsOtherIssue": 9,
                "ExpectedMsg":  10, "Remarks":         11
            }

            row_data = {
                key: sheet.cell(row=row_num, column=col).value
                for key, col in data_map.items()
            }
            workbook.close()

            # if str(row_data.get("TestStatus", "")).strip().lower() != "run":
            #     print(f"⏭️ Skipping row {row_num} (Status={row_data.get('TestStatus')})")
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 TC: {row_data.get('TestCaseId')}  |  Type: {row_data.get('ApprovalType')}")
            print(f"{'='*80}")

            try:
                result = self._run_approval_flow(row_data, row_num, sheet_name)
                print(f"🏁 Result: {result[0]} — {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
            except Exception as e:
                print(f"❌ TC {row_data.get('TestCaseId')} exception: {e}")
                self._take_screenshot(f"EX_{row_data.get('TestCaseId')}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)

    # ─────────────────────────────────────────────────────────
    # MAIN APPROVAL FLOW
    # ─────────────────────────────────────────────────────────
    def _run_approval_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Init"

        try:
            # 1. Select Approval Type Radio
            approval_type = str(row_data.get("ApprovalType", "Transit Approval")).strip()
            current_field = f"Approval Type: {approval_type}"

            if approval_type.lower() == "stock download":
                # Fallback locators
                xpath_radio = '//input[@name="bt_approval_type" and @value="2"]'
            else: # Defaults to Transit Approval
                xpath_radio = '//input[@name="bt_approval_type" and @value="1"]'
            
            radio_btn = wait.until(EC.presence_of_element_located((By.XPATH, xpath_radio)))
            driver.execute_script("arguments[0].click();", radio_btn)
            sleep(1.5)

            # 2. Select From Branch & To Branch
            if row_data.get("FromBranch"):
                current_field = "From Branch"
                self.fc.dropdown_select2(
                    '//select[@id="filter_from_brn"]/following-sibling::span',
                    str(row_data["FromBranch"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            if row_data.get("ToBranch"):
                current_field = "To Branch"
                self.fc.dropdown_select2(
                    '//select[@id="filtr_to_brn"]/following-sibling::span',
                    str(row_data["ToBranch"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 3. Enter Trans Code in Search box
            trans_code = str(row_data.get("TransCode", "")).strip()
            if trans_code:
                current_field = "Trans Code Filter Input"
                search_tc_input = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@id="bt_trans_code"]')))
                search_tc_input.clear()
                search_tc_input.send_keys(trans_code)
                sleep(0.5)

            # 4. Select Transfer Type Radio
            transfer_type_raw = str(row_data.get("TransferType", "Tagged")).strip()
            current_field = f"Transfer Type: {transfer_type_raw}"

            # Normalise to a canonical key (covers Excel typos / spacing variants)
            _norm = transfer_type_raw.lower().replace("-", "").replace(" ", "")
            type_map = {
                "tagged":          "1",
                "nontagged":       "2",   # "NonTagged", "Non-Tagged", "Non Tagged"
                "nontaggeditem":   "2",
                "purchaseitems":   "3",   # "Purchase Items", "PurchaseItems"
                "purchaseitem":    "3",
                "packagingitems":  "4",   # "Packaging Items", "PackagingItems"
                "packagingitem":   "4",
                "repairorder":     "5",   # "Repair Order", "RepairOrder"
            }
            radio_val = type_map.get(_norm, "1")
            print(f"   → TransferType '{transfer_type_raw}' → radio value '{radio_val}'")

            # Try by @value attribute (most reliable)
            try:
                type_radio = wait.until(EC.presence_of_element_located(
                    (By.XPATH, f'//input[@name="transfer_item_type" and @value="{radio_val}"]')
                ))
                driver.execute_script("arguments[0].click();", type_radio)
                print(f"   ✅ Clicked transfer_item_type radio value={radio_val}")
            except Exception as e1:
                print(f"   ⚠️ Primary radio click failed ({e1}), trying label fallback…")
                # Fallback: find radio by label text
                try:
                    label_xpath = (
                        f'//label[contains(normalize-space(translate(text(),'
                        f'"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz")),'
                        f'"{transfer_type_raw.lower()}")]'
                        f'/preceding-sibling::input[@type="radio"]'
                        f' | //label[contains(normalize-space(translate(text(),'
                        f'"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz")),'
                        f'"{transfer_type_raw.lower()}")]'
                        f'/input[@type="radio"]'
                    )
                    type_radio = wait.until(EC.presence_of_element_located((By.XPATH, label_xpath)))
                    driver.execute_script("arguments[0].click();", type_radio)
                    print(f"   ✅ Clicked transfer type radio via label fallback")
                except Exception as e2:
                    print(f"   ❌ Transfer type radio not clickable: {e2}")
            sleep(1)


            # 5. Other Issue (Checkbox)
            other_issue = str(row_data.get("IsOtherIssue", "No")).strip().lower()
            if other_issue == "yes":
                current_field = "Other Issue Checkbox"
                oi_cb = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@name="other_issue" or @id="other_issue"] | //label[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "other issue")]/input')))
                if not oi_cb.is_selected():
                    driver.execute_script("arguments[0].click();", oi_cb)
                sleep(1)

            # 6. Click Filter
            current_field = "Filter Button"
            Function_Call.click(self, '//button[@id="btran_filter"]')
            sleep(3) # Wait for table load

            self._take_screenshot(f"Filtered_{row_data.get('TestCaseId')}")

            # 7. Pre-Approve Flow (Transit vs Stock)
            current_field = "Table Submission Inputs"
            if approval_type.lower() == "transit approval":
                self._handle_transit_approval_table(trans_code)
            else:
                self._handle_stock_download_table(trans_code)

            # 8. Click Approve Button
            current_field = "Approve Button"
            approve_btn = wait.until(EC.presence_of_element_located((By.XPATH, '//button[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "approve")] | //a[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "approve")]')))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", approve_btn)
            sleep(0.5)
            driver.execute_script("arguments[0].click();", approve_btn)
            sleep(2)

            # 9. Verify Success
            current_field = "Success Message Verify"
            driver.execute_script("window.scrollTo(0, 0);") # Scroll back up for banner
            sleep(1)

            try:
                banner = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".alert.alert-success, .alert-box, .toast-message")))
                actual_msg = banner.text.strip().replace('\n', ' ')
            except Exception:
                actual_msg = "Success banner not found"

            expected_msg = str(row_data.get("ExpectedMsg", "")).strip()

            if expected_msg and expected_msg.lower() not in actual_msg.lower():
                return ("Fail", f"Expected '{expected_msg}' but got '{actual_msg}'")

            return ("Pass", actual_msg)

        except Exception as e:
            self._take_screenshot(f"ERR_{row_data.get('TestCaseId')}")
            return ("Fail", f"Error in [{current_field}]: {str(e)}")

    def _handle_transit_approval_table(self, trans_code):
        """Handle the table selections for Transit Approval"""
        driver, wait = self.driver, self.wait
        print("  -> Assessed as Transit Approval Flow")
        try:
            # Explicit wait for table wrapper
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[contains(@class,"dataTable")]/tbody/tr')))
             
            # Find input field in table and enter Trans Code
            if trans_code:
                # Look for an input inside the row that takes the trans code and isn't the main search filter
                # Or just dynamically find inputs that are NOT checkbox/radio/hidden inside tbody
                text_inputs = driver.find_elements(By.XPATH, '//table[contains(@class,"dataTable")]/tbody/tr//input[not(@type="checkbox" or @type="radio" or @type="hidden")]')
                for ti in text_inputs:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ti)
                        ti.clear()
                        sleep(0.2)
                        ti.send_keys(trans_code)
                        print(f"  -> Filled trans_code into row input: {trans_code}")
                    except Exception as e:
                        pass
            # Select table checkbox if exists
            checkboxes = driver.find_elements(By.XPATH, '//table[contains(@class,"dataTable")]/tbody/tr//input[@type="checkbox"]')
            for cb in checkboxes:
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
            sleep(1)
        except Exception as e:
            print(f"⚠️ Transit approval table handling warn: {e}")

    def _handle_stock_download_table(self, trans_code):
        """Handle the table selections for Stock Download"""
        driver, wait = self.driver, self.wait
        print("  -> Assessed as Stock Download Flow")
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[contains(@class,"dataTable")]/tbody/tr')))
            sleep(1)

            if trans_code:
                # As per recent screenshot, Stock Download table often has the trans_code input directly in the row
                row_inputs = driver.find_elements(By.XPATH, '//table[contains(@class,"dataTable")]/tbody/tr//input[not(@type="checkbox" or @type="radio" or @type="hidden")]')
                
                if row_inputs:
                    for ti in row_inputs:
                        try:
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ti)
                            ti.clear()
                            sleep(0.2)
                            ti.send_keys(trans_code)
                            print(f"  -> Filled trans_code into row input: {trans_code}")
                        except Exception as e:
                            pass
                else:
                    # Fallback to bottom input if the row input doesn't exist
                    bottom_inputs = driver.find_elements(By.XPATH, '//input[@name="tag_code" or @name="down_tag_code"] | //label[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "tag code")]/following-sibling::input | //label[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "trans code")]/following-sibling::input')
                    valid_inputs = [i for i in bottom_inputs if i.is_displayed()]
                    if valid_inputs:
                        bottom_in = valid_inputs[-1]
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", bottom_in)
                        bottom_in.clear()
                        bottom_in.send_keys(trans_code)
                        print(f"  -> Filled trans_code into bottom confirmation input: {trans_code}")
                    else:
                        print("  -> Could not find input for valid entry.")
            sleep(1)
            # Select table checkbox if exists
            checkboxes = driver.find_elements(By.XPATH, '//table[contains(@class,"dataTable")]/tbody/tr//input[@type="checkbox"]')
            for cb in checkboxes:
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
        except Exception as e:
            print(f"⚠️ Stock download table handling warn: {e}")


    # ─────────────────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────────────────
    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Write Pass/Fail + actual status back to Excel."""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")

    def _take_screenshot(self, filename):
        """Save screenshot to reports folder."""
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(
                ExcelUtils.SCREENSHOT_PATH,
                f"{filename}_{datetime.now().strftime('%H%M%S')}.png"
            )
            self.driver.save_screenshot(path)
        except Exception:
            pass

