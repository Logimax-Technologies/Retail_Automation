import os
import sys
import datetime
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
# Import established framework utilities
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from openpyxl.styles import Font
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class OrderLink(unittest.TestCase):
    """
    Automates the Order Link flow under Tagging > Order Link
    """
    def __init__(self, driver):
        super().__init__('test_order_link')
        self.driver = driver
        self.wait = WebDriverWait(driver, 15)
        self.fc = Function_Call(driver)


    def test_order_link(self):
        driver, wait = self.driver, self.wait

        # Navigate to Order Link
        print("🧭 Navigating to Order Link...")
        try:
            if "admin_ret_tagging/order_link" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                self.fc.click(self, "//span[contains(text(), 'Inventory')]")
                sleep(1)
                self.fc.click(self, "(//span[contains(normalize-space(), 'Order Link')])") 
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            pass
        
        # Always force via URL to be perfectly safe as menu structures vary
        driver.get(BASE_URL + "index.php/admin_ret_tagging/tagging/tag_link")
        sleep(2)

        sheet_name = "OrderLink"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 2} test cases in '{sheet_name}' sheet\n")
        except Exception as e:
            print(f"❌ Failed to read valid rows from '{sheet_name}': {e}")
            return []

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]

            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus":3,
                "Branch":     4, "FinYear":    5, "OrderNo":    6, 
                "TagNo":      7, "OldTagNo":   8, "ExpectedMsg": 9, 
                "Remarks":    10
            }

            row_data = {
                key: sheet.cell(row=row_num, column=col).value
                for key, col in data_map.items()
            }
            workbook.close()

            print(f"\n{'='*80}")
            print(f"🧪 TC: {row_data.get('TestCaseId')}  |  Order: {row_data.get('OrderNo')}")
            print(f"{'='*80}")

            try:
                result = self._run_order_link(row_data)
                print(f"🏁 Result: {result[0]} — {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
            except Exception as e:
                print(f"❌ TC {row_data.get('TestCaseId')} exception: {e}")
                self._take_screenshot(f"EX_{row_data.get('TestCaseId')}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)

    def _run_order_link(self, row_data):
        """Core automation for a single row's Order Link test case."""
        driver, wait = self.driver, self.wait
        current_field = "Initialization"
        
        try:
            # Refresh if we're doing a second test case to get clean state
            # driver.refresh()
            # sleep(2)

            # 1. Select Branch
            branch = str(row_data.get("Branch", "")).strip()
            if branch:
                current_field = "Branch"
                # Using standard select2 locating mechanism
                self.fc.dropdown_select2(
                    '//select[@id="branch_select"]/following-sibling::span',
                    branch,
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 2. Select Financial Year
            fin_year = str(row_data.get("FinYear", "")).strip()
            if fin_year:
                current_field = "FinYear"
                fin_select_el = wait.until(EC.presence_of_element_located((By.XPATH, '//select[@id="order_fin_year_select"]')))
                Select(fin_select_el).select_by_visible_text(fin_year)
                sleep(1)

            # 3. Search and Select Order
            order_no = str(row_data.get("OrderNo", "")).strip()
            if not order_no:
                raise ValueError("OrderNo cannot be empty.")
            
            current_field = "Search Order"
            self.fc.dropdown_select2(
                '//select[@id="select_order"]/following-sibling::span',
                order_no,
                '//span[@class="select2-search select2-search--dropdown"]/input'
            )
            print(f"  -> Searched and selected Order: {order_no}")
            
            # Wait for table to dynamically load AJAX
            sleep(2)
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[contains(@class,"dataTable") or contains(@class,"table")]/tbody/tr')))

            # 4. Table Interactions
            current_field = "Table Row Processing"
            self._take_screenshot(f"TableLoaded_{row_data.get('TestCaseId')}")

            # Make sure main checkbox is selected (Targeting first visual row)
            checkboxes = driver.find_elements(By.XPATH, '//table/tbody/tr//input[@type="checkbox"]')
            if checkboxes:
                cb = checkboxes[0]
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
            
            tag_val = row_data.get("TagNo")
            tag_no = str(tag_val).strip() if tag_val is not None else ""
            
            old_tag_val = row_data.get("OldTagNo")
            old_tag_no = str(old_tag_val).strip() if old_tag_val is not None else ""

            if tag_no or old_tag_no:
                # Find all valid text inputs in the row
                row_inputs = driver.find_elements(By.XPATH, '//table/tbody/tr[1]//input[@type="text" or not(@type)]')
                
                # We assume Tag No is the first text input and Old Tag No is the second or specific by name
                if tag_no and len(row_inputs) >= 1:
                    xpath_tag_input = '(//table/tbody/tr[1]//input[@type="text" or not(@type)])[1]'
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row_inputs[0])
                    self.fc.fill_autocomplete_field2(xpath_tag_input, tag_no)
                    print(f"  -> Autocompleted TagNo: {tag_no}")
                    
                    
                if old_tag_no and len(row_inputs) >= 2:
                    xpath_old_tag_input = '(//table/tbody/tr[1]//input[@type="text" or not(@type)])[2]'
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row_inputs[1])
                    self.fc.fill_autocomplete_field2(xpath_old_tag_input, old_tag_no)
                    print(f"  -> Autocompleted OldTagNo: {old_tag_no}")
            
            # 5. Save All
            current_field = "Save All Button"
            sleep(1)
            save_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "save all")] | //button[@id="save_link"]')))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", save_btn)
            sleep(0.5)
            driver.execute_script("arguments[0].click();", save_btn)

            # 6. Verify Success
            current_field = "Success Message Banner"
            expected_msg = str(row_data.get("ExpectedMsg", "")).strip()
            
            try:
                banner = wait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "alert-success")] | //div[@class="toast-message"]')))
                actual_msg = banner.text.strip()
                driver.execute_script("arguments[0].scrollIntoView(true);", banner)
                self._take_screenshot(f"Success_{row_data.get('TestCaseId')}")
                print(f"  -> Validated Banner: {actual_msg}")

                if expected_msg and expected_msg.lower() not in actual_msg.lower():
                    return ("Warning", f"Expected '{expected_msg}', got '{actual_msg}'")
                return ("Pass", actual_msg)
            except TimeoutException:
                self._take_screenshot(f"NoBanner_{row_data.get('TestCaseId')}")
                return ("Fail", "Success banner did not appear after save.")

        except Exception as e:
            print(f"❌ Error at '{current_field}': {e}")
            self._take_screenshot(f"Error_{row_data.get('TestCaseId', 'Unknown')}")
            return ("Fail", f"Error in [{current_field}]: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Write Pass/Fail + actual status back to Excel."""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")

    def _take_screenshot(self, name):
        """Helper to capture full screenshots during critical execution states."""
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"OrderLink_{name}_{timestamp}.png"
            os.makedirs("Screenshots/OrderLink", exist_ok=True)
            path = os.path.join("Screenshots/OrderLink", filename)
            self.driver.save_screenshot(path)
            print(f"📸 Screen saved: {filename}")
        except Exception as e:
            print(f"⚠️ Screenshot Failed: {e}")
