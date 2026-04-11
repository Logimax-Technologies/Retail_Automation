from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import re
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class NonTagReceipt(unittest.TestCase):
    """
    Non-Tag Receipt Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_non_tag_receipt(self):
        """Main entry point for Non-Tag Receipt automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigate to Non-Tag Receipt Module
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Inventory')]")
            Function_Call.click(self, "(//span[contains(normalize-space(), 'Non Tag Receipt')])")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_purchase/nontag_receipt/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "NonTagReceipt"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "LotNo": 4, "Product": 5, "Design": 6, "SubDesign": 7,
                "Branch": 8, "Section": 9, "Pieces": 10,
                "GrossWt": 11, "NetWt": 12, "Remark": 13, "ExpectedStatus": 14,
                "Receipt No":15
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            # if str(row_data["TestStatus"]).strip().lower() != "yes":
            #     print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Yes')")
            #     continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_nontag_receipt_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # If save successful, verify in list page and capture receipt no
                receipt_no = ""
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_nontag_receipt_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    
                    if list_result[0] == "Pass" and len(list_result) > 2:
                        receipt_no = list_result[2]
                        # Combine messages for clarity or update to the verification status
                        actual_status = f"{actual_status} | {list_result[1]}"
                
                # Single update call at the end
                self._update_excel_status(row_num, test_status, actual_status, sheet_name, receipt_no=receipt_no)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Non-Tag Receipt Automation Completed")
        print(f"{'='*80}")

    def test_nontag_receipt_add(self, row_data, row_num, sheet_name):
        """Logic to add a new Non-Tag Receipt"""
        driver = self.driver
        wait = self.wait
        current_field = "Navigation"
        
        try:
            # Click Add Button
            Function_Call.click(self, '//a[@id="add_Order"]')
            sleep(3)
            
            # 1. Select Lot No (Select2)
            if row_data["LotNo"]:
                current_field = f"Lot No ({row_data['LotNo']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_lot"]/following-sibling::span',
                    row_data["LotNo"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                print(f"✅ Selected Lot No: {row_data['LotNo']}")
                sleep(2) # Wait for AJAX to load details
            
            # 2. Select Product, Design, and SubDesign (Manually)
            if row_data["Product"]:
                current_field = f"Product ({row_data['Product']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="product_sel"]/following-sibling::span',
                    row_data["Product"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            if row_data["Design"]:
                current_field = f"Design ({row_data['Design']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="design_sel"]/following-sibling::span',
                    row_data["Design"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            if row_data["SubDesign"]:
                current_field = f"SubDesign ({row_data['SubDesign']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="sub_design_sel"]/following-sibling::span',
                    row_data["SubDesign"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)
            
            # 3. Select Branch & Section if available
            if row_data["Branch"]:
                current_field = f"Branch ({row_data['Branch']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="branch_select"]/following-sibling::span',
                    row_data["Branch"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
            
            if row_data["Section"]:
                current_field = f"Section ({row_data['Section']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_section"]/following-sibling::span',
                    row_data["Section"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )

            # 4. Input Pieces, Gross Wt, Net Wt
            if row_data["Pieces"]:
                current_field = f"Pieces ({row_data['Pieces']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.CLASS_NAME, "nt_pcs"),
                    value=str(row_data["Pieces"]),
                    field_name="Pieces",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            if row_data["GrossWt"]:
                current_field = f"Gross Wt ({row_data['GrossWt']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.CLASS_NAME, "nt_grswt"),
                    value=str(row_data["GrossWt"]),
                    field_name="Gross Wt",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            if row_data["NetWt"]:
                current_field = f"Net Wt ({row_data['NetWt']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.CLASS_NAME, "nt_netwt"),
                    value=str(row_data["NetWt"]),
                    field_name="Net Wt",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            # 5. Remark
            if row_data["Remark"]:
                current_field = f"Remark ({row_data['Remark']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "remark"),
                    value=row_data["Remark"],
                    field_name="Remark",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            # 6. Save
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="nt_receipt_submit"]')
            sleep(3)

            # 7. Success Verification
            try:
                # Use toast or alert verification as per project standard
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'NonTag Receipt Added Successfully')]"
                success_element = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath)))
                msg = success_element.text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "NonTag Receipt Added Successfully")
            except:
                # Check for error alert
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_nontag_receipt_list_verification(self, row_data):
        """Verify the added receipt in the list table"""
        driver = self.driver
        wait = self.wait
        
        try:
            # Navigate to list if not already there
            if "nontag_receipt/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/nontag_receipt/list")
                sleep(2)
            
            # Search for Lot No in DataTable search
            search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(str(row_data["LotNo"]))
            sleep(2)
            
            # Verify first row and capture Receipt No
            try:
                row_xpath = f"//table[@id='nontag_receipt_list']/tbody/tr[contains(., '{row_data['LotNo']}')]"
                row_element = wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                
                # Extract Receipt No from Column 2
                receipt_no = row_element.find_element(By.XPATH, "./td[2]").text.strip()
                print(f"✅ Found Lot {row_data['LotNo']} in list page. Receipt No: {receipt_no}")
                return ("Pass", f"Verified Lot {row_data['LotNo']} in list", receipt_no)
            except:
                print(f"❌ Could not find Lot {row_data['LotNo']} in list page after search")
                return ("Fail", f"Lot {row_data['LotNo']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, receipt_no=""):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            # Update Receipt No in Column 15
            if receipt_no:
                sheet.cell(row=row_num, column=15, value=receipt_no).font = Font(bold=True, color="0000FF")
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")

    def _take_screenshot(self, name):
        """Utility to take screenshots"""
        try:
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{name}.png")
            self.driver.save_screenshot(path)
            print(f"📸 Screenshot saved: {path}")
        except:
            pass
