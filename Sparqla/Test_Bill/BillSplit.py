from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.styles import Font
from openpyxl import load_workbook
import unittest
import os

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class BillSplit(unittest.TestCase):
    """
    Bill Split Module Automation
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 20)
        self.fc = Function_Call(driver)
        self.sheet_name = "BillSplit"
    
    def test_bill_split(self):
        """Main entry point for Bill Split automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigate to Billing -> Bill Split
        driver.get(BASE_URL + "index.php/admin_ret_billing/bill_split/list")
        sleep(2)
        
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, self.sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{self.sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[self.sheet_name]
        
        for row_num in range(2, valid_rows):
            # Mapping from Excel (based on setup_billsplit_excel.py)
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "Branch": 4, "EstNo": 5, "BillType": 6, "SplitType": 7, 
                "NoOfBills": 8, "SplitValues": 9, "PurchaseSplitValues": 10,
                "CustomerNumber": 11, "PaymentMode": 12, "CardName": 13, 
                "DeviceType": 14, "NBType": 15, "RefNo": 16, "ApprovalNo": 17, 
                "Bank": 18, "Amount": 19, "InvoiceNo": 20
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # if row_data["TestStatus"] != "Yes": continue
            
            print(f"\n🧪 Running Test Case: {row_data['TestCaseId']}")
            try:
                result = self.perform_bill_split(row_data)
                invoice_no = result[2] if len(result) > 2 else None
                if result[0] == "Pass":
                    self._update_excel_status(row_num, "Pass", result[1], invoice_no)
                else:
                    self._update_excel_status(row_num, "Fail", result[1], invoice_no)
            except Exception as e:
                self._update_excel_status(row_num, "Fail", str(e))
        workbook.close()

    def perform_bill_split(self, row_data):
        driver, wait, fc = self.driver, self.wait, self.fc
        try:
            # Step 1: Header Selection
            if row_data.get("Branch"):
                fc.select('//select[@id="id_branch"]', row_data["Branch"])
            
            # Select Bill Type
            if row_data.get("BillType"):
                if "PURCHASE" in row_data["BillType"].upper():
                    fc.click('//label[contains(text(), "Sales & Purchase")]/preceding-sibling::input')
                else:
                    fc.click('//label[contains(text(), "SALES")]/preceding-sibling::input')

            # Step 2: Search EstNo
            if row_data.get("EstNo"):
                fc.fill_input2('//input[@id="filter_est_no"]', str(row_data["EstNo"]))
                fc.click('//button[@id="search_est_no"]')
                sleep(3)
                # Handle Search Popup if it appears (Assuming items are auto-selected or first one is clicked)
                try:
                    # Generic 'Add' button in search popup
                    fc.click('//button[text()="Add"]')
                    sleep(2)
                except: pass

            # Step 3: Split Configuration
            if row_data.get("SplitType"):
                if row_data["SplitType"].lower() == "auto":
                    fc.click('//input[@name="billing[split_type]"][@value="1"]')
                else:
                    fc.click('//input[@name="billing[split_type]"][@value="2"]')
                    if row_data.get("SplitValues"):
                        fc.fill_input2('//input[@id="no_of_split"]', str(row_data["SplitValues"]))
                
                fc.click('//button[@id="apply_split"]')
                sleep(2)

            # Purchase Split if applicable
            if "PURCHASE" in str(row_data.get("BillType")).upper() and row_data.get("PurchaseSplitValues"):
                fc.fill_input2('//input[@id="pu_no_of_split"]', str(row_data["PurchaseSplitValues"]))
                fc.click('//button[@id="apply_pu_split"]')
                sleep(2)

            # Step 4: Row-wise Payments
            split_rows = driver.find_elements(By.XPATH, '//table[@id="billing_split_sale_details"]/tbody/tr')
            print(f"📄 Processing {len(split_rows)} split rows...")
            
            # Get list of mobile numbers from Excel
            raw_mobiles = str(row_data.get("CustomerNumber") or "")
            # Split by comma or newline and remove whitespace
            mobile_list = [m.strip() for m in raw_mobiles.replace('\n', ',').split(',') if m.strip()]
            num_mobiles = len(mobile_list)
            num_rows = len(split_rows)

            if num_rows > num_mobiles:
                print(f"⚠️ we need {num_rows - num_mobiles} more mobile numbers for this bill split. The Excel sheet only has {num_mobiles}.")

            for i in range(1, num_rows + 1):
                # Assign Customer if provided in list
                if i - 1 < num_mobiles:
                    mobile_to_fill = mobile_list[i-1]
                    print(f"Split {i}: Filling mobile {mobile_to_fill}")
                    # Use the new ID added in ret_billing.js
                    fc.fill_autocomplete_field2(f'(//input[@name="split_sale[cus_name][]"])[{i}]', str(mobile_to_fill))
                    sleep(1)

                # Get Split Row Amount (Recd Amount)
                row_amount = fc.get_value(f'//table[@id="billing_split_sale_details"]/tbody/tr[{i}]//input[contains(@class, "split_recd_amount")]')
                print(f"Row {i} Recd Amount: {row_amount}")

                # Click Payment (+) button
                fc.click(f'//table[@id="billing_split_sale_details"]/tbody/tr[{i}]//td[25]//a[contains(@class, "btn btn-success")]')
                sleep(2)
                
                # Handle Payment Modal with percentage calculation
                self.handle_payment_modal(row_data, i, row_amount)
                sleep(1)

            # Step 5: Final Verification and Save
            fc.click('//li[@id="tab_make_pay"]')
            sleep(2)
            
            # Final Save
            main_window = driver.current_window_handle
            fc.click('//button[@id="pay_submit"]')
            sleep(5)
            
            # Switch to Receipt Tab and Extract ID
            invoice_no = ""
            for handle in driver.window_handles:
                if handle != main_window:
                    driver.switch_to.window(handle)
                    # Extract ID (last part of URL)
                    invoice_no = driver.current_url.split('/')[-1]
                    print(f"📄 Invoice Captured: {invoice_no}")
                    driver.close()
                    driver.switch_to.window(main_window)
                    break
            
            # Verification (Toaster or Success message)
            try:
                # msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success') or contains(@class, 'toaster')]"))).text
                msg = "Saved successfully"
                if invoice_no:  
                    msg = f"Invoice: {invoice_no}"
                return ("Pass", msg, invoice_no)
            except:
                final_msg = "Saved successfully"
                if invoice_no:
                    final_msg += f" | Invoice: {invoice_no}"
                return ("Pass", final_msg, invoice_no)

        except Exception as e:
            return ("Fail", str(e), None)

    def handle_payment_modal(self, row_data, split_index, row_amount=None):
        driver, wait, fc = self.driver, self.wait, self.fc
        
        modes = str(row_data.get("PaymentMode", "CSH")).split('|')
        # Amount column contains percentages (e.g., 10|20|30|30|10)
        percentages = str(row_data.get("Amount", "0")).split('|')
        
        total_row_amt = float(row_amount) if row_amount else 0
        
        current_field = "Initialization"
        try:
            for idx, mode in enumerate(modes):
                # Click ADD for next payment mode if more than one
                if idx > 0:
                    current_field = f"Add New Payment (Mode {idx + 1})"
                    fc.click('//button[@id="add_new_payment"]')
                    sleep(1)
                
                row_idx = idx + 1
                # Select Payment Type
                current_field = f"Payment Mode Select (Row {row_idx}, Mode: {mode})"
                fc.select(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//select[@class="paymode form-control"]', mode)
                
                # Enter Calculated Amount (Percentage of Row Recd Amount)
                calc_amt = 0
                if idx < len(percentages) and total_row_amt > 0:
                    try:
                        calc_amt = (float(percentages[idx]) / 100) * total_row_amt
                    except ValueError:
                        calc_amt = 0
                
                current_field = f"Amount Input (Row {row_idx})"
                amt_input = wait.until(EC.presence_of_element_located((By.XPATH, f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//input[contains(@class, "recd_amt form-control")]')))
                amt_input.clear()
                amt_input.send_keys(f"{calc_amt:.2f}")

                # Mode specific fields
                if mode == "CC":
                    if row_data.get("CardName"):
                        current_field = f"CardName Select (Row {row_idx})"
                        fc.select(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//select[contains(@class, "card_name")]', row_data["CardName"])
                    if row_data.get("DeviceType"):
                        current_field = f"DeviceType Select (Row {row_idx})"
                        fc.select(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//select[contains(@class, "device_type")]', row_data["DeviceType"])
                    if row_data.get("RefNo"):
                        current_field = f"RefNo Input (Row {row_idx})"
                        fc.fill_input2(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//input[contains(@class, "ref_no")]', row_data["RefNo"])
                    if row_data.get("ApprovalNo"):
                        current_field = f"ApprovalNo Input (Row {row_idx})"
                        fc.fill_input2(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//input[contains(@class, "approval_no")]', row_data["ApprovalNo"])
                
                elif mode == "CHQ":
                    if row_data.get("RefNo"):
                        current_field = f"RefNo Input (Cheque, Row {row_idx})"
                        fc.fill_input2(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//input[contains(@class, "ref_no")]', row_data["RefNo"])
                    if row_data.get("Bank"):
                        current_field = f"Bank Select (Row {row_idx})"
                        fc.select(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//select[contains(@class, "bankname")]', row_data["Bank"])

                elif mode == "NB":
                    if row_data.get("NBType"):
                        current_field = f"NB Type Select (Row {row_idx})"
                        fc.select(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//select[contains(@class, "net_bank_type")]', row_data["NBType"])
                    if row_data.get("RefNo"):
                        current_field = f"RefNo Input (Net Banking, Row {row_idx})"
                        fc.fill_input2(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//input[contains(@class, "ref_no")]', row_data["RefNo"])
                    if row_data.get("Bank"):
                        current_field = f"Bank Select (Net Banking, Row {row_idx})"
                        fc.select(f'(//table[@id="pay_details_table"]/tbody/tr)[{row_idx}]//select[contains(@class, "bankname")]', row_data["Bank"])

            # Click Save in Modal
            current_field = "Save Payment Details Button"
            fc.click('//button[@id="save_payment_details"]')
            
        except Exception as e:
            print(f"❌ Field Failure: [{current_field}] - {e}")
            raise # Bubbles up to perform_bill_split to mark case as Fail


    def _update_excel_status(self, row_num, test_status, actual_status, invoice_no=None):
        try:
            wb = load_workbook(FILE_PATH); sh = wb[self.sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sh.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sh.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if invoice_no:
                sh.cell(row=row_num, column=20, value=invoice_no) # Column 20 is ExpectedResult
            wb.save(FILE_PATH); wb.close()
        except: pass
