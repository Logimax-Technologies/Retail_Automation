from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.styles import Font
from openpyxl import load_workbook
import re
import unittest
import os
from datetime import datetime

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class BillingReceipt(unittest.TestCase):
    """
    Billing Receipt Module Automation
    Handles multiple scenarios: Advance, Petty Cash, Credit Collection, etc.
    Now includes Advance Transfer.
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 20)
        self.fc = Function_Call(driver)
    
    def run_all_tests(self):
        """Main entry point for all Billing Receipt related tests"""
        self.test_billing_receipt()
        self.test_advance_transfer()

    def test_billing_receipt(self):
        """Main entry point for Billing Receipt automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigate to Billing -> Receipt
        driver.get(BASE_URL + "index.php/admin_ret_billing/receipt/list")
        sleep(2)
        
        # Read Excel data
        sheet_name = "BillingReceipt"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        data_map = {
            "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
            "Branch": 4, "Customer": 5, "ReceiptType": 6, "Amount": 7, "Employee": 8,
            "Narration": 9, "CashAmount": 10, 
            "Card(Yes/No)": 11, "CardType": 12, "CardBank": 13, "CardNo": 14, "ExpDate": 15, "CardAmount": 16,
            "Cheque(Yes/No)": 17, "ChqDate": 18, "ChqBank": 19, "ChqNo": 20, "ChqAmount": 21,
            "NB(Yes/No)": 22, "NBType": 23, "NBBank": 24, "NBDate": 25, "NBRefNo": 26, "NBAmount": 27,
            "AdvAdj(Yes/No)": 28, "AdvAdjRecNo": 29, "AdvAdjAmount": 30,
            "EmpCreditRecNo": 31, "EmpCreditAmt": 32,
            "BillNo": 33, "FromDate": 34, "ToDate": 35
        }

        for row_num in range(2, valid_rows):
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            if row_data["TestStatus"] != "Yes": continue
            
            print(f"\n🧪 Running Test Case: {row_data['TestCaseId']}")
            try:
                result = self.test_receipt_creation(row_data, row_num, sheet_name)
                if result[0] == "Pass":
                    captured_id = result[2] if len(result) > 2 else ""
                    bill_no = self.test_list_verification(captured_id, row_data.get("FromDate"), row_data.get("ToDate"))
                    self._update_excel_status(row_num, "Pass", "Success", sheet_name, bill_no)
                else:
                    self._update_excel_status(row_num, "Fail", result[1], sheet_name)
            except Exception as e:
                self._update_excel_status(row_num, "Fail", str(e), sheet_name)
        workbook.close()

    def test_receipt_creation(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        try:
            Function_Call.click(self, '//a[@id="add_billing"]')
            sleep(2)
            if row_data.get("Branch"):
                Function_Call.dropdown_select(self, '//select[@id="branch_select"]/following-sibling::span', row_data["Branch"], '//span[@class="select2-search select2-search--dropdown"]/input')
            
            if row_data.get("Customer"):
                Function_Call.fill_autocomplete_field(self, "name", str(row_data["Customer"]))
            
            rtype = str(row_data.get("ReceiptType"))
            if rtype:
                Function_Call.click(self, f"//input[@name='receipt[receipt_type]'][@value='{rtype}']")
            
            if row_data.get("TestCaseId") == "TC_BR_05": # Against Est
                Function_Call.click(self, '//input[@id="is_aganist_est_yes"]')
                Function_Call.fill_input(self, wait, (By.ID, "esti_no"), "1")
                Function_Call.click(self, '//button[@id="est_search"]')
                sleep(2)

            if rtype == "8": # Petty
                Function_Call.click(self, '//input[@id="receipt_to_emp"]')
                Function_Call.click(self, '//input[@id="is_aganist_est_no"]')

            if row_data.get("Amount"):
                Function_Call.fill_input(self, wait, (By.ID, "amount"), str(row_data["Amount"]))
            
            if row_data.get("Employee"):
                Function_Call.dropdown_select(self, '//select[@id="emp_select"]/following-sibling::span', row_data["Employee"], '//span[@class="select2-search select2-search--dropdown"]/input')

            if float(row_data.get("Amount", 0)) > 200000 and str(row_data.get("TestCaseId")) != "TC_BR_ERROR_02":
                Function_Call.fill_input(self, wait, (By.ID, "pan_no"), "ABCDE1234F")
                Function_Call.fill_input(self, wait, (By.ID, "aadhar_no"), "123456789012")

            if row_data.get("Narration"):
                Function_Call.fill_input(self, wait, (By.ID, "narration"), row_data["Narration"])

            if row_data.get("EmpCreditRecNo"): self.handle_employee_credit(row_data)
            if row_data.get("AdvAdj(Yes/No)") == "Yes": self.handle_advance_adjustment(row_data)

            if float(row_data.get("CashAmount") or 0) > 0:
                Function_Call.fill_input(self, wait, (By.ID, "make_pay_cash"), str(row_data["CashAmount"]))
            
            if row_data.get("Card(Yes/No)") == "Yes": self.handle_card(row_data)
            if row_data.get("Cheque(Yes/No)") == "Yes": self.handle_cheque(row_data)
            if row_data.get("NB(Yes/No)") == "Yes": self.handle_nb(row_data)

            original_window = driver.current_window_handle
            Function_Call.click(self, '//button[@id="save_receipt"]')
            sleep(3)
            
            # High-value validation test
            if float(row_data.get("Amount", 0)) > 200000 and str(row_data.get("TestCaseId")) == "TC_BR_ERROR_02":
                try:
                    alert = driver.switch_to.alert
                    alert.accept()
                    return ("Pass", "Validation worked")
                except: return ("Fail", "Validation alert missed")

            # ID extraction
            captured_id = ""
            try:
                wait.until(EC.number_of_windows_to_be(2))
                for window_handle in driver.window_handles:
                    if window_handle != original_window:
                        driver.switch_to.window(window_handle)
                        captured_id = self._extract_id_from_url()
                        driver.close(); driver.switch_to.window(original_window)
                        break
            except: captured_id = self._extract_id_from_url()

            msg_el = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]")))
            if "successfully" in msg_el.text.lower():
                return ("Pass", "Success", captured_id)
            return ("Fail", "Success message not found")
        except Exception as e: return ("Fail", str(e))

    def handle_card(self, row_data):
        Function_Call.click(self, '//a[@id="card_detail_modal"]')
        sleep(2)
        Function_Call.select_visible_text(self, '(//select[contains(@id, "card_type")])[1]', row_data.get("CardType", "Visa"))
        Function_Call.select_visible_text(self, '(//select[@name="card_details[id_bank][]"])[1]', row_data.get("CardBank", "IOB 4795"))
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="card_details[card_no][]"])[1]'), str(row_data.get("CardNo", "1234")))
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="card_details[card_amount][]"])[1]'), str(row_data.get("CardAmount", 0)))
        Function_Call.click(self, '//a[@id="add_card"]'); sleep(1)

    def handle_cheque(self, row_data):
        Function_Call.click(self, '//a[@id="cheque_modal"]')
        sleep(2)
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="cheque_details[cheque_date][]"])[1]'), str(row_data.get("ChqDate", "21-03-2026")))
        Function_Call.select_visible_text(self, '(//select[@name="cheque_details[id_bank][]"])[1]', row_data.get("ChqBank", "IOB 4795"))
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="cheque_details[cheque_no][]"])[1]'), str(row_data.get("ChqNo", "111")))
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="cheque_details[payment_amount][]"])[1]'), str(row_data.get("ChqAmount", 0)))
        Function_Call.click(self, '//a[@id="save_chq"]'); sleep(1)

    def handle_nb(self, row_data):
        Function_Call.click(self, '//a[@id="net_bank_modal"]')
        sleep(2)
        Function_Call.select_visible_text(self, '(//select[contains(@id, "nb_type")])[1]', row_data.get("NBType", "IMPS"))
        Function_Call.select_visible_text(self, '(//select[@name="nb_details[nb_bank][]"])[1]', row_data.get("NBBank", "IOB 4795"))
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="nb_details[ref_no][]"])[1]'), str(row_data.get("NBRefNo", "REF123")))
        Function_Call.fill_input(self, self.wait, (By.XPATH, '(//input[@name="nb_details[amount][]"])[1]'), str(row_data.get("NBAmount", 0)))
        Function_Call.click(self, '//a[@id="save_net_banking"]'); sleep(1)

    def handle_employee_credit(self, row_data):
        try:
            modal = self.driver.find_element(By.ID, "credit_collection")
            if modal.is_displayed():
                Function_Call.click(self, '//table[@id="issue_list"]/tbody/tr[1]/td[1]/input')
                Function_Call.fill_input(self, self.wait, (By.XPATH, '//table[@id="issue_list"]/tbody/tr[1]/td[7]/input'), str(row_data.get("EmpCreditAmt", 0)))
                Function_Call.click(self, '//button[@id="save_credit_collection"]'); sleep(1)
        except: pass

    def handle_advance_adjustment(self, row_data):
        Function_Call.click(self, '//button[@id="adv_adj_modal"]')
        sleep(2)
        Function_Call.click(self, '//table[@id="bill_adv_adj"]/tbody/tr[1]/td[1]/input')
        Function_Call.fill_input(self, self.wait, (By.XPATH, '//table[@id="bill_adv_adj"]/tbody/tr[1]/td[4]/input'), str(row_data.get("AdvAdjAmount", 0)))
        Function_Call.click(self, '//button[@id="save_receipt_adv_adj"]'); sleep(1)

    def test_list_verification(self, captured_id="", from_date=None, to_date=None):
        driver, wait = self.driver, self.wait
        try:
            if from_date and to_date:
                driver.find_element(By.ID, "dt_range").click()
                start = wait.until(EC.element_to_be_clickable((By.NAME, "daterangepicker_start")))
                start.clear(); start.send_keys(str(from_date))
                end = wait.until(EC.element_to_be_clickable((By.NAME, "daterangepicker_end")))
                end.clear(); end.send_keys(str(to_date))
                Function_Call.click(self, '//button[contains(@class, "applyBtn")]'); sleep(1)
                Function_Call.click(self, '//button[@id="receipt_search"]'); sleep(2)

            search_input = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@type="search"]')))
            search_input.clear(); search_input.send_keys(captured_id); search_input.send_keys(Keys.ENTER); sleep(2)
            bill_no = driver.find_element(By.XPATH, '//table[@id="receipt_list"]/tbody/tr[1]/td[3]').text
            return bill_no
        except: return "Unknown"

    def test_advance_transfer(self):
        """Automation for Advance Transfer module"""
        driver, wait = self.driver, self.wait
        driver.get(BASE_URL + "index.php/admin_ret_billing/advance_transfer/list")
        sleep(2)
        
        sheet_name = "AdvanceTransfer"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
        workbook = load_workbook(FILE_PATH); sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows):
            data_map = {"TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3, "Branch": 4, "FromCustomer": 5, "ToCustomer": 6, "TransferAmount": 7, "OTP": 8}
            row_data = {k: sheet.cell(row=row_num, column=v).value for k, v in data_map.items()}
            if row_data["TestStatus"] != "Yes": continue
            
            try:
                Function_Call.click(self, '//a[@id="add_advance_transfer"]')
                sleep(2)
                if row_data.get("Branch"):
                    Function_Call.dropdown_select(self, '//select[@id="branch_select"]/following-sibling::span', row_data["Branch"], '//span[@class="select2-search select2-search--dropdown"]/input')
                
                Function_Call.fill_autocomplete_field(self, "adv_trns_from_cust", str(row_data["FromCustomer"]))
                sleep(2) # Wait for table to load
                
                # Select first receipt in table advance_trns_list
                Function_Call.click(self, '//table[@id="advance_trns_list"]/tbody/tr[1]/td[1]/input')
                Function_Call.fill_input(self, wait, (By.XPATH, '//table[@id="advance_trns_list"]/tbody/tr[1]/td[4]/input'), str(row_data["TransferAmount"]))
                
                Function_Call.fill_autocomplete_field(self, "adv_trns_to_cust", str(row_data["ToCustomer"]))
                Function_Call.click(self, '//button[@id="submit_advance_transfer"]')
                sleep(2)
                
                # Handle OTP if modal appears
                try:
                    otp_modal = driver.find_element(By.ID, "otp_modal")
                    if otp_modal.is_displayed():
                        Function_Call.fill_input(self, wait, (By.ID, "adv_trns_otp"), "123456") # Example OTP
                        Function_Call.click(self, '//button[@id="verify_advance_transfer_otp"]')
                        sleep(1)
                        Function_Call.click(self, '//button[contains(@class, "submit_advance_transfer")]')
                except: pass

                # Verify
                msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text
                self._update_excel_status(row_num, "Pass", "Success", sheet_name)
            except Exception as e:
                self._update_excel_status(row_num, "Fail", str(e), sheet_name)
        workbook.close()

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, bill_no=None):
        try:
            wb = load_workbook(FILE_PATH); sh = wb[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sh.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sh.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if bill_no:
                col = 33 if sheet_name == "BillingReceipt" else 9 # Column for capture
                sh.cell(row=row_num, column=col, value=bill_no)
            wb.save(FILE_PATH); wb.close()
        except: pass

    def _extract_id_from_url(self):
        parts = self.driver.current_url.split('/')
        return parts[-1] if parts[-1].isdigit() else ""
