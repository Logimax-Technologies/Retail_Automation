from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class SearchBill(unittest.TestCase):
    """
    Search Bill Module Automation
    Handles searching, viewing, and cancelling existing billing records.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_search_bill(self):
        """Main entry point for Search Bill automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Billing List
        try:
            if "admin_ret_billing/billing/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Billing')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Search Bill')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_billing/billing/list")
            sleep(2)

        # Read Excel data
        sheet_name = "SearchBill"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Mapping based on SearchBillPrompt.md
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 11,
                "FromDate": 3, "ToDate": 4, "BillNo": 5,
                "Action": 6, "Remarks": 7, "UseOTP": 8, "OTP": 9,
                "VerifyStatus": 10
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            if str(row_data.get("TestStatus")).strip().lower() == "skip":
                print(f"⏭️ Skipping Test Case: {row_data['TestCaseId']}")
                continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                self.driver.refresh()
                sleep(2)
                
                result = self.execute_search_bill_flow(row_data, row_num, sheet_name)
                
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_search_bill_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Search Section"
        try:
            Function_Call.alert(self)
            
            # --- Date Range Filter ---
            if row_data.get("FromDate") and row_data.get("ToDate"):
                current_field = "Date Range"
                print(f"📅 Setting Date Range: {row_data['FromDate']} to {row_data['ToDate']}")
                
                # Directly setting hidden spans if possible, or using picker
                driver.execute_script(f"$('#rpt_from_date').html('{row_data['FromDate']}');")
                driver.execute_script(f"$('#rpt_to_date').html('{row_data['ToDate']}');")
                # Triggering the display update if needed by range
                # However, for automation, it's safer to click the picker if it affects CSRF or triggers
            else:
                # Default to "Today" via picker
                Function_Call.click(self, "//button[@id='rpt_date_picker']")
                sleep(1)
                Function_Call.click(self, "//li[contains(text(), 'Today')]")
                sleep(1)

            # --- Bill No Filter ---
            if row_data.get("BillNo"):
                current_field = "Bill No Filter"
                # Check if it's visible (might be commented out in some versions)
                try:
                    bill_input = driver.find_element(By.ID, "filter_bill_no")
                    if bill_input.is_displayed():
                        bill_input.clear()
                        bill_input.send_keys(str(row_data["BillNo"]))
                except:
                    print("⚠️ Bill No filter input not found or hidden.")

            # --- Click Search ---
            current_field = "Search Button"
            Function_Call.click(self, "//button[@id='bill_search']")
            sleep(3) # Wait for AJAX

            # --- Handle Action ---
            action = str(row_data.get("Action")).strip().lower()
            if action == "print":
                return self.handle_print_action(row_data)
            elif action == "cancel":
                return self.handle_cancel_action(row_data, row_num, sheet_name)
            else:
                # Default: Just verify existence
                return self.verify_presence(row_data)

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def handle_print_action(self, row_data):
        driver = self.driver
        bill_no = str(row_data.get("BillNo"))
        target_xpath = f"//table[@id='billing_list']/tbody/tr[contains(., '{bill_no}')]//a[contains(@class, 'btn-info')]"
        
        # If BillNo is not provided, take the first row
        if not row_data.get("BillNo"):
            target_xpath = "//table[@id='billing_list']/tbody/tr[1]//a[contains(@class, 'btn-info')]"
            
        print("🖨️ Clicking Print Receipt...")
        Function_Call.click(self, target_xpath)
        sleep(3)
        
        # Handle new tab
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])
            print(f"📄 Switched to print tab: {driver.current_url}")
            sleep(2)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return ("Pass", "Print receipt tab opened and closed successfully.")
        else:
            return ("Fail", "Print receipt tab did not open.")

    def handle_cancel_action(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        bill_no = str(row_data.get("BillNo"))
        
        # Find the row and click Cancel (Warning/Close button)
        cancel_btn_xpath = f"//table[@id='billing_list']/tbody/tr[contains(., '{bill_no}')]//button[contains(@class, 'btn-warning')]"
        if not row_data.get("BillNo"):
            cancel_btn_xpath = "//table[@id='billing_list']/tbody/tr[1]//button[contains(@class, 'btn-warning')]"
            
        print("🛑 Clicking Cancel Bill...")
        Function_Call.click(self, cancel_btn_xpath)
        sleep(2)
        
        # Handle Modal
        try:
            wait.until(EC.visibility_of_element_located((By.ID, "confirm-billcancell")))
            
            # Remarks
            if row_data.get("Remarks"):
                Function_Call.fill_input(self, wait, (By.ID, "cancel_remark"), str(row_data["Remarks"]), "Cancel Remark", row_num, Sheet_name=sheet_name)
                sleep(1)
            
            # OTP
            if row_data.get("UseOTP") == "Yes" and row_data.get("OTP"):
                Function_Call.fill_input(self, wait, (By.ID, "cancel_otp"), str(row_data["OTP"]), "Cancel OTP", row_num, Sheet_name=sheet_name)
                sleep(1)
                Function_Call.click(self, "//button[@id='verify_otp']")
                sleep(2)
            
            # Confirm
            Function_Call.click(self, "//button[@id='cancell_delete']")
            sleep(3)
            
            return self.verify_presence(row_data, expected_status="Cancelled")
            
        except Exception as e:
            return ("Fail", f"Cancellation modal error: {str(e)}")

    def verify_presence(self, row_data, expected_status=None):
        driver = self.driver
        bill_no = str(row_data.get("BillNo"))
        
        # Search for BillNo in datatable search box for quick isolation
        if row_data.get("BillNo"):
            try:
                search_box = driver.find_element(By.XPATH, "//div[@id='billing_list_filter']//input")
                search_box.clear()
                search_box.send_keys(bill_no)
                sleep(2)
            except:
                pass

        try:
            # Check if row exists
            row_xpath = "//table[@id='billing_list']/tbody/tr[1]"
            first_row_text = driver.find_element(By.XPATH, row_xpath).text
            
            if "No matching records found" in first_row_text:
                return ("Fail", f"Bill {bill_no} not found in list.")
            
            # Verify status if provided
            if expected_status:
                status_text = driver.find_element(By.XPATH, row_xpath + "/td[10]").text.strip()
                if expected_status.lower() in status_text.lower():
                    return ("Pass", f"Bill found and verified with status: {status_text}")
                else:
                    return ("Fail", f"Status mismatch. Expected: {expected_status}, Found: {status_text}")
            
            return ("Pass", "Record present in billing list.")
        except Exception as e:
            return ("Fail", f"Presence verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=11, value=actual_status).font = Font(bold=True, color=color)
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass
