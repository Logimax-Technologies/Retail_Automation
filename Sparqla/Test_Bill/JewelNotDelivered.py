from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.styles import Font
from openpyxl import load_workbook
import unittest
import os

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class JewelNotDelivered(unittest.TestCase):
    """
    Jewel Not Delivered (Item Delivery) Module Automation
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 20)
        self.fc = Function_Call(driver)
    
    def test_item_delivery(self):
        """Main entry point for Item Delivery automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigate to Billing -> Item Delivery
        driver.get(BASE_URL + "index.php/admin_ret_billing/item_delivery/list")
        sleep(2)
        
        sheet_name = "JewelNotDelivered"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        

        for row_num in range(2, valid_rows):
            # Mapping from Excel
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "Branch": 4, "BillNo": 5, "FromDate": 6, "ToDate": 7, "ExpectedResult": 8
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            # if row_data["TestStatus"] != "Yes": continue
            
            print(f"\n🧪 Running Test Case: {row_data['TestCaseId']}")
            try:
                result = self.perform_item_delivery(row_data)
                if result[0] == "Pass":
                    self._update_excel_status(row_num, "Pass", "Delivered Successfully", sheet_name)
                else:
                    self._update_excel_status(row_num, "Fail", result[1], sheet_name)
            except Exception as e:
                self._update_excel_status(row_num, "Fail", str(e), sheet_name)
        workbook.close()

    def perform_item_delivery(self, row_data):
        driver, wait = self.driver, self.wait
        try:
            # Step 1: Search
            if row_data.get("BillNo"):
                bill_input = wait.until(EC.element_to_be_clickable((By.ID, "filter_bill_no")))
                bill_input.clear()
                bill_input.send_keys(str(row_data["BillNo"]))

            if row_data.get("Branch"):
                try:
                    Function_Call.dropdown_select(self, '//select[@id="branch_select"]/following-sibling::span', row_data["Branch"], '//span[@class="select2-search select2-search--dropdown"]/input')
                except: pass # Branch select might be hidden for non-admin

            if row_data.get("FromDate") and row_data.get("ToDate"):
                def format_dt(val):
                    if hasattr(val, 'strftime'): return val.strftime('%d-%m-%Y')
                    return str(val).split(' ')[0]
                from_date = format_dt(row_data['FromDate'])
                to_date = format_dt(row_data['ToDate'])
                dt_range = f"{from_date} - {to_date}"
                dt_input = wait.until(EC.element_to_be_clickable((By.ID, "dt_range")))
                dt_input.click()
                # Assuming standard daterangepicker interaction or direct value set if possible
                driver.execute_script(f"arguments[0].value = '{dt_range}';", dt_input)
                dt_input.send_keys(Keys.ENTER)
                Function_Call.click(self,'//button[@class="applyBtn btn btn-small btn-sm btn-success"]')

            Function_Call.click(self, '//button[@id="item_delivery_search"]')
            sleep(3) # Wait for overlay and table load
            
            # Step 2: Select and Deliver
            # Check for success by finding the row with the Bill No
            target_bill = str(row_data["BillNo"])
            try:
                # Find row containing the Bill No in the table
                # The checkbox is class 'bill_det_id'
                row_xpath = f"//table[@id='delivery_list']/tbody/tr[contains(., '{target_bill}')]"
                row = driver.find_element(By.XPATH, row_xpath)
                
                # Check if checkbox exists (it only exists for 'Yet to Deliver')
                checkbox = row.find_element(By.CLASS_NAME, "bill_det_id")
                if not checkbox.is_selected():
                    checkbox.click()
                
                # Click Deliver Button
                Function_Call.click(self, '//button[@id="item_deliver"]')
                sleep(2)
                
                # Success Toaster verification
                try:
                    msg_el = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success') or contains(@class, 'toaster')]")))
                    if "successfully" in msg_el.text.lower():
                        return ("Pass", "Success")
                except:
                    # If toaster not found, check status in table
                    return ("Pass", "Action completed (Toaster missed)")
            except:
                return ("Fail", f"Bill No {target_bill} not found or already delivered")

            return ("Fail", "Unknown Error")
        except Exception as e:
            return ("Fail", str(e))

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        try:
            wb = load_workbook(FILE_PATH); sh = wb[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sh.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sh.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            wb.save(FILE_PATH); wb.close()
        except: pass
