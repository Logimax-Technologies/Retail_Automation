from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.styles import Font
from openpyxl import load_workbook
import re
import unittest
import os
from datetime import datetime

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class BillingIssue(unittest.TestCase):
    """
    Billing Issue Module Automation
    Follows Sparqla framework rules.
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_billing_issue(self):
        """Main entry point for Billing Issue automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigate to Billing -> Issue
        try:
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
            Function_Call.click(self, "//span[contains(text(), 'Billing')]")
            Function_Call.click(self, "//span[text()='Issue']/parent::a")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}. Using base URL fallback.")
           
            driver.get(BASE_URL + "index.php/admin_ret_billing/issue/list")
           
        
        # Read Excel data
        sheet_name = "BillingIssue"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows):
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "Branch": 4, "IssueTo": 5, "IssueType": 6, "AccountHead": 7,
                "Name": 8, "Amount": 9, "ReferenceNo": 10, "Employee": 11,
                "Narration": 12, "CashAmount": 13, 
                "Cheque(Yes/No)": 14, "ChequeDate": 15, "Bank": 16, "ChequeNo": 17, "ChequeAmount": 18, "ChequeAction": 19,
                "NB(Yes/No)": 20, "NBType": 21, "NBBankDevice": 22, "NBPaymentDate": 23, "NBRefNo": 24, "NBAmount": 25, "NBAction": 26,
                "CapturedBillNo": 27, "FromDate": 28, "ToDate": 29
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # if row_data["TestStatus"] != "Yes":
            #     continue
            
            print(f"\n🧪 Running Test Case: {row_data['TestCaseId']}")
            
            try:
                result = self.test_issue_creation(row_data, row_num, sheet_name)
                print(f"🏁 Creation Result: {result[0]} - {result[1]}")
                
                if result[0] == "Pass":
                    captured_id = result[2] if len(result) > 2 else ""
                    bill_no = self.test_issue_list_verification(
                        captured_id, 
                        row_data.get("FromDate"), 
                        row_data.get("ToDate")
                    )
                    self._update_excel_status(row_num, "Pass", "Success", sheet_name, bill_no)
                else:
                    self._update_excel_status(row_num, "Fail", result[1], sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", str(e), sheet_name)
        
        workbook.close()
    
    def test_issue_creation(self, row_data, row_num, sheet_name):
        """Core logic for filling and saving the Issue form"""
        driver = self.driver
        wait = self.wait
        
        try:
            # Click ADD button
            Function_Call.click(self, '//a[@id="add_billing"]')
            sleep(2)
            
            # 1. Branch (if exists)
            if row_data.get("Branch"):
                try:
                    Function_Call.dropdown_select(
                        self, '//select[@id="branch_select"]/following-sibling::span',
                        row_data["Branch"],
                        '//span[@class="select2-search select2-search--dropdown"]/input'
                    )
                except:
                    pass # Ignore if not multiple branches

            # 2. Issue To
            if row_data.get("IssueTo"):
                Function_Call.dropdown_select(
                    self, '//select[@id="issue_to"]/following-sibling::span',
                    row_data["IssueTo"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 3. Issue Type
            if row_data.get("IssueType"):
                Function_Call.dropdown_select(
                    self, '//select[@id="issue_type"]/following-sibling::span',
                    row_data["IssueType"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 4. Account Head
            if row_data.get("IssueType") == "Expenses":
                if row_data.get("AccountHead"):
                    try:
                        Function_Call.dropdown_select(
                            self, '//select[@id="acc_head"]/following-sibling::span',
                            row_data["AccountHead"],
                            '//span[@class="select2-search select2-search--dropdown"]/input'
                        )
                        sleep(1)
                    except Exception as e:
                        print(f"⚠️ Could not select Account Head: {e}")

            # 5. Name (Autocomplete)
            if row_data.get("Name"):
                Function_Call.fill_autocomplete_field(self, "name", row_data["Name"])
                sleep(1)

            # 6. Amount
            if row_data.get("Amount"):
                Function_Call.fill_input(
                    self, wait, locator=(By.ID, "issue_amount"),
                    value=str(row_data["Amount"]), field_name="Amount", 
                    row_num=row_num, Sheet_name=sheet_name
                )

            # 7. Employee
            if row_data.get("Employee"):
                Function_Call.dropdown_select(
                    self, '//select[@id="emp_select"]/following-sibling::span',
                    row_data["Employee"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )

            # 8. Narration
            if row_data.get("Narration"):
                Function_Call.fill_input(
                    self, wait, locator=(By.ID, "narration"),
                    value=row_data["Narration"], field_name="Narration",
                    row_num=row_num, Sheet_name=sheet_name
                )

            # 9. Payment Sections
            cash = float(row_data.get("CashAmount") or 0)
           
            
            # Cash
            if cash > 0:
                Function_Call.fill_input(
                    self, wait, locator=(By.ID, "cash_pay"),
                    value=str(cash), field_name="Cash",
                    row_num=row_num, Sheet_name=sheet_name
                )
            
            # Cheque
            if row_data.get("Cheque(Yes/No)") == "Yes":
                self.Cheque(row_data, row_num, sheet_name)
            
            # Net Banking
            if row_data.get("NB(Yes/No)") == "Yes":
                self.NetBanking(row_data, row_num, sheet_name)
            
            # 10. Save
            original_window = driver.current_window_handle
            Function_Call.click(self, '//button[@id="save_issue"]')
            sleep(3)
            
            # Extract ID from Print URL (Handle new tab)
            captured_id = ""
            try:
                # Wait for new window to open
                wait.until(EC.number_of_windows_to_be(2))
                for window_handle in driver.window_handles:
                    if window_handle != original_window:
                        driver.switch_to.window(window_handle)
                        captured_id = self._extract_id_from_url()
                        print(f"🔗 Extracted ID from new tab URL: {captured_id}")
                        driver.close()
                        driver.switch_to.window(original_window)
                        break
            except Exception as e:
                print(f"⚠️ Window handling error or no new tab: {e}")
                captured_id = self._extract_id_from_url() # Fallback to current window

            # Verify success
            try:
                msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text
                if "successfully" in msg.lower():
                    return ("Pass", "Issue Given successfully", captured_id)
                else:
                    return ("Fail", f"Unexpected message: {msg}")
            except:
                return ("Fail", "Success message not found")

        except Exception as e:
            return ("Fail", str(e))

    def test_issue_list_verification(self, captured_id="", from_date=None, to_date=None):
        """Verifies the entry in the list and returns the Bill No"""
        driver = self.driver
        wait = self.wait
        try:
            # 1. Apply Date Range Filter if provided
            if from_date and to_date:
                print(f"📅 Applying date range: {from_date} to {to_date}")
                try:
                    # Click date range button/input to open picker
                    dt_btn = None
                    dt_btn = driver.find_element(By.XPATH, '//input[@id="dt_range"]')
                    if dt_btn.is_displayed():
                        pass
                    if dt_btn:
                        dt_btn.click()
                        sleep(1)
                        # Fill From and To dates
                        from_input = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@name="daterangepicker_start"]')))
                        from_input.clear()
                        from_input.send_keys(str(from_date))
                        
                        to_input = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@name="daterangepicker_end"]')))
                        to_input.clear()
                        to_input.send_keys(str(to_date))
                        
                        # Click Apply
                        Function_Call.click(self, '//button[contains(@class, "applyBtn")]')
                        sleep(1)
                        # Click Search if button exists
                        try:
                            Function_Call.click(self, '//button[@id="search_button"] | //button[text()="Search"]')
                        except:
                            pass
                        sleep(2)
                except Exception as e:
                    print(f"⚠️ Date range filter error: {e}")

            # 2. Search by ID if available, otherwise stay in list search
            search_input = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@type="search"]')))
            search_input.clear()
            search_input.send_keys(captured_id)
            search_input.send_keys(Keys.ENTER)
            sleep(2)
            
            # Get first row voucher no (usually column 2)
            bill_no = driver.find_element(By.XPATH, '//table[contains(@id, "list")]/tbody/tr[1]/td[3]').text
            print(f"✅ Captured Bill No: {bill_no}")
            return bill_no
        except:
            print("⚠️ Could not verify in list")
            return "Unknown"
        
    def Cheque(self, row_data, row_num, Sheet_name):
        driver = self.driver
        wait = self.wait
        
        Function_Call.click(self, '//a[@id="cheque_modal"]')
        sleep(2)
        
        if row_data.get("ChequeDate"):
            Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//input[@name="cheque_details[cheque_date][]"]'),
                value=row_data["ChequeDate"],
                pattern=r"^(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-\d{4}$",
                field_name="ChequeDate",
                screenshot_prefix="ChequeDate",
                row_num=row_num,
                Sheet_name=Sheet_name,
                extra_keys=Keys.TAB,
                Date_range="Yes"
            )

        # Use index [1] since row is undefined
        Function_Call.select_visible_text(self, f"(//select[@name='cheque_details[id_bank][]'])[1]", row_data.get("Bank", "IOB 4795"))
        
        if row_data.get("ChequeNo"):
            Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//input[@name="cheque_details[cheque_no][]"]'),
                value=row_data["ChequeNo"],
                pattern=r"\d{1,11}?$",
                field_name="ChequeNo",
                screenshot_prefix="ChequeNo",
                row_num=row_num,
                Sheet_name=Sheet_name
            )
        
        
        if row_data.get("ChequeAmount"):
            chq_amt = row_data["ChequeAmount"]
            # try:
            #     # If it's a percentage (less than or equal to 100), calculate it
            #     val = float(row_data["ChequeAmount"])
            #     if val <= 100:
            #         chq_amt = float((Received * val) / 100)
            #     else:
            #         chq_amt = val
            # except:
            #     pass

            Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//input[@name="cheque_details[payment_amount][]"]'),
                value=chq_amt,
                pattern=r"\d{1,11}?$",
                field_name="Amount",
                screenshot_prefix="Amount",
                row_num=row_num,
                Sheet_name=Sheet_name
            )
        
        # Save & Close
        Function_Call.click(self, '//a[@id="save_issue_chq"]')
        sleep(1)
        
        return
        
    def NetBanking(self, row_data, row_num, Sheet_name):
        driver = self.driver
        wait = self.wait
        
        Function_Call.click(self, '//a[@id="net_bank_modal"]')
        sleep(2)
                
        Function_Call.select_visible_text(self, "(//select[contains(@id, 'nb_type')])[1]", row_data.get("NBType", "Online"))
        Function_Call.select_visible_text(self, "(//select[@name='nb_details[nb_bank][]'])[1]", row_data.get("NBBankDevice", "IOB 4795"))
        
        if row_data.get("NBPaymentDate"):
            Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//input[@name="nb_details[nb_date][]"]'),
                value=row_data["NBPaymentDate"],
                pattern=r"^(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-\d{4}$",
                field_name="PaymentDate",
                screenshot_prefix="PaymentDate",
                row_num=row_num,
                Sheet_name=Sheet_name,
                extra_keys=Keys.TAB,
                Date_range="Yes"
            )
        
        if row_data.get("NBRefNo"):
            Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//input[@name="nb_details[ref_no][]"]'),
                value=row_data["NBRefNo"],
                pattern=r"\d{1,11}?$",
                field_name="RefNo",
                screenshot_prefix="RefNo",
                row_num=row_num,
                Sheet_name=Sheet_name
            )
        
        
        if row_data.get("NBAmount"):
            nb_amt = row_data["NBAmount"]
            # try:
            #     val = float(row_data["NBAmount"])
            #     if val <= 100:
            #         nb_amt = float((Received * val) / 100)
            #     else:
            #         nb_amt = val
            # except:
            #     pass

            Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//input[@name="nb_details[amount][]"]'),
                value=nb_amt,
                pattern=r"\d{1,11}?$",
                field_name="Amount",
                screenshot_prefix="Amount",
                row_num=row_num,
                Sheet_name=Sheet_name
            )
        
        # Save & Close
        Function_Call.click(self, '//a[@id="save_issue_net_banking"]')
        sleep(1)
        return
        
    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, bill_no=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if bill_no:
                # Store captured Ref No in column 21
                sheet.cell(row=row_num, column=27, value=bill_no)
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update failed: {e}")  
            

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass

    def _extract_id_from_url(self):
        url = self.driver.current_url
        parts = url.split('/')
        if parts[-1].isdigit():
            return parts[-1]
        return ""
       
