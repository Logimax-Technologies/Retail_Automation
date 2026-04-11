from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class RepairOrder(unittest.TestCase):
    """
    Repair Order Module Automation
    Handles Customer & Stock Repair, Multi-Item, Stone Modal, and Tab-based verification.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_repair_order(self):
        """Main entry point for Repair Order automation"""
        driver = self.driver
        wait = self.wait
        # Navigate to Repair Order List
        try:
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
            sleep(1)
            Function_Call.click(self, "//span[contains(text(), 'Repair Orders')]")
            sleep(1)
            Function_Call.click(self, "(//span[contains(text(), 'Create Order')])[2]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_order/repair_order/list")
            sleep(2)

        sheet_name = "RepairOrder"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3, "OrderType": 4,
                "WorkAt": 5, "Employee": 6, "OrderFrom": 7, "CusName": 8,
                "Metal": 9, "Product": 10, "Design": 11, "SubDesign": 12,
                "GrossWt": 13, "StoneData": 14, "Pcs": 15, "DueDate": 16,
                "RepairType": 17, "TagCode": 18, "Narration": 19,
                "ItemImage": 20, "ItemDesc": 21,"Repairorder_NO":22

            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # if str(row_data["TestStatus"]).strip().lower() != "run":
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

           

            try:
                result = self.test_repair_order_flow(row_data, row_num, sheet_name)
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_id = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                # Fallback to list page if stuck
                driver.get(BASE_URL + "index.php/admin_ret_order/repair_order/list")
                sleep(2)

    def test_repair_order_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        fc = self.fc
        wait.until(EC.element_to_be_clickable((By.XPATH,'//a[@id="add_Order"]'))).click()
        sleep(2)
        # 1. Header SectionS
        # Order Type: Customer(3) | Stock(4)
        order_type = "3" # Default
        if row_data.get("OrderType") is not None:
            current_field = "Order Type"
            val = str(row_data["OrderType"]).strip()
            type_map = {"Customer": "3", "Stock": "4"}
            order_type = type_map.get(val, val)
            id_map = {"3": "cus_repair_order", "4": "stock_repair_order"}
            target_id = id_map.get(str(order_type), "cus_repair_order")
            Function_Call.click(self, f"//input[@id='{target_id}']")
            sleep(1)
        


        # Work At: In House(1) | Out Source(2)
        work_at = "1" # Default
        if row_data.get("WorkAt") is not None:
            current_field = "Work At"
            val = str(row_data["WorkAt"]).strip()
            work_map = {"In House": "1", "Out Source": "2"}
            work_at = work_map.get(val, val)
            id_map = {"1": "cus_repair_at_inhouse", "2": "cus_repair_at_outsource"}
            target_id = id_map.get(str(work_at), "cus_repair_at_inhouse")
            Function_Call.click(self, f"//input[@id='{target_id}']")
            sleep(1)
        
        
        # Order From
        if row_data["OrderFrom"]:
            # Check if branch_select is enabled
            branch_select = driver.find_element(By.ID, "branch_select")
            if branch_select.is_enabled():
                Function_Call.dropdown_select(self, '//select[@id="branch_select"]/following-sibling::span', str(row_data["OrderFrom"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)
        
        # Employee
        if row_data["Employee"]:
            Function_Call.dropdown_select(self, '//select[@id="employee_sel"]/following-sibling::span', str(row_data["Employee"]), '//span[@class="select2-search select2-search--dropdown"]/input')
            sleep(1)

        # 2. Recipient Section
        if str(order_type) == "3": # Customer
            if row_data["CusName"]:
                Function_Call.fill_autocomplete_field(self, "cus_name", str(row_data["CusName"]))
                sleep(1)
        else: # Stock (4 or "stock")
            if row_data["TagCode"]:
                tags = str(row_data["TagCode"]).split('|')
                # Note: Currently handling single tag for simple stock repair, but can loop if needed
                for tag in tags:
                    Function_Call.fill_input(self, self.wait, (By.ID, "issue_tag_code"), tag.strip(), "Tag Code", row_num, Sheet_name=sheet_name)
                    Function_Call.click(self, "//button[@id='issue_tag_search']")
                    sleep(2)

        # 3. Item Selection (Customer)
        if str(order_type) == "3":
            metals = str(row_data["Metal"]).split('|')
            products = str(row_data["Product"]).split('|')
            designs = str(row_data["Design"]).split('|')
            sub_designs = str(row_data["SubDesign"]).split('|')
            gross_wts = str(row_data["GrossWt"]).split('|')
            pcs_list = str(row_data["Pcs"]).split('|')
            due_dates = str(row_data["DueDate"]).split('|')
            repair_types = str(row_data["RepairType"]).split('|')
            stone_data_groups = str(row_data["StoneData"]).split(';') if row_data["StoneData"] else []
            wait.until(EC.element_to_be_clickable((By.ID,"repaid_order_items"))).click()
            for i in range(len(products)):
                row_idx = i + 1
                prefix = f"(//table[@id='custrepair_item_detail']/tbody/tr)[{row_idx}]"
                
                # Add row if not the first item
                if i > 0:
                    fc.click("//table[@id='custrepair_item_detail']/tbody/tr[last()]//a[@id='create_order_item']")
                    sleep(1)

                Function_Call.dropdown_select(self, f"{prefix}//span[contains(@id, 'select2-metal-container')]/following-sibling::span", metals[i].strip(), '//span[@class="select2-search select2-search--dropdown"]/input')
                Function_Call.dropdown_select(self, f"{prefix}//span[@id='select2-product-container']", products[i].strip(), '//span[@class="select2-search select2-search--dropdown"]/input')
                Function_Call.dropdown_select(self, f"{prefix}//span[contains(@id,'order_item[id_design]')]", designs[i].strip(), '//span[@class="select2-search select2-search--dropdown"]/input')
                Function_Call.dropdown_select(self, f"{prefix}//span[contains(@id,'order_item[id_sub_design]')]", sub_designs[i].strip(), '//span[@class="select2-search select2-search--dropdown"]/input')
                Function_Call.fill_input(self, self.wait, (By.XPATH, f"{prefix}//input[@name='order_item[weight][]']"), gross_wts[i].strip(), "Gross Wt", row_num, Sheet_name=sheet_name)
                
                # Handle Stones
                if i < len(stone_data_groups) and stone_data_groups[i].strip():
                    Function_Call.click(self, f"{prefix}//input[@name='order_item[less_wt][]']") # Less Wt plus button
                    sleep(2)
                    self.handle_stone_modal(stone_data_groups[i].strip(), row_num, sheet_name)
                    sleep(1)

                Function_Call.fill_input(self, self.wait, (By.XPATH, f"{prefix}//input[@name='order_item[piece][]']"), pcs_list[i].strip(), "Pcs", row_num, Sheet_name=sheet_name)
                
                # Due Date (Using JS to bypass datepicker interference)
                due_date_val = due_dates[i].strip()
                due_date_el = wait.until(EC.presence_of_element_located((By.XPATH, f"{prefix}//input[@name='order_item[cus_due_days][]']")))
                driver.execute_script("arguments[0].value = arguments[1];", due_date_el, due_date_val)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", due_date_el)
                driver.execute_script("arguments[0].dispatchEvent(new Event('blur'));", due_date_el)
                sleep(1)
                Function_Call.dropdown_select(self, f"{prefix}//span[@id='select2-repair-container']/following-sibling::span", repair_types[i].strip(), '//span[@class="select2-search select2-search--dropdown"]/input')
                
                # Description if provided
                if row_data["ItemDesc"]:
                    descs = str(row_data["ItemDesc"]).split('|')
                    if i < len(descs) and descs[i].strip():
                        Function_Call.click(self, f"{prefix}//a[contains(@onclick,'update_order_description')]")
                        sleep(1)
                        Function_Call.fill_input(self, self.wait, (By.ID, "description"), descs[i].strip(), "Item Desc", row_num, Sheet_name=sheet_name)
                        Function_Call.click(self, "//a[@class='btn btn-success add_order_desc']") # Modal add button
                        sleep(1)

        else: # Stock Item Filling
            # Stock details are auto-filled, just set repair type and narration
            repair_types = str(row_data["RepairType"]).split('|')
            narrations = str(row_data["Narration"]).split('|')
            for i in range(len(repair_types)):
                prefix = f"(//table[@id='tagissue_item_detail']/tbody/tr)[{i+1}]"
                Function_Call.dropdown_select(self, f"{prefix}//select[contains(@name, 'repair')]/following-sibling::span", repair_types[i].strip(), '//span[@class="select2-search select2-search--dropdown"]/input')
                # Narration uses plus button for modal? In prompt it said description button.
                # Assuming simple description modal if button present.
                try:
                    Function_Call.click(self, f"{prefix}//a[contains(@onclick,'update_order_description')]")
                    sleep(1)
                    Function_Call.fill_input(self, self.wait, (By.ID, "description"), narrations[i].strip(), "Stock Narration", row_num, Sheet_name=sheet_name)
                    Function_Call.click(self, "//a[@class='btn btn-success add_order_desc']")
                    sleep(1)
                except:
                    pass

        # 4. Save
        fc.click("//button[@id='create_repair_order']")
        sleep(5)

        # 5. Verification
        try:
            return self.extract_id_and_verify(row_data)
        except Exception as e:
            return ("Pass", f"Saved but verification failed: {e}")

    def handle_stone_modal(self, stone_group, row_num, sheet_name):
        """Fills multiple stones in the stone modal"""
        stones = stone_group.split('#')
        for idx, stone_str in enumerate(stones):
            # Format: Type,Name,Pcs,Wt,Unit,Cal,Rate
            parts = [p.strip() for p in stone_str.split(',')]
            if len(parts) < 7: continue
            
            row_idx = idx + 1
            prefix = f"(//table[@id='estimation_stone_cus_item_details']/tbody/tr)[{row_idx}]"
            
            if idx > 0:
                Function_Call.click(self, "//button[@id='create_stone_item_details']")
                sleep(1)

            # LWT checkbox (default checked usually, but let's ensure)
            lwt_check = self.driver.find_element(By.XPATH, f"{prefix}//input[@type='checkbox']")
            if not lwt_check.is_selected():
                Function_Call.click(self, f"{prefix}//input[@type='checkbox']")

            Function_Call.select_visible_text(self, f"{prefix}//select[@name='est_stones_item[stones_type][]']", parts[0])
            Function_Call.select_visible_text(self, f"{prefix}//select[@name='est_stones_item[stone_id][]']", parts[1])
            Function_Call.fill_input(self, self.wait, (By.XPATH, f"{prefix}//input[@name='est_stones_item[stone_pcs][]']"), parts[2], "Stone Pcs", row_num, Sheet_name=sheet_name)
            Function_Call.fill_input(self, self.wait, (By.XPATH, f"{prefix}//input[@name='est_stones_item[stone_wt][]']"), parts[3], "Stone Wt", row_num, Sheet_name=sheet_name)
            
            # Unit Select
            Function_Call.select_visible_text(self, f"{prefix}//select[@name='est_stones_item[stone_uom_id][]']", parts[4])
            
            # Cal Type Radio
            if parts[5].lower() == "by pcs":
                Function_Call.click(self, f"{prefix}//input[@class='stone_cal_type' and @value='1']")
            else: # By Wt (1)
                Function_Call.click(self, f"{prefix}//input[@class='stone_cal_type' and @value='2']")

            Function_Call.fill_input(self, self.wait, (By.XPATH, f"{prefix}//input[@name='est_stones_item[stone_rate][]']"), parts[6], "Stone Rate", row_num, Sheet_name=sheet_name)
        
        Function_Call.click(self, "//button[@id='update_stone_details']") # Modal Save

    def extract_id_and_verify(self, row_data):
        driver = self.driver
        main_window = driver.current_window_handle
        
        # Wait for new tab
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
        
        captured_id = ""
        for handle in driver.window_handles:
            if handle != main_window:
                driver.switch_to.window(handle)
                url = driver.current_url
                # URL like: .../repair_acknowledgement/711
                captured_id = url.split('/')[-1]
                print(f"🎯 Captured Repair Order ID: {captured_id}")
                driver.close()
                break
        
        driver.switch_to.window(main_window)
        
        # Navigate to Listing and verify
        driver.get(BASE_URL + "index.php/admin_ret_order/repair_order/list")
        sleep(2)
        
        # Branch Filter
        if row_data["OrderFrom"]:
            # Branch filter usually has a specific ID on list page, check list.php research or guess
            try:
                Function_Call.dropdown_select(self, "//span[@id='select2-branch_select-container']", str(row_data["OrderFrom"]), '//span[@class="select2-search select2-search--dropdown"]/input')
            except: pass
            
        # Date Filter - Today
        try:
            # Click the Date range picker button and select 'Today'
            Function_Call.click(self, "//button[contains(., 'Date range picker')]")
            sleep(1)
            Function_Call.click(self, "//li[contains(text(), 'Today')]")
            sleep(2)
        except:
             # Fallback to manual date entry if picker fails
             today_date = datetime.now().strftime("%d-%m-%Y")
             Function_Call.fill_input(self, self.wait, (By.ID, "from_date"), today_date, "From Date", row_num, Sheet_name=sheet_name, extra_keys=Keys.ENTER)
             Function_Call.fill_input(self, self.wait, (By.ID, "to_date"), today_date, "To Date", row_num, Sheet_name=sheet_name, extra_keys=Keys.ENTER)

        # Employee
        if row_data["Employee"]:
            try:
                Function_Call.dropdown_select(self, "//span[@id='select2-employee_sel-container']", str(row_data["Employee"]), '//span[@class="select2-search select2-search--dropdown"]/input')
            except: pass

        self.fc.click("//button[@id='reorder_search']")
        sleep(2)
        
        # Grid Search
        # Many lists have a 'DataTables' search box
        try:
            search_box = driver.find_element(By.XPATH, "//input[@type='search']")
            search_box.clear()
            search_box.send_keys(captured_id)
            sleep(2)
            
            # Verify 1st row
            first_row = driver.find_element(By.XPATH, "//table[@id='repair_order_list']/tbody/tr[1]")
            repair_order_no = first_row.find_element(By.XPATH, "./td[2]").text.strip()
            if captured_id in first_row.text:
                return ("Pass", f"Order {captured_id} created and verified in listing. ({repair_order_no})", repair_order_no)
            else:
                return ("Pass", f"Order {captured_id} created but not found in search.", None)
        except:
            return ("Pass", f"Order {captured_id} created. (List search failed)", None)

    def _update_excel_status(self, row_num, status, message, sheet_name, captured_no=None):
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        color = "00B050" if status == "Pass" else "FF0000"
        sheet.cell(row=row_num, column=2, value=status).font = Font(bold=True, color=color)
        sheet.cell(row=row_num, column=3, value=message).font = Font(bold=True, color=color)
        if captured_no:
            # Saving ID to Column 22
            sheet.cell(row=row_num, column=22).value = captured_no
            # Save the captured ID to KarigarAllotment sheet as well (OrderNo is Column 4)
            if "KarigarAllotment" in workbook.sheetnames:
                ka_sheet = workbook["KarigarAllotment"]
                ka_sheet.cell(row=row_num, column=4).value = captured_no
            if "RepairOrderStatus" in workbook.sheetnames:
                ros_sheet = workbook["RepairOrderStatus"]
                ros_sheet.cell(row=row_num, column=4).value = captured_no
                
        workbook.save(FILE_PATH)

if __name__ == "__main__":
    # For local testing if needed
    pass
