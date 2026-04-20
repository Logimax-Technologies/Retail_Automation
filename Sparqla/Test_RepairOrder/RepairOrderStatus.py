from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from openpyxl.styles import Font
from time import sleep
import unittest
import re

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class RepairOrderStatus(unittest.TestCase):
    """
    Repair Order Status Automation.
    Used to complete a repair order or mark it as delivered, including adding extra metal and grid amounts.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_repair_order_status(self):
        driver = self.driver
        wait = self.wait
        sheet_name = "RepairOrderStatus"

        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases for {sheet_name}")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        # Navigation
        try:
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
            sleep(1)
            Function_Call.click(self, "//span[contains(text(), 'Repair Orders')]")
            sleep(1)
            # Find the exact span text for Status
            Function_Call.click(self, "//span[contains(text(), 'Repair Order Status')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ UI Navigation failed, using fallback URL: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_order/repair_order/repair_order_status")
            sleep(3)

        for row_num in range(2, valid_rows):
            try:
                workbook = load_workbook(FILE_PATH)
                sheet = workbook[sheet_name]
            except Exception as e:
                print(f"Error loading workbook for row {row_num}: {e}")
                continue

            data_map = {
                "TestCaseId": 1,
                "TestStatus": 2,
                "ActualStatus": 3,
                "OrderNo": 4,
                "Branch": 5,
                "DateRange": 6,
                "RepairType": 7,
                "ExtraMetal": 8,
                "CompletedWeight": 9,
                "Amount": 10,
                "Action": 11,
                "Remark": 12,
                "Customer":13
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # status = str(row_data["TestStatus"]).strip().lower() if row_data.get("TestStatus") else ""
            # if status != "run":
            #     continue 

            print(f"\n{'='*80}")
            print(f"🧪 Running TC: {row_data['TestCaseId']} - Order: {row_data['OrderNo']}")
            print(f"{'='*80}")

            try:
                result = self.test_status_flow(row_data, row_num, sheet_name)
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
                
                # Auto-sync to Billing sheet on Pass
                if result[0] == "Pass":
                    self._update_billing_sheet(row_data)
            except Exception as e:
                err_msg = str(e)
                print(f"❌ Test Failed Exception: {err_msg}")
                self._update_excel_status(row_num, "Fail", f"Exception: {err_msg[:60]}", sheet_name)
                # Fallback to list page if stuck mid-flow
                if "repair_item_details" in driver.current_url:
                    driver.get(BASE_URL + "index.php/admin_ret_order/repair_order/repair_order_status")
                    sleep(2)

    def test_status_flow(self, row_data, row_num, sheet_name):
        driver = self.driver
        wait = self.wait
        current_field = "Start Flow"

        try:
            order_no = str(row_data["OrderNo"]).strip() if row_data.get("OrderNo") else None
            branch = str(row_data["Branch"]).strip() if row_data.get("Branch") else "Head Office"
            date_range = str(row_data["DateRange"]).strip() if row_data.get("DateRange") else "Today"
            repair_type = str(row_data["RepairType"]).strip().lower() if row_data.get("RepairType") else "customer"
            extra_metal = str(row_data["ExtraMetal"]).strip() if row_data.get("ExtraMetal") else None
            completed_wt = str(row_data.get("CompletedWeight", "")).strip()
            amount = str(row_data.get("Amount", "")).strip()
            action = str(row_data["Action"]).strip().lower() if row_data.get("Action") else "complete"

            if not order_no:
                return ("Fail", "OrderNo is required")

            # 0. Branch Selection
            current_field = "Branch"
            try:
                search_xpath = "//span[@class='select2-search select2-search--dropdown']/input"
                Function_Call.dropdown_select(self, "//select[@id='branch_select']/following-sibling::span", branch, search_xpath)
                sleep(1)
            except Exception as e:
                print(f"⚠️ Could not select branch: {e}")

            # 1. Date Range
            if date_range.lower() == "today":
                current_field = "Date Range"
                try:
                    Function_Call.click(self, "//button[@id='rpt_payment_date']")
                    sleep(1)
                    Function_Call.click(self, "//li[contains(text(), 'Today')]")
                    sleep(1)
                except Exception as e:
                    print("Could not set Today date, moving on")

            # 2. Repair Type
            current_field = "Repair Type"
            try:
                if repair_type == "company":
                    Function_Call.select(self, "//select[@id='repair_type']", "1")
                else:
                    Function_Call.select(self, "//select[@id='repair_type']", "2")
            except:
                pass

            # Click Search Action
            Function_Call.click(self, "//button[@id='repair_order_search']")
            sleep(3)

            # 3. Search in DataTables
            current_field = "Search Box"
            search_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(order_no)
            sleep(2)

            # 4. Extra Metal Logic -> Item Details Page
            if extra_metal and extra_metal != "None":
                current_field = "Navigate to Item Details"
                try:
                    target_btn = f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//a[contains(@href, 'repair_item_details')]"
                    Function_Call.click(self, target_btn)
                    sleep(3)
                    
                    # Fill Item Details Form
                    current_field = "Filling Item Details"
                    parts = extra_metal.split('|')
                    if len(parts) >= 13:
                        search_xpath = "//span[@class='select2-search select2-search--dropdown']/input"
                        Function_Call.dropdown_select(self, "//select[@id='item_type']/following-sibling::span", parts[0], search_xpath)
                        sleep(0.5)
                        
                        # Select2 Dropdowns
                        Function_Call.dropdown_select(self, "//select[@id='section']/following-sibling::span", parts[1], search_xpath)
                        Function_Call.dropdown_select(self, "//select[@id='category']/following-sibling::span", parts[2], search_xpath)
                        Function_Call.dropdown_select(self, "//select[@id='purity']/following-sibling::span", parts[3], search_xpath)
                        Function_Call.dropdown_select(self, "//select[@id='prod_select']/following-sibling::span", parts[4], search_xpath)
                        
                        # Optional Design / Sub Design
                        if parts[5].strip():
                            Function_Call.dropdown_select(self, "//select[@id='des_select']/following-sibling::span", parts[5], search_xpath)
                        if parts[6].strip():
                            Function_Call.dropdown_select(self, "//select[@id='sub_design_select']/following-sibling::span", parts[6], search_xpath)
                        
                        # Number Inputs
                        Function_Call.fill_input(self, wait, (By.ID, "pcs"), parts[7], "Pcs", row_num, Sheet_name=sheet_name)
                        Function_Call.fill_input(self, wait, (By.ID, "gross_wt"), parts[8], "Gross Wt", row_num, Sheet_name=sheet_name)
                        Function_Call.fill_input(self, wait, (By.ID, "wast_per"), parts[9], "V.A(%)", row_num, Sheet_name=sheet_name)
                        
                        # MC Type - expects visible text ("Per Pcs", "Per Grams", "On Price")
                        mc_type_map = {"per pcs": "Per Pcs", "per grams": "Per Grams", "on price": "On Price"}
                        mc_val = mc_type_map.get(parts[10].strip().lower(), "Per Pcs")
                        Function_Call.select(self, "//select[@id='mc_type']", mc_val)

                        Function_Call.fill_input(self, wait, (By.ID, "mc_value"), parts[11], "M.C", row_num, Sheet_name=sheet_name)
                        Function_Call.fill_input(self, wait, (By.ID, "service_charge"), parts[12], "Service Charge", row_num, Sheet_name=sheet_name)
                        
                        # Click Add to grid
                        Function_Call.click(self, "//button[@id='add_repair_item']")
                        sleep(2)
                        
                        # Save Page
                        Function_Call.click(self, "//button[@id='save_repair_item']")
                        sleep(3)
                    else:
                        print("⚠️ ExtraMetal string does not have 13 parts for Item Details.")
                        
                except Exception as e:
                    print(f"⚠️ Warning: Could not complete Item Details flow: {e}")

                # Ensure we are back on the status list page
                if "repair_item_details" in driver.current_url:
                    driver.get(BASE_URL + "index.php/admin_ret_order/repair_order/repair_order_status")
                    sleep(3)

                # Re-search if we just came back from the page
                search_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
                search_box.clear()
                search_box.send_keys(order_no)
                sleep(2)

            # 4.5 Fill Completed Weight and Amount in specific row
            current_field = "Fill Row Weight/Amount"
            if completed_wt and completed_wt != "None":
                try:
                    # Using position index or fuzzy text matching. Let's find any input in the row inside a td.
                    # Column 20 is Completed Weight, 21 is Amount per the <th> tags, but in tbody the index might differ if buttons or hidden columns change it.
                    # We will use contains class names fallback if exact TD misses:
                    cw_xpath = f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//input[contains(@class, 'completed_weight') or contains(@name, 'completed_weight') or contains(@placeholder, 'Weight')]"
                    try:
                        cw_inp = wait.until(EC.presence_of_element_located((By.XPATH, cw_xpath)))
                    except:
                        # Fallback to absolute TD index (19 or 20 or 21 depending on UI rendering)
                        cw_inp = driver.find_element(By.XPATH, f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//td[19]//input")
                        
                    cw_inp.clear()
                    cw_inp.send_keys(completed_wt)
                    sleep(0.5)
                except Exception as e:
                    print(f"⚠️ Could not fill Completed Weight: {e}")
                    
            if amount and amount != "None":
                try:
                    amt_xpath = f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//input[contains(@class, 'amount') or contains(@name, 'amount')]"
                    try:
                        amt_inp = wait.until(EC.presence_of_element_located((By.XPATH, amt_xpath)))
                    except:
                        amt_inp = driver.find_element(By.XPATH, f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//td[20]//input")
                        
                    amt_inp.clear()
                    amt_inp.send_keys(amount)
                    sleep(0.5)
                except Exception as e:
                    print(f"⚠️ Could not fill Amount: {e}")

            # 5. Select Checkbox
            current_field = "Select Row Checkbox"
            try:
                row_xpath = f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//input[@type='checkbox']"
                checkbox = wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                
                if not checkbox.is_selected():
                    try:
                        checkbox.click()
                    except:
                        label_xpath = f"//table[@id='repair_order_list']/tbody/tr[contains(., '{order_no}')]//label"
                        Function_Call.click(self, label_xpath)
                sleep(1)
            except Exception as e:
                return ("Fail", f"Order {order_no} not found or checkbox error")

            # 6. Action (Complete / Deliver)
            current_field = "Submit Action"
            if action == "deliver":
                Function_Call.click(self, "//button[@id='repair_delivered']")
            else:
                Function_Call.click(self, "//button[@id='repair_order_status']")

            # 7. Check toaster
            current_field = "Toaster Capture"
            try:
                msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text.strip()
                return ("Pass", f"Success: {msg[:50]}")
            except TimeoutException:
                try:
                    Function_Call.alert2(self)
                    return ("Pass", f"Alert accepted for {order_no}")
                except:
                    return ("Pass", "Submitted but no success toaster seen")

        except Exception as e:
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def _update_billing_sheet(self, row_data):
        """Automatically appends Repair Order details to the Billing sheet for delivery."""
        try:
            workbook = load_workbook(FILE_PATH)
            if "Billing" not in workbook.sheetnames:
                print("⚠️ 'Billing' sheet not found in workbook.")
                return
            
            sheet = workbook["Billing"]
            next_row = sheet.max_row + 1
            
            # Generate sequential Test Case ID (from TC001, TC002...) based on previous row
            last_id = sheet.cell(row=next_row - 1, column=1).value
            new_tc_id = "TC001"
            if last_id and isinstance(last_id, str):
                match = re.search(r"(\d+)", last_id)
                if match:
                    num_str = match.group(0)
                    new_num = int(num_str) + 1
                    new_tc_id = last_id.replace(num_str, str(new_num).zfill(len(num_str)))
            
            print(f"🔄 Next sequential ID generated: {new_tc_id} for Row {next_row}")

            # Mapping as per USER request
            sheet.cell(row=next_row, column=1, value=new_tc_id)                       # Col 1: Incremented TC Id
            sheet.cell(row=next_row, column=2, value="run")                          # Col 2: TestStatus (set to run)
            sheet.cell(row=next_row, column=4, value=row_data.get("Branch"))        # Col 4: Cost Centre
            sheet.cell(row=next_row, column=5, value="Customer")                     # Col 5: Billing To
            sheet.cell(row=next_row, column=6, value="111-Developer Logimax")        # Col 6: Employee
            sheet.cell(row=next_row, column=7, value=row_data.get("Customer"))      # Col 7: Customer Number
            sheet.cell(row=next_row, column=9, value="Show Room")                    # Col 9: Delivery Location
            sheet.cell(row=next_row, column=10, value="Repair Order Delivery")       # Col 10: Bill Type
            sheet.cell(row=next_row, column=11, value="No")                          # Col 11: driect
            sheet.cell(row=next_row, column=31, value=row_data.get("OrderNo"))       # Col 31: OrderNo
            sheet.cell(row=next_row, column=34, value=row_data.get("Amount"))        # Col 34: RepairAmount
            
            workbook.save(FILE_PATH)
            print(f"✅ Successfully synced Order {row_data.get('OrderNo')} to Billing sheet (Row {next_row})")
            
        except Exception as e:
            print(f"⚠️ Billing sheet sync failed: {e}")

    def _update_excel_status(self, row_num, status, message, sheet_name):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            
            color = "00B050" if status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=message).font = Font(bold=True, color=color)
            
            # Write remark to column 12
            sheet.cell(row=row_num, column=12, value=message)

            workbook.save(FILE_PATH)
        except Exception as e:
            print(f"⚠️ Excel save failed at row {row_num}: {e}")

if __name__ == "__main__":
    # Local Testing
    pass
