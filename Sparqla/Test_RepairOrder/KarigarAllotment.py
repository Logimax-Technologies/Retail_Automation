from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from openpyxl.styles import Font
from time import sleep
import unittest

FILE_PATH = ExcelUtils.file_path

class KarigarAllotment(unittest.TestCase):
    """
    Karigar Allotment Module Automation.
    Used to assign specific orders to a Karigar or Employee in bulk list page.
    Follows Sparqla framework standards.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_karigar_allotment(self):
        driver = self.driver
        wait = self.wait
        sheet_name = "KarigarAllotment"

        # Read Excel data
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases for {sheet_name}")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        # Navigate to New Orders explicitly at start of all cases
        try:
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
            Function_Call.click(self, "//span[contains(text(), 'Repair Orders')]")
            Function_Call.click(self, "(//span[contains(text(), 'Karigar Allotment')])[2]")
            sleep(3)
        except Exception as e:
            # print(f"⚠️ Initial UI Navigation Warning: {e}")
            print(f"⚠️ Navigation failed: {e}")
            driver.get(ExcelUtils.BASE_URL + "index.php/admin_ret_order/repair_order/neworders")
               


        for row_num in range(2, valid_rows):
            # Reload workbook per row
            try:
                workbook = load_workbook(FILE_PATH)
                sheet = workbook[sheet_name]
            except Exception as e:
                print(f"Error loading workbook for row {row_num}: {e}")
                continue

            data_map = {
                "TestCaseId": 1,
                "TestStatus": 2,
                "ActualStatus": 3,
                "OrderNo": 4,
                "AssignTo": 5,      # Karigar / Employee
                "AssignName": 6,    # Name from Select2
                "SmithDueDate": 7,  # Due Date
                "Branch": 8,        # Branch
                "Action": 9,        # Assign / Reject
                "Remark": 10
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # if not row_data["TestCaseId"]:
            #     continue 
            
            # Skip if disabled (using framework string test for 'yes')
            # status = str(row_data["TestStatus"]).strip().lower()
            # if status != "run":
            #     if status not in ["pass", "fail"]:
            #         continue # Ignore empty rows or non-execution rows

            print(f"\n{'='*80}")
            print(f"🧪 Running TC: {row_data['TestCaseId']} - Order: {row_data['OrderNo']}")
            print(f"{'='*80}")

            try:
                result = self.test_allotment_flow(row_data, row_num, sheet_name)
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
            except Exception as e:
                err_msg = str(e)
                print(f"❌ Test Failed Exception: {err_msg}")
                self._update_excel_status(row_num, "Fail", f"Exception: {err_msg[:60]}", sheet_name)

    def test_allotment_flow(self, row_data, row_num, sheet_name):
        driver = self.driver
        wait = self.wait
        current_field = "Start Flow"

        try:

            # 0. Filter Branch and Search
            branch = str(row_data.get("Branch", "")).strip()
            if branch and branch.lower() != "none" and branch != "":
                current_field = "Filter Branch"
                search_xpath = "//span[@class='select2-search select2-search--dropdown']/input"
                try:
                    # Select branch from dropdown
                    Function_Call.dropdown_select(self, "//span[@id='select2-branch_filter-container']", branch, search_xpath)
                    sleep(1)
                    # Click main search button above grid
                    Function_Call.click(self, "//button[@id='search_new_order_list']")
                    sleep(3)
                except Exception as e:
                    print(f"⚠️ Warning: Could not filter branch {branch}: {e}")

            order_no = str(row_data["OrderNo"]) if row_data.get("OrderNo") else None
            assign_to = str(row_data["AssignTo"]).strip().lower() if row_data.get("AssignTo") else ""
            assign_name = str(row_data["AssignName"]) if row_data.get("AssignName") else None
            smith_due_date = str(row_data.get("SmithDueDate", "")).strip()[:10] if row_data.get("SmithDueDate") else None # Get first 10 chars (YYYY-MM-DD)
            action = str(row_data["Action"]).strip().lower() if row_data.get("Action") else "assign"

            if not order_no:
                return ("Fail", "OrderNo is required")

            # 1. Search for the specific Order in DataTables
            current_field = "Search Order"
            search_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(order_no)
            sleep(2) # Wait for table filter

            # 2. Select the specific row checkbox
            current_field = "Select Order Checkbox"
            try:
                # Find the row containing the order number and click its checkbox label
                row_xpath = f"//table[@id='neworder_list']/tbody/tr[contains(., '{order_no}')]//input[@type='checkbox']"
                checkbox = wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                
                if not checkbox.is_selected():
                    # Checkbox might be obscured, use wrapper click if native fails
                    try:
                        checkbox.click()
                    except ElementClickInterceptedException:
                        # Find the parent label and click that
                        label_xpath = f"//table[@id='neworder_list']/tbody/tr[contains(., '{order_no}')]//label"
                        Function_Call.click(self, label_xpath)
                        
                sleep(1)
            except Exception as e:
                return ("Fail", f"Order {order_no} not found in list or cannot click row checkbox")


            # 3. Assign Role (Karigar or Employee)
            current_field = "Assign Role Radio"
            if assign_to == "employee":
                Function_Call.click(self, "//input[@name='order[assign_to]' and @value='2']")
                target_select_id = "employee_sel"
            else:
                Function_Call.click(self, "//input[@name='order[assign_to]' and @value='1']")
                target_select_id = "karigar_sel"
            
            sleep(1)

            # 4. Pick Name from Select2
            current_field = "Assign Name Select2"
            if assign_name:
                trigger_xpath = f"//select[@id='{target_select_id}']/following-sibling::span"
                search_xpath = "//span[@class='select2-search select2-search--dropdown']/input"
                Function_Call.dropdown_select(self, trigger_xpath, assign_name, search_xpath)
                sleep(1)

            # 4.5. Enter Smith Due Date (if provided)
            if smith_due_date and smith_due_date != "None":
                current_field = "Smith Due Date"
                try:
                    # Target the first text input inside the selected order's row
                    row_input_xpath = f"//table[@id='neworder_list']/tbody/tr[contains(., '{order_no}')]//input[@type='text' or contains(@class, 'datepicker')]"
                    due_date_el = wait.until(EC.presence_of_element_located((By.XPATH, row_input_xpath)))
                    driver.execute_script("arguments[0].value = arguments[1];", due_date_el, smith_due_date)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", due_date_el)
                    sleep(0.5)
                except Exception as e:
                    print(f"⚠️ Warning: Could not set Smith Due Date for {order_no}: {e}")

            # 5. Submit Action
            current_field = "Submit Button"
            if action == "reject":
                Function_Call.click(self, "//label[@id='reject']") # This label contains the radio
            else:
                Function_Call.click(self, "//label[@id='approve']") # Assign/Approve
            
            # 6. Capture Success Alert
            current_field = "Capture Success Toaster"
            try:
                success_xpath = "//div[contains(@class, 'alert-success')]"
                msg = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath))).text.strip()
                if "success" in msg.lower():
                    return ("Pass", f"Successfully Assigned: {order_no}")
                else:
                    return ("Fail", f"Unexpected alert message: {msg}")
            except TimeoutException:
                # Ensure we handle potential browser native alerts that sometimes preempt toaster
                try:
                    Function_Call.alert(self) # if a classic alert popped
                    return ("Pass", f"Assigned Alert Accepted: {order_no}")
                except:
                    return ("Fail", "No success message appeared after submission")


        except Exception as e:
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def _update_excel_status(self, row_num, status, message, sheet_name):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            
            color = "00B050" if status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=message).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
            print(f"📊 Extracted and updated Excel at row {row_num}")
        except Exception as e:
            print(f"⚠️ Excel save failed at row {row_num}: {e}")
