import pandas as pd
from openpyxl import load_workbook
import os
from Utils.Excel import ExcelUtils

file_path = ExcelUtils.file_path

if not os.path.exists(file_path):
    print(f"File {file_path} not found.")
    exit(1)

wb = load_workbook(file_path)

sheet_name = "RepairOrderStatus"

if sheet_name not in wb.sheetnames:
    ws = wb.create_sheet(sheet_name)
    headers = ["TestCaseId", "TestStatus", "ActualStatus", "OrderNo", "DateRange", "RepairType", "ExtraMetal", "Action", "Remark"]
    
    # Write Headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write Dummy Data
    dummy_data = ["TC001", "Run", "", "ATM25-RE-00013", "Today", "Customer", "Gold|22K|Chain|2.5|10|0|5000", "Complete", ""]
    for col_num, val in enumerate(dummy_data, 1):
        ws.cell(row=2, column=col_num, value=val)
    
    wb.save(file_path)
    print(f"Successfully added sheet '{sheet_name}' with dummy data.")
else:
    print(f"Sheet '{sheet_name}' already exists.")
