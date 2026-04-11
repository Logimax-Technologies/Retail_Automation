from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path

class StockIssue(unittest.TestCase):
    """
    Stock Issue / Receipt Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_stock_issue(self):
        """Main entry point for Stock Issue/ Receipt automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Stock Issue List
        try:
            if "admin_ret_stock_issue/stock_issue/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Stock Issue / Receipt')]")
                sleep(1)
                Function_Call.click(self, "(//span[contains(text(), 'Stock Issue')])[3]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(ExcelUtils.base_url + "index.php/admin_ret_stock_issue/stock_issue/list")
            sleep(2)

        sheet_name = "StockIssue"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3, "ProcessType": 4,
                "IssueFrom": 5, "SelEmp": 6, "IssueType": 7, "StockType": 8, "IssuedTo": 9,
                "CusName": 10, "IssueEmployee": 11, "Karigar": 12, "Metal": 13,
                "Section": 14, "TagCode": 15, "OldTagCode": 16, "RatePerGram": 17,
                "NT_Product": 18, "NT_Design": 19, "NT_SubDesign": 20, "NT_Pcs": 21,
                "NT_GrossWt": 22, "NT_NetWt": 23, "ReceiptIssueNo": 24,
                "ReceiptTagCode": 25, "ReceiptOldTagCode": 26, "Remark": 27
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # # Execution Guard
            # if str(row_data["TestStatus"]).strip().lower() == "done":
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']} ({row_data['ProcessType']})")
            print(f"{'='*80}")

            try:
                process_type = str(row_data["ProcessType"]).strip().lower()
                if process_type == "issue":
                    result = self.test_issue_flow(row_data, row_num, sheet_name)
                elif process_type == "receipt":
                    result = self.test_receipt_flow(row_data, row_num, sheet_name)
                else:
                    result = ("Fail", f"Unknown ProcessType: {process_type}")

                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_invoice = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_invoice)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def test_issue_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Issue Flow Start"
        try:
            Function_Call.alert(self)
            Function_Call.click(self, '//a[@id="add_Order"]')
            sleep(3)

            # 1. Type (Issue)
            current_field = "Type Selection"
            Function_Call.click(self, '//input[@id="type_issue"]')
            sleep(1)

            # 2. Issue From (Branch)
            if row_data.get("IssueFrom"):
                current_field = "Issue From"
                self.fc.dropdown_select2('//select[@id="branch_select"]/following-sibling::span', str(row_data["IssueFrom"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # 3. Select Employee
            if row_data.get("SelEmp"):
                current_field = "Sel Employee"
                self.fc.dropdown_select2('//select[@id="sel_emp"]/following-sibling::span', str(row_data["SelEmp"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # 4. Issue Type
            if row_data.get("IssueType"):
                current_field = "Issue Type"
                self.fc.dropdown_select2('//select[@id="issue_type"]/following-sibling::span', str(row_data["IssueType"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # 5. Stock Type
            if row_data.get("StockType"):
                current_field = "Stock Type"
                self.fc.select('//select[@id="stock_type"]', str(row_data["StockType"]))
                sleep(1)

            # 6. Select Metal
            if row_data.get("Metal"):
                current_field = "Select Metal"
                self.fc.dropdown_select2('//select[@id="metal"]/following-sibling::span', str(row_data["Metal"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # 7. Issue to
            if row_data.get("IssuedTo"):
                current_field = "Issue To"
                self.fc.select('//select[@id="issued_to"]', str(row_data["IssuedTo"]))
                sleep(2) # Wait for dynamic fields

                # 8. Recipient dynamic fields
                issued_to = str(row_data["IssuedTo"]).strip()
                if issued_to == "Customer": # Customer
                    current_field = "Customer lookup"
                    self.fc.fill_autocomplete_field("est_cus_name", str(row_data["CusName"]))
                    sleep(1)
                elif issued_to == "Employee": # Employee
                    current_field = "Issue Employee"
                    self.fc.dropdown_select2('//select[@id="issue_employee"]/following-sibling::span', str(row_data["IssueEmployee"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                elif issued_to == "karigar": # Karigar
                    current_field = "Karigar"
                    self.fc.dropdown_select2('//select[@id="karigar"]/following-sibling::span', str(row_data["Karigar"]), '//span[@class="select2-search select2-search--dropdown"]/input')

            stock_type = str(row_data["StockType"]).strip()
            if stock_type == "Taged": # Tagged
                if row_data.get("Section"):
                   self.fc.dropdown_select2('//select[@id="section_select"]/following-sibling::span', str(row_data["Section"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                
                current_field = "Tag Scan"
                wait.until(EC.presence_of_element_located((By.ID, "issue_tag_code"))).send_keys(str(row_data["TagCode"]))
                # Function_Call.fill_input(self, wait, (By.ID, "issue_tag_code"), str(row_data["TagCode"]), "TagCode", row_num, sheet_name)
                Function_Call.click(self, '//button[@id="issue_tag_search"]')
                sleep(3)
                
                current_field = "Verify Tag Loaded"
                wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="tagissue_item_detail"]/tbody/tr')))
            else: # Non-Tagged (Refactored for Multi-Item Match & Fill)
                current_field = "Non-Tag Grid Interaction"
                Function_Call.click(self, '//input[@id="search_non_tag"]')
                sleep(3) # Wait for grid to populate
                
                # Support multi-item via pipe separator |
                target_products = [p.strip() for p in str(row_data["NT_Product"]).split('|')]
                target_pcs = [p.strip() for p in str(row_data["NT_Pcs"]).split('|')]
                target_gwgt = [p.strip() for p in str(row_data["NT_GrossWt"]).split('|')]
                target_nwgt = [p.strip() for p in str(row_data["NT_NetWt"]).split('|')]

                # Iterate through grid rows to match and fill
                rows = driver.find_elements(By.XPATH, '//table[@id="nontagissue_item_detail"]/tbody/tr')
                for row in rows:
                    try:
                        # Product Name is typically in the 3rd column
                        product_name = row.find_element(By.XPATH, './td[3]').text.strip()
                        if product_name in target_products:
                            idx = target_products.index(product_name)
                            
                            # 1. Check the checkbox
                            checkbox = row.find_element(By.CLASS_NAME, 'nt_item_sel')
                            if not checkbox.is_selected():
                                checkbox.click()
                            
                            # 2. Fill Pcs
                            pcs_input = row.find_element(By.CLASS_NAME, 'ntpcs')
                            pcs_input.clear()
                            pcs_input.send_keys(target_pcs[idx])
                            
                            # 3. Fill GWgt
                            gwt_input = row.find_element(By.CLASS_NAME, 'nt_gross_wt')
                            gwt_input.clear()
                            gwt_input.send_keys(target_gwgt[idx])
                            
                            # 4. Fill NWgt (May be auto-calculated or manual depend on JS)
                            nwt_input = row.find_element(By.CLASS_NAME, 'nt_net_wt')
                            if nwt_input.is_enabled():
                                nwt_input.clear()
                                nwt_input.send_keys(target_nwgt[idx])
                            
                            print(f"✅ Updated Grid Row for Product: {product_name}")
                    except Exception as e:
                        print(f"⚠️ Skipping row due to error: {e}")
                sleep(1)

            if row_data.get("RatePerGram"):
                current_field = "Rate per Gram"
               #Function_Call.fill_input(self, wait, (By.ID, "rate_per_gram"), str(row_data["RatePerGram"]), "RatePerGram", row_num, sheet_name)
                wait.until(EC.presence_of_element_located((By.ID, "rate_per_gram"))).send_keys(str(row_data["RatePerGram"]))
                sleep(1)
            if row_data.get("Remark"):
                Function_Call.fill_input(self, wait, (By.ID, "remark"), str(row_data["Remark"]), "Remark", row_num, Sheet_name=sheet_name)

            current_field = "Submit Issue"
            self._take_screenshot(f"BeforeIssue_TC{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="stock_issue_submit"]')
            sleep(5)

            # 9. Post-Save: ID Extraction from Print Tab
            current_field = "Capture ID and Invoice"
            main_handle = driver.current_window_handle
            wait.until(EC.number_of_windows_to_be(2))
            for handle in driver.window_handles:
                if handle != main_handle:
                    driver.switch_to.window(handle)
                    break
            
            print(f"✅ Switched to Print Tab: {driver.current_url}")
            current_url = driver.current_url
            issue_id = current_url.split('/')[-1]
            
            
            
            print(f"✅ Captured Issue ID: {issue_id}")
            driver.close()
            driver.switch_to.window(main_handle)

            # 10. List Verification
            return self.test_list_verification_flow(row_data, issue_id)

        except Exception as e:
            self._take_screenshot(f"ErrorIssue_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_receipt_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Receipt Flow Start"
        try:
            Function_Call.alert(self)
            Function_Call.click(self, '//a[@id="add_Order"]')
            sleep(3)

            # 1. Type (Receipt)
            current_field = "Type Selection"
            Function_Call.click(self, '//input[@id="type_receipt"]')
            sleep(1)

            # 2. Issue From
            if row_data.get("IssueFrom"):
                current_field = "Issue From"
                self.fc.dropdown_select2('//select[@id="branch_select"]/following-sibling::span', str(row_data["IssueFrom"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # 3. Select Employee
            if row_data.get("SelEmp"):
                current_field = "Sel Employee"
                self.fc.dropdown_select2('//select[@id="sel_emp"]/following-sibling::span', str(row_data["SelEmp"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # 4. Stock Type
            if row_data.get("StockType"):
                current_field = "Stock Type"
                Function_Call.select(self,'//select[@id="stock_type"]', str(row_data["StockType"]))
                sleep(1)

            # 5. Select Issue No
            if row_data.get("ReceiptIssueNo"):
                current_field = "Select Issue No"
                self.fc.dropdown_select2('//select[@id="select_issue_no"]/following-sibling::span', str(row_data["ReceiptIssueNo"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(2)

            stock_type = str(row_data["StockType"]).strip()
            ReceiptIssueNO=str(row_data["ReceiptIssueNo"]).strip()
            if ReceiptIssueNO == "" or None:
                if stock_type == "Taged": # Tagged
                    Function_Call.fill_input(self, wait, (By.ID, "receipt_tag_code"), str(row_data["ReceiptTagCode"]), "ReceiptTagCode", row_num, Sheet_name=sheet_name)
                    Function_Call.click(self, '//button[@id="receipt_tag_search"]')
                    sleep(3)
                    wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="tag_receipt_item_detail"]/tbody/tr')))
                else: # Non-Tagged
                    current_field = "Non-Tagged Receipt Grid"
                    # Select checkboxes for items loaded by Issue No
                    items = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//table[@id="nontag_receipt_item_detail"]//input[@type="checkbox"]')))
                    for item in items:
                        if not item.is_selected():
                            item.click()
                    sleep(1)

           
            current_field = "Submit Receipt"
            self._take_screenshot(f"BeforeReceipt_TC{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="stock_issue_submit"]')
            
            # Success is defined as reaching the list page
            try:
                wait.until(EC.url_contains("stock_issue/list"))
                return ("Pass", "Receipt saved successfully")
            except:
                return ("Fail", "Failed to reach list page after submit")

          

        except Exception as e:
            self._take_screenshot(f"ErrorReceipt_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_list_verification_flow(self, row_data, issue_id):
        driver, wait = self.driver, self.wait
        current_field = "List Verification"
        try:
            # Navigate to list if not there
            if "stock_issue/list" not in driver.current_url:
                driver.get(ExcelUtils.base_url + "index.php/admin_ret_stock_issue/stock_issue/list")
                sleep(2)
            
            # Filter by "Issued" status
            current_field = "Status Filter"
            if row_data["ProcessType"]=="Issue":
                Status="Issued"
            else:
                Status="Received"               
            Function_Call.select(self,'//select[@id="issue_status"]', Status)
            Function_Call.click(self, '//button[contains(text(), "Search")]')
            sleep(3)

            # Search for ID in dataTable search box (only if provided)
            if issue_id:
                current_field = "DT Search"
                search_box = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="DataTables_Table_0_filter"]//input | //input[@type="search"]')))
                search_box.clear()
                search_box.send_keys(issue_id)
                sleep(2)
            else:
                print("ℹ️ No specific ID provided, verifying the latest entry in the list.")

            # Verify row presence and correct Invoice No
            current_field = "Row Verification"
            row_text = driver.find_element(By.XPATH, '//table[contains(@class, "dataTable")]/tbody/tr[1]').text
            if issue_id in row_text:
                issue_no = driver.find_element(By.XPATH, '//table[contains(@class, "dataTable")]/tbody/tr[1]/td[2]').text
                return ("Pass", f"Verified in list. ID: {issue_id}, Invoice: {issue_no}", issue_no)
            else:    
                return ("Fail", f"Verification failed. Grid row text: {row_text}")

        except Exception as e:
            return ("Fail", f"List Error in {current_field}: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_invoice=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            if captured_invoice:
                # Link to next row if it's a receipt
                next_row = row_num + 1
                try:
                    next_process = str(sheet.cell(row=next_row, column=4).value).strip().lower()
                    if next_process == "receipt":
                        sheet.cell(row=next_row, column=24, value=captured_invoice).font = Font(bold=True)
                except:
                    pass

            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass

    def _cancel_form(self):
        try:
            Function_Call.click(self, '//button[contains(@class, "btn-cancel")]')
        except:
            pass
