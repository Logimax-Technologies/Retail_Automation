from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils


class StoneRateSettings:
    def __init__(self, driver):
        self.driver = driver
        self.fc = Function_Call(driver)

    def test_stone_rate_settings(self):
        FILE_PATH = ExcelUtils.file_path
        BASE_URL = ExcelUtils.BASE_URL
        driver = self.driver
        wait = WebDriverWait(driver, 15)
        
        # 1. Navigation
        try:
            if "admin_ret_catalog/ret_stone_rate_settings" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                self.fc.click(self, "//span[contains(text(), 'Retail catalog')]")
                sleep(1)
                self.fc.click(self, "//span[contains(text(), 'Stone Rate Settings')]")
                sleep(2)
            
            # Click Add button in list view
            try:
                add_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, '/add')]")))
                add_btn.click()
                sleep(2)
            except:
                # If already on add page or button not found, fallback to direct URL
                driver.get(driver.current_url.split('/list')[0] + "/add")
                sleep(2)

        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_catalog/ret_stone_rate_settings/add")
            sleep(3)

        # 2. Data Processing
        excel_path = FILE_PATH
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb["StoneRateSettings"]
        
        results = []
        for row in range(2, sheet.max_row + 1):
            row_data = {
                'TestCaseId': sheet.cell(row, 1).value,
                'TestStatus': sheet.cell(row, 2).value,
                'Tab': sheet.cell(row, 4).value,
                'Branch': sheet.cell(row, 5).value,
                'StoneType': sheet.cell(row, 6).value,
                'StoneName': sheet.cell(row, 7).value,
                'Product': sheet.cell(row, 8).value,
                'Design': sheet.cell(row, 9).value,
                'SubDesign': sheet.cell(row, 10).value,
                'Quality': sheet.cell(row, 11).value,
                'UOM': sheet.cell(row, 12).value,
                'CalcType': sheet.cell(row, 13).value,
                'FromCent': sheet.cell(row, 14).value,
                'ToCent': sheet.cell(row, 15).value,
                'MinRate': sheet.cell(row, 16).value,
                'MaxRate': sheet.cell(row, 17).value,
                'ExpectedMsg': sheet.cell(row, 18).value
            }

            if row_data['TestStatus'] != "Run":
                continue

            # 2. Data Splitting (Support for Range Flow via pipe separation)
            data_fields = [
                'Branch', 'StoneType', 'StoneName', 'Product', 'Design', 
                'SubDesign', 'Quality', 'UOM', 'CalcType', 'FromCent', 
                'ToCent', 'MinRate', 'MaxRate'
            ]
            
            # Split each field into a list by |
            split_data = {}
            max_rows = 1
            for field in data_fields:
                val = str(row_data.get(field) or "").strip()
                items = [i.strip() for i in val.split('|')]
                split_data[field] = items
                max_rows = max(max_rows, len(items))

            print(f"🚀 Executing {row_data['TestCaseId']} on {row_data['Tab']} ({max_rows} rows)...")
            
            try:
                # Tab Switching
                if row_data['Tab'] == "Loose Products":
                    self.fc.click(self, "//a[@href='#tab_2' and contains(text(), 'Loose Products')]")
                    sleep(1)
                    add_row_xpath = "//span[@class='add_stoneproduct_info']"
                    save_btn_xpath = "//button[@id='add_stnproduct_rate_settings']"
                    table_id = "total_stnproduct_rate_items"
                else:
                    self.fc.click(self, "//a[@href='#tab_1' and contains(text(), 'Loose Stone')]")
                    sleep(1)
                    add_row_xpath = "//span[@class='add_stonerate_info']"
                    save_btn_xpath = "//button[@id='add_stn_rate_settings']"
                    table_id = "total_stone_rate_items"
                
                sleep(1)
                
                # Loop through rows for Range Flow
                for i in range(max_rows):
                    # Click Add Row (only if not already provided or for additional rows)
                    # The UI usually has one empty row if fresh, but better to click + for each
                    self.fc.click(self, add_row_xpath)
                    sleep(1)
                    
                    # Identify newly added row (the last row in the table)
                    row_xpath = f"(//table[@id='{table_id}']/tbody/tr)[last()]"
                    
                    # Helper function to get value by index (fallback to first item if list is shorter)
                    def get_val(f, idx):
                        lst = split_data.get(f, [""])
                        return lst[idx] if idx < len(lst) else lst[0]

                    # 1. Branch
                    branch_val = get_val('Branch', i)
                    if branch_val:
                        self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[id_branch]')]/following-sibling::span", branch_val)
                    
                    # 2. Tab Specific Fields
                    if row_data['Tab'] == "Loose Products":
                        prod_val = get_val('Product', i)
                        if prod_val:
                            self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[id_product]')]/following-sibling::span", prod_val)
                        
                        des_val = get_val('Design', i)
                        if des_val:
                            self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[id_design]')]/following-sibling::span", des_val)
                        
                        sub_val = get_val('SubDesign', i)
                        if sub_val:
                            self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[id_sub_design]')]/following-sibling::span", sub_val)
                    else:
                        stype_val = get_val('StoneType', i)
                        if stype_val:
                            self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[stn_type]')]/following-sibling::span", stype_val)
                        
                        sname_val = get_val('StoneName', i)
                        if sname_val:
                            self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[stn_name]')]/following-sibling::span", sname_val)
                    
                    # 3. Quality, UOM, CalcType
                    qual_val = get_val('Quality', i)
                    if qual_val:
                        self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[stn_quality]') or contains(@name, '[quality_code]')]/following-sibling::span", qual_val)
                    
                    uom_val = get_val('UOM', i)
                    if uom_val:
                        self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[uom_id]')]/following-sibling::span", uom_val)
                    
                    calc_val = get_val('CalcType', i)
                    if calc_val:
                        self.fc.dropdown_select2(self, f"{row_xpath}//select[contains(@name, '[stn_calc_type]')]/following-sibling::span", calc_val)
                    
                    # 4. Inputs (handle disabled/readonly)
                    self.fc.fill_input(self, f"{row_xpath}//input[contains(@name, '[from_wt]')]", get_val('FromCent', i))
                    self.fc.fill_input(self, f"{row_xpath}//input[contains(@name, '[to_wt]')]", get_val('ToCent', i))
                    self.fc.fill_input(self, f"{row_xpath}//input[contains(@name, '[min_rate]')]", get_val('MinRate', i))
                    self.fc.fill_input(self, f"{row_xpath}//input[contains(@name, '[max_rate]')]", get_val('MaxRate', i))
                
                # 7. Final Save after all rows added
                self.fc.click(self, save_btn_xpath)
                sleep(1)
                
                # 8. Verification
                alert_msg = self.fc.alert2(self)
                if alert_msg and (row_data['ExpectedMsg'] or "success" in alert_msg.lower()):
                    status = "Pass"
                else:
                    status = "Fail"
                
                # Update excel and collect result
                self.fc._update_excel_status(row, "Pass", alert_msg or "Saved", Sheet_name="StoneRateSettings")
                results.append((row_data['TestCaseId'], status, alert_msg))
                
                # If passed, we usually navigate back to add for next row if needed, 
                # but the system might redirect to list. 
                if "list" in driver.current_url:
                    driver.get(driver.current_url.split('/list')[0] + "/add")
                    sleep(2)

            except Exception as e:
                print(f"❌ Error in {row_data['TestCaseId']}: {e}")
                self.fc._update_excel_status(row, "Fail", str(e), Sheet_name="StoneRateSettings")
                results.append((row_data['TestCaseId'], "Fail", str(e)))
                driver.get(BASE_URL + "index.php/admin_ret_catalog/ret_stone_rate_settings/add")
                sleep(2)

        wb.close()
        return results
