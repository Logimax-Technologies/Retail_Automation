from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from time import sleep
import math
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call

FILE_PATH = ExcelUtils.file_path

class Stone:
    """Helper class for handling Less Weight / Stone details."""
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)
        self.fc = Function_Call(driver)

    def test_tagStone(self, sheet_name, test_case_id):
        """Main method to fill stone details for a specific test case."""
        wait = self.wait
        sleep(2)
        
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            count = ExcelUtils.Test_case_id_count(FILE_PATH, sheet_name, test_case_id)
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
        except Exception as e:
            print(f"❌ Failed to load stone data for {sheet_name}: {e}")
            return None

        try:
            row_idx = 1
            processed_count = 0
            all_stones_data = []

            for row_num in range(2, valid_rows + 1):
                current_id = sheet.cell(row=row_num, column=1).value
                if str(current_id).strip() == str(test_case_id).strip():
                    # Adjusted mapping for Customer_LWT sheet based on provided screenshot
                    data_map = {
                        "Type": 2, "Name": 3, "Pcs": 4, 
                        "Wt": 5, "Wt Type": 6, "Cal.Type": 7, "Rate": 8, "Amount": 9
                    }
                    row_data = {k: sheet.cell(row=row_num, column=v).value for k, v in data_map.items()}
                    
                    print(f"💎 Adding Actual Less Weight: {row_data['Name']} (Row {row_idx})")
                    
                    # Check LWT checkbox
                    lwt_xpath = f"(//input[@name='est_stones_item[show_in_lwt][]'])[{row_idx}]"
                    try:
                        lwt_checkbox = wait.until(EC.presence_of_element_located((By.XPATH, lwt_xpath)))
                        if not lwt_checkbox.is_selected():
                            self.fc.click(lwt_xpath)
                    except Exception as e:
                        print(f"⚠️ LWT Checkbox click for row {row_idx} failed: {e}")

                    # Fill Dropdowns
                    Stone._select_dropdown(self, f"(//select[@name='est_stones_item[stones_type][]'])[{row_idx}]", row_data["Type"])
                    Stone._select_dropdown(self, f"(//select[@name='est_stones_item[stone_id][]'])[{row_idx}]", row_data["Name"])

                    # Fill Inputs
                    Stone._fill_input(self, f"(//input[@name='est_stones_item[stone_pcs][]'])[{row_idx}]", row_data["Pcs"])
                    Stone._fill_input(self, f"(//input[@name='est_stones_item[stone_wt][]'])[{row_idx}]",  row_data["Wt"])
                    Stone._select_dropdown(self, f"(//select[@name='est_stones_item[stone_uom_id][]'])[{row_idx}]", row_data["Wt Type"])

                    # Cal.Type Radio
                    stone_row_param = row_idx - 1
                    cal_type_val = 1 if str(row_data["Cal.Type"]).strip().lower() == "wt" else 2
                    cal_xpath = f"(//input[@name='est_stones_item[cal_type][{stone_row_param}]' and @value='{cal_type_val}'])"
                    self.fc.click(cal_xpath)

                    # Rate
                    Stone._fill_input(self, f"(//input[@name='est_stones_item[stone_rate][]'])[{row_idx}]", row_data["Rate"])
                    
                    # Get calculation from UI (since Excel might be empty)
                    sleep(1) # Wait for UI calculation
                    amt_xpath = f'(//input[@name="est_stones_item[stone_price][]"])[{row_idx}]'
                    table_amt = self.fc.get_value(amt_xpath)
                    row_data["Amount"] = table_amt  # Update with actual UI value
                    
                    Stone._validate_calc(self, row_data, table_amt)

                    processed_count += 1
                    all_stones_data.append(row_data)

                    # Add more rows if needed
                    if processed_count < count:
                        wait.until(EC.element_to_be_clickable((By.ID, "create_stone_item_details"))).click()
                        row_idx += 1
                        sleep(1)
        except Exception as e:
            print(f"❌ Error during stone processing: {e}")
        # Calculate totals (Wt, Amt) manually with unit conversion
        wt_total = 0.0
        amt_total = 0.0
        for stone in all_stones_data:
            try:
                wt = float(str(stone.get("Wt") or 0).replace(",", "").strip())
                if str(stone.get("Wt Type")).strip().lower() == "carat":
                    wt = wt * 0.2  # Convert carat to grams
                wt_total += wt
                
                # Handle potential commas in amount string
                amt_str = str(stone.get("Amount") or 0).replace(",", "").strip()
                amt_total += float(amt_str)
            except Exception as e:
                print(f"⚠️ Calculation error for row: {e}")
        
        wt_total = round(wt_total, 3)
        amt_total = round(amt_total, 2)

        # Save and return
        if processed_count > 0:
            wait.until(EC.element_to_be_clickable((By.ID, "update_stone_details"))).click()
            print(f"✅ Successfully added {processed_count} stone(s) | Total Wt: {wt_total}g, Total Amt: {amt_total}")
            
            return True, wt_total, amt_total, all_stones_data
        
        workbook.close()
        return None

    def _select_dropdown(self, locator, value):
        if value is None: return
        try:
            el = self.wait.until(EC.presence_of_element_located((By.XPATH, locator)))
            Select(el).select_by_visible_text(str(value))
        except Exception as e:
            print(f"⚠️ Dropdown failed: {locator} | {e}")

    def _fill_input(self, locator, value):
        if value is None: return
        try:
            el = self.wait.until(EC.visibility_of_element_located((By.XPATH, locator)))
            el.clear()
            el.send_keys(str(value))
        except Exception as e:
            print(f"⚠️ Input failed: {locator} | {e}")

    def _validate_calc(self, data, table_amt):
        try:
            rate = float(data["Rate"] or 0)
            pcs = float(data["Pcs"] or 0)
            wt = float(data["Wt"] or 0)
            cal_type = str(data["Cal.Type"]).strip().lower()
            
            expected = math.ceil(wt * rate) if cal_type == "wt" else math.ceil(pcs * rate)
            found = math.ceil(float(table_amt or 0))
            
            if abs(expected - found) > 1:
                print(f"❌ Stone Calc mismatch: Expected {expected}, Found {found}")
        except:
            pass

    def _get_table_total(self, name):
        """Sum up values from all inputs with given name."""
        sleep(1)
        els = self.driver.find_elements(By.NAME, name)
        total = 0.0
        for el in els:
            try:
                val = el.get_attribute("value")
                if val: total += float(val)
            except: pass
        return round(total, 3)