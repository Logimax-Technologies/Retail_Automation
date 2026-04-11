import openpyxl
from Utils.Excel import ExcelUtils

file_path = ExcelUtils.file_path

try:
    wb = openpyxl.load_workbook(file_path)
    if "RepairOrderStatus" in wb.sheetnames:
        ws = wb["RepairOrderStatus"]
        
        # New headers with Branch added
        headers = ["TestCaseId", "TestStatus", "ActualStatus", "OrderNo", "Branch", "DateRange", "RepairType", "ExtraMetal", "CompletedWeight", "Amount", "Action", "Remark"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            
        # Dummy data matching: ItemType|Section|Category|Purity|Product|Design|SubDesign|Pcs|GWt|VA%|MCType|MC|ServiceCharge
        dummy_metal = "Repair|GOLD RING|22KT GOLD ORNAMENTS|91.6000|GOLD BANGLES|CASTING|CAST BANGLE|1|4|1|Per Grams|200|500"
        
        # Test Data mapping to new 12 cols
        dummy_data = ["TC001", "Run", "", "ATM25-RE-00013", "Head Office", "Today", "Customer", dummy_metal, "11.000", "1000", "Complete", ""]
        
        for col_num, val in enumerate(dummy_data, 1):
            ws.cell(row=2, column=col_num, value=val)
            
        wb.save(file_path)
        print("Excel dummy data updated with new Branch column (12 columns total).")
    else:
        print("Sheet RepairOrderStatus not found.")
except Exception as e:
    print(f"Error updating excel: {e}")
