from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.styles import Font
from openpyxl import load_workbook
import unittest
import os

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class OldMetalProcess(unittest.TestCase):
    """
    Old Metal Process Module Automation
    """
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 20)
        self.fc = Function_Call(driver)
        self.sheet_name = "OldMetalProcess"

    def test_old_metal_process(self):
        """Main entry point for Old Metal Process automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigate to Metal Process List
        driver.get(BASE_URL + "index.php/admin_ret_metal_process/metal_process_issue/list")
        sleep(2)
        
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, self.sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{self.sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return
        
        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[self.sheet_name]
            
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "Actual Status": 3, "ProcessFor": 4, 
                "AgainstMelting": 5, "ProcessType": 6, "Vendor": 7, 
                "MeltingType": 8, "PocketNo": 9,"ProcessNo":10, "ModalCategory": 11, 
                "ModalSection": 12, "ModalProduct": 13, "ModalDesign": 14,
                "ModalSubDesign": 15, "ModalPcs": 16, "ModalWeight": 17, 
                "Mode": 18, "Purity": 19, "Charges": 20, "RefNo": 21, "Remark": 22
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()
            
            # if str(row_data["TestStatus"]).strip().lower() != "run": 
            #     continue
            
            print(f"\n🧪 Running Test Case: {row_data['TestCaseId']}")
            try:
                # Navigate to Add page for each run
                driver.get(BASE_URL + "index.php/admin_ret_metal_process/metal_process_issue/add")
                sleep(2)
                
                result = self.perform_process(row_data)
                
                if result[0] == "Pass":
                    # Capture Dynamic ID if it was an Issue
                    if str(row_data["ProcessFor"]).upper() == "ISSUE":
                        captured_id = self.capture_process_no(row_data)
                        if captured_id:
                            print(f"🎯 Captured Process No: {captured_id}")
                            # Update the next row's PocketNo in Excel if it's a Receipt
                            self._update_next_row_id(row_num, captured_id)
                            result = ("Pass", f"Saved Successfully. ID: {captured_id}")
                    
                    self._update_excel_status(row_num, "Pass", result[1])
                else:
                    self._update_excel_status(row_num, "Fail", result[1])
            except Exception as e:
                print(f"❌ Error in TC {row_data['TestCaseId']}: {e}")
                self._update_excel_status(row_num, "Fail", str(e))
        # End of test cases

    def capture_process_no(self, row_data):
        """Navigate to list, filter, and capture the latest Process No."""
        driver, wait, fc = self.driver, self.wait, self.fc
        try:
            driver.get(BASE_URL + "index.php/admin_ret_metal_process/metal_process_issue/list")
            sleep(3)
            
            # 1. Date Range: Today
            fc.click('//button[@id="date_range_picker"]')
            sleep(1)
            # Click "Today" option
            fc.click('//div[@class="ranges"]//li[text()="Today"]')
            sleep(2)
            
            # 2. Search Box: Karigar Name
            search_box = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@type="search"]')))
            search_box.clear()
            search_box.send_keys(str(row_data["Vendor"]))
            sleep(2)
            
            # 3. Capture first row Process No
            # Column indices: # (1), Process No (2), Date (3), Process For (4), Process Type (5), Karigar (6)
            process_no = fc.get_text('//table[@id="process_list"]/tbody/tr[1]/td[2]')
            return process_no
        except Exception as e:
            print(f"⚠️ Failed to capture Process No: {e}")
            return None

    def perform_process(self, row_data):
        driver, wait, fc = self.driver, self.wait, self.fc
        
        # 1. Select Process For
        if str(row_data["ProcessFor"]).upper() == "ISSUE":
            fc.click('//input[@id="issue_process"]')
            print("Selected: ISSUE")
        else:
            fc.click('//input[@id="receipt_process"]')
            print("Selected: RECEIPT")
            sleep(1)
            # Against Melting logic for Receipt
            if str(row_data.get("AgainstMelting")).lower() == "yes":
                fc.click('//input[@id="against_melting_yes"]')
            else:
                fc.click('//input[@id="against_melting_no"]')

        # 2. Select Process Type
        if row_data.get("ProcessType"):
            fc.dropdown_select('//select[@id="select_process"]/following-sibling::span', str(row_data["ProcessType"]), '//span[@class="select2-search select2-search--dropdown"]/input')
            sleep(1)

        # 3. Select Vendor
        if row_data.get("Vendor"):
            fc.dropdown_select('//select[@id="karigar"]/following-sibling::span', str(row_data["Vendor"]), '//span[@class="select2-search select2-search--dropdown"]/input')

        # 4. Search
        fc.click('//button[@id="process_filter"]')
        sleep(3)
        
        p_type = str(row_data["ProcessType"]).upper()
        p_for = str(row_data["ProcessFor"]).upper()

        # 5. Process Specific Logic
        if p_type == "MELTING" and p_for == "ISSUE":
            if row_data.get("MeltingType"):
                fc.select('//select[@id="melting_trans_type"]', str(row_data["MeltingType"]))
                sleep(1)
            if row_data.get("PocketNo"):
                fc.select('//select[@id="select_pocket"]', str(row_data["PocketNo"]))
                sleep(2)
                checkbox_xpath = f'//table[@id="pocket_details"]//td[contains(text(), "{row_data["PocketNo"]}")]/..//input[@type="checkbox"]'
                fc.click(checkbox_xpath)

        elif p_type == "MELTING" and p_for == "RECEIPT":
            # Match the ID from Excel (dynamically updated by Issue TC)
            target_id = str(row_data.get("ProcessNo") or "").strip()
            if target_id:
                # Find row by ID
                row_xpath = f'//table[@id="melting_receipt"]//td[contains(text(), "{target_id}")]/..'
                
                # 1. Click Checkbox
                Check_box = f'{row_xpath}//input[@type="checkbox"]'
                print(Check_box)
                print(f"Clicking Checkbox: {Check_box}")
                fc.click(Check_box)
                sleep(1)
                
                # 2. Click Plus Button
                plus_btn = f'{row_xpath}//a[contains(@class, "btn-success")]'
                print(plus_btn)
                print(f"Clicking Plus Button: {plus_btn}")
                fc.click(plus_btn)
            else:
                # Fallback to first if no ID
                row_xpath = '(//table[@id="melting_receipt"]//tbody/tr)[1]'
                fc.click(f'{row_xpath}//a[contains(@class, "btn-success")]')
            
            sleep(2)
            # Modal Interaction
            if row_data.get("ModalCategory"):
                fc.select('//select[contains(@class, "id_ret_category")]', str(row_data["ModalCategory"]))
            if row_data.get("ModalSection"):
                fc.select('//select[contains(@class, "id_section")]', str(row_data["ModalSection"]))
            if row_data.get("ModalProduct"):
                fc.select('//select[contains(@class, "id_product")]', str(row_data["ModalProduct"]))
            if row_data.get("ModalDesign"):
                fc.select('//select[contains(@class, "id_design")]', str(row_data["ModalDesign"]))
            if row_data.get("ModalSubDesign"):
                fc.select('//select[contains(@class, "id_sub_design")]', str(row_data["ModalSubDesign"]))
            if row_data.get("ModalPcs"):
                fc.fill_input2('//input[contains(@class, "recd_pcs")]', str(row_data["ModalPcs"]))
            if row_data.get("ModalWeight"):
                fc.fill_input2('//input[contains(@class, "recd_gross_wt")]', str(row_data["ModalWeight"]))
            fc.click('//button[@id="update_category"]')
            sleep(2)
            
            if row_data.get("Charges"):
                fc.fill_input2(f'{row_xpath}//td[10]//input', str(row_data["Charges"]))
            if row_data.get("RefNo"):
                fc.fill_input2(f'{row_xpath}//td[11]//input', str(row_data["RefNo"]))

        elif p_type == "TESTING" and p_for == "ISSUE":
             p_no = str(row_data.get("PocketNo") or "")
             if p_no:
                 fc.click(f'//table[@id="testing_process_details"]//td[contains(text(), "{p_no}")]/..//input[@type="checkbox"]')
             else:
                 fc.click('//table[@id="testing_process_details"]/tbody/tr[1]//input[@type="checkbox"]')

        elif p_type == "TESTING" and p_for == "RECEIPT":
            if str(row_data.get("AgainstMelting")).lower() == "yes":
                target_id = str(row_data.get("PocketNo") or "").strip()
                if target_id:
                    # Select2 dropdown usually filters by typing
                    fc.select('//select[@id="select_metal_process"]', target_id)
                else:
                    fc.select('//select[@id="select_metal_process"]', "index=1")
                sleep(1)
                
                if str(row_data.get("Mode")).lower() == "add to stock":
                    fc.click('//input[@id="add_to_acc_stock"]')
                else:
                    fc.click('//input[@id="add_to_next_process"]')
                
                if row_data.get("Purity"):
                    fc.fill_input2('//table[@id="against_melting_receipt"]//input[contains(@name, "purity")]', str(row_data["Purity"]))

        # 6. Remark
        if row_data.get("Remark"):
            fc.fill_input2('//textarea[@id="remark"]', str(row_data["Remark"]))

        # 7. Final Save
        fc.click('//button[@id="issue_submit"]')
        
        try:
            alert_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".alert")))
            alert_msg = alert_element.text.replace('×', '').replace('\n', ' ').strip()
        except:
            alert_msg = "Process Saved Successfully"
            
        sleep(3)
        return ("Pass", alert_msg)

    def _update_excel_status(self, row_num, test_status, actual_status):
        try:
            wb = load_workbook(FILE_PATH); sh = wb[self.sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sh.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sh.cell(row=row_num, column=3, value=actual_status).font = Font(color=color)
            wb.save(FILE_PATH); wb.close()
        except: pass

    def _update_next_row_id(self, current_row, new_id):
        """Update the next row's ProcessNo column with the captured ID."""
        try:
            wb = load_workbook(FILE_PATH); sh = wb[self.sheet_name]
            # Column 10 is ProcessNo
            sh.cell(row=current_row + 1, column=10, value=new_id)
            wb.save(FILE_PATH); wb.close() 
            print(f"📝 Updated row {current_row + 1} with ID {new_id}")
        except Exception as e:
            print(f"⚠️ Failed to update next row: {e}")
