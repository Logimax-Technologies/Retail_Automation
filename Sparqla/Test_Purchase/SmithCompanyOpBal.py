from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class SmithCompanyOpBal(unittest.TestCase):
    """
    Smith / Company Opening Balance Module Automation
    Handles entering initial stock/balance for Smith/Karigars.
    Focuses exclusively on Smith Stock flow.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_smith_company_op_bal(self):
        """Main entry point for Smith Company Op Bal automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Add Page
        try:
            if "admin_ret_purchase/smith_cmpy_op_bal/add" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Purchase')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Smith/Company Opening Balance')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/smith_cmpy_op_bal/list")
            sleep(2)

        # Read Excel data
        sheet_name = "SmithCompanyOpBal"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to reach Excel or sheet: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Mapping based on SmithCompanyOpBalPrompt.md
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,"StockType": 4,
                "SmithType": 5, "BalType": 6, "MetalType": 7,
                "KarigarName": 8, "MetalName": 9, "CategoryName": 10,
                "ProductName": 11, "GrossWgt": 12,
                "NetWgt": 13, "DiaWgt": 14, "PureWgt": 15,
                "WtType": 16, "Amount": 17, "AmtType": 18,
                 "Ref No": 19,"Remark": 20
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            if str(row_data.get("TestStatus")).strip().lower() == "skip":
                print(f"⏭️ Skipping Test Case: {row_data['TestCaseId']}")
                continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                self.driver.refresh()
                sleep(1)
                Function_Call.click(self, "//a[@id='add_Order']")
                sleep(2)
                result = self.execute_flow(row_data, row_num, sheet_name)
                
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_id = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Initial Setup"
        try:
            Function_Call.alert(self)
            
            # --- 1. O/P Balance For (Smith Stock) ---
            # Click only if not already selected
            smith_stock_radio = driver.find_element(By.ID, "smth_stock")
            if not smith_stock_radio.is_selected():
                print("smith_stock_radio is not selected")
                Function_Call.click(self, "//input[@id='smth_stock']")
                sleep(1)
            else:
                print("smith_stock_radio is selected")

            # --- 2. Smith Type ---
            if row_data.get("SmithType"):
                current_field = "Smith Type"
                val = str(row_data["SmithType"]).strip().lower()
                type_map = {"supplier": "1", "smith": "2", "approval supplier": "3", "stone supplier": "4"}
                final_val = type_map.get(val, val)
                Function_Call.click(self, f"//input[@name='smth_cmpy_stk[smith_type]' and @value='{final_val}']")
                sleep(1)

            # --- 3. Balance Type ---
            if row_data.get("BalType"):
                current_field = "Balance Type"
                val = str(row_data["BalType"]).strip().lower()
                bal_map = {"metal": "1", "diamond/stone": "3", "diamond": "3", "stone": "3"}
                final_bal = bal_map.get(val, val)
                Function_Call.click(self, f"//input[@name='smth_cmpy_stk[bal_type]' and @value='{final_bal}']")
                sleep(1)

            # --- 4. Metal Type ---
            if str(row_data.get("BalType")) == "1" or (row_data.get("BalType") and str(final_bal) == "1"): # Metal
                current_field = "Metal Type"
                val = str(row_data.get("MetalType", "1")).strip().lower()
                metal_map = {"oranment": "1", "old metal": "2"}
                final_metal = metal_map.get(val, val)
                Function_Call.click(self, f"//input[@name='smth_cmpy_stk[metal_type]' and @value='{final_metal}']")
                sleep(1)

            # --- 5. Selection Dropdowns (Select2) ---
            # Karigar
            if row_data.get("KarigarName"):
                current_field = "Karigar"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_karigar-container']", 
                                             str(row_data["KarigarName"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # Metal
            if row_data.get("MetalName"):
                current_field = "Metal"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_metal-container']", 
                                             str(row_data["MetalName"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # Category / Old Metal Category
            if row_data.get("CategoryName"):
                current_field = "Category"
                if str(row_data.get("MetalType")) == "2": # Old Metal
                    Function_Call.dropdown_select(self, "//span[@id='select2-oldcategory-container']", 
                                                 str(row_data["CategoryName"]), 
                                                 '//span[@class="select2-search select2-search--dropdown"]/input')
                else: # Ornament
                    Function_Call.dropdown_select(self, "//span[@id='select2-select_category-container']", 
                                                 str(row_data["CategoryName"]), 
                                                 '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # Product
            if row_data.get("ProductName"):
                current_field = "Product"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_product-container']", 
                                             str(row_data["ProductName"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # --- 6. Quantity Fields ---
            if str(row_data.get("BalType")) == "1": # Metal
                current_field = "Metal Quantity Fields"
                if row_data.get("GrossWgt"):
                    Function_Call.fill_input2(self, "//input[@name='smth_cmpy_stk[op_wgt]']", str(row_data["GrossWgt"]))
                if row_data.get("NetWgt"):
                    Function_Call.fill_input2(self, "//input[@name='smth_cmpy_stk[op_net_wgt]']", str(row_data["NetWgt"]))
                if row_data.get("DiaWgt"):
                    Function_Call.fill_input2(self, "//input[@name='smth_cmpy_stk[op_dia_wgt]']", str(row_data["DiaWgt"]))
                if row_data.get("PureWgt"):
                    Function_Call.fill_input2(self, "//input[@name='smth_cmpy_stk[op_pure_wgt]']", str(row_data["PureWgt"]))
                
                # Weight Type (Credit/Debit)
                if row_data.get("WtType"):
                    val = str(row_data["WtType"]).strip().lower()
                    wt_map = {"credit": "1", "debit": "2"}
                    final_wt = wt_map.get(val, val)
                    Function_Call.click(self, f"//input[@name='smth_cmpy_stk[wt_receipt_type]' and @value='{final_wt}']")

            # --- 7. Value Fields (Amount) ---
            if row_data.get("Amount"):
                current_field = "Amount Value Fields"
                Function_Call.fill_input2(self, "//input[@name='smth_cmpy_stk[op_amt]']", str(row_data["Amount"]))
                if row_data.get("AmtType"):
                    val = str(row_data["AmtType"]).strip().lower()
                    amt_map = {"credit": "1", "debit": "2"}
                    final_amt = amt_map.get(val, val)
                    Function_Call.click(self, f"//input[@name='smth_cmpy_stk[amt_receipt_type]' and @value='{final_amt}']")

            # --- 8. Remark & Save ---
            current_field = "Remark and Save"
            if row_data.get("Remark"):
                Function_Call.fill_input2(self, "//textarea[@name='smth_cmpy_stk[remark]']", str(row_data["Remark"]))
            
            Function_Call.click(self, "//button[@id='smth_cmpy_bal_submit']")
            sleep(3)

            # Verify response
            try:
                # 1. Check for Flash Success Message on page (as seen in screenshot)
                try:
                    success_xpath = "//div[contains(@class, 'alert-success')] | //div[contains(@class, 'success')]"
                    success_el = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath)))
                    success_text = success_el.text.strip()
                    if "Successfully" in success_text:
                        return self.verify_in_list(row_data)
                except:
                    pass # No flash success message yet

                # 2. Try browser alert first (classic confirmation)
                try:
                    alert = driver.switch_to.alert
                    alert_text = alert.text
                    alert.accept()
                    if "Successfully" in alert_text:
                        return self.verify_in_list(row_data)
                    else:
                        return ("Fail", f"Alert: {alert_text}")
                except:
                    pass # No browser alert

                # 3. Fallback: Check for redirect to list page
                if "list" in driver.current_url:
                    return self.verify_in_list(row_data)
                else:
                    return ("Fail", "Save clicked but no confirmation or redirect received.")

            except Exception as e:
                return ("Fail", f"Error during verification: {str(e)}")

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def verify_in_list(self, row_data):
        """Verify the recorded entry in the list page table."""
        driver, wait = self.driver, self.wait
        try:
            # Navigate to list if not already there
            if "smith_cmpy_op_bal/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/smith_cmpy_op_bal/list")
                sleep(2)
            
            # --- 1. Select Date Range (Today) ---
            # Click the Date range picker button
            Function_Call.click(self, "//button[contains(., 'Date range picker')]")
            sleep(1)
            # Select 'Today' from the daterangepicker list
            Function_Call.click(self, "//li[contains(text(), 'Today')]")
            sleep(2)
            
            # --- 2. Click Search ---
            # Blue search button to refresh the list
            Function_Call.click(self, "//button[contains(text(), 'Search')]")
            sleep(2)

            # Wait for the summary table to load
            wait.until(EC.presence_of_element_located((By.XPATH, "//table[contains(@id, 'list')]")))
            
            # Check the top row for matching data
            # XPath for row 1, various columns based on screenshot
            row1_xpath = "//table[contains(@id, 'list')]/tbody/tr[1]"
            
            ui_ref_no = driver.find_element(By.XPATH, f"{row1_xpath}/td[2]").text.strip()
            ui_karigar = driver.find_element(By.XPATH, f"{row1_xpath}/td[6]").text.strip()
            ui_metal = driver.find_element(By.XPATH, f"{row1_xpath}/td[7]").text.strip()
            ui_category = driver.find_element(By.XPATH, f"{row1_xpath}/td[8]").text.strip()
            ui_product = driver.find_element(By.XPATH, f"{row1_xpath}/td[9]").text.strip()
            
            print(f"🔍 List Verification: Karigar={ui_karigar}, Metal={ui_metal}, Category={ui_category}, Product={ui_product}")
            
            # Validate identifying fields
            mismatches = []
            if row_data.get("KarigarName") and str(row_data["KarigarName"]).strip().lower() not in ui_karigar.lower():
                mismatches.append(f"Karigar: {ui_karigar}")
            
            if row_data.get("MetalName") and str(row_data["MetalName"]).strip().lower() not in ui_metal.lower():
                mismatches.append(f"Metal: {ui_metal}")

            if mismatches:
                return ("Fail", f"List Verification Failed: {', '.join(mismatches)}", ui_ref_no)
            
            return ("Pass", f"Verified in List: {ui_ref_no} | {ui_karigar} | {ui_metal}", ui_ref_no)

        except Exception as e:
            return ("Fail", f"Error in List Verification: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if captured_id:
                # Store captured Ref No in column 21
                sheet.cell(row=row_num, column=19, value=captured_id)
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass
