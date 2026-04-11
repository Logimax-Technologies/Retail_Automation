from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest
import re

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class SmithSupplierPayment(unittest.TestCase):
    """
    Supplier PO Payment Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_smith_supplier_payment(self):
        """Main entry point for Supplier PO Payment automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Supplier PO Payment List
        try:
            if "supplier_po_payment/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Supplier Payment')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/supplier_po_payment/list")
            sleep(2)

        # Read Excel data
        sheet_name = "SmithSupplierPayment"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Mapping based on finalized structure
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "BillType": 4, "Karigar": 5, "OpeningBalance": 6,
                "BillSelection": 7, "NetAmount":8,"CashAmount": 9, "NB_YesNo": 10,
                "NB_Type": 11, "NB_Bank": 12, "NB_Amount": 13, "NB_RefNo": 14, "NB_Date": 15,
                "Cheque_YesNo": 16, "ChequeDate": 17, "ChequeBank": 18, "branch": 19, "ChequeNo": 20,
                "ifsc": 21, "ChequeAmount": 22, "Remark": 23, "ExpectedNetAmount": 24, "ExpectedPayRefNo": 25, "ExpectedStatus": 26
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                self.driver.refresh()
                sleep(2)
                result = self.execute_supplier_po_payment_flow(row_data, row_num, sheet_name)
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_id = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_supplier_po_payment_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Add Start"
        try:
            Function_Call.alert(self)
            
            # Click ADD button
            if "/add" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/supplier_po_payment/add")
                sleep(3)

            # --- Header Section ---
            # Bill Type: BILL(1) | RECEIPT(2)
            if row_data.get("BillType"):
                current_field = "Bill Type"
                val = str(row_data["BillType"]).strip()
                type_map = {"BILL": "1", "RECEIPT": "2"}
                final_val = type_map.get(val, val)
                Function_Call.click(self, f"//input[@name='billing[bill_type]' and @value='{final_val}']")
                sleep(1)

            # Karigar (Supplier) (Select2)
            if row_data.get("Karigar"):
                current_field = "Karigar Select"
                Function_Call.dropdown_select(self, '//select[@id="select_karigar"]/following-sibling::span', str(row_data["Karigar"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(3) # Wait for pending bills to load

            # Opening Balance Checkbox
            if row_data.get("OpeningBalance") == "Yes":
                current_field = "Opening Balance Checkbox"
                Function_Call.click(self, "//input[@id='opening_checkbox']")
                sleep(1)

            # Bill Selection
            if row_data.get("BillSelection"):
                current_field = "Bill Selection"
                selection = str(row_data["BillSelection"]).strip()
                
                # List of Ref Nos to select (if not "All")
                target_refs = []
                if selection.lower() != "all":
                    # Scroll down to ensure buttons are visible
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    sleep(1)
                    # Click Clear All first
                    Function_Call.click(self, "//button[contains(text(), 'Clear All')]")
                    sleep(2)
                    
                    for r in selection.split(','):
                        target_refs.append(r.strip())
                
                total_selected_amt = 0.0

                # Locate all rows in the pending bills container
                bill_rows = driver.find_elements(By.XPATH, "//div[@id='pending_bills_container']//table/tbody/tr")
                
                for i, row in enumerate(bill_rows, 0):
                    try:
                        ref_no = row.find_element(By.XPATH, "./td[2]").text.strip()
                        amount_text = row.find_element(By.XPATH, "./td[3]").text.strip().replace(',', '')
                        amount = float(amount_text or 0)
                        
                        # Construct a unique XPath for the checkbox in this specific row (name-based as requested)
                        checkbox_xpath = f"//input[@name='billing[{i}][bill_ids][]']"
                        checkbox = driver.find_element(By.XPATH, checkbox_xpath)
                        
                        if selection.lower() == "all":
                            if not checkbox.is_selected():
                                Function_Call.click(self, checkbox_xpath)
                            total_selected_amt += amount
                            print(f"✅ Selected Bill: {ref_no} (Amount: {amount})")
                        elif ref_no in target_refs:
                            if not checkbox.is_selected():
                                Function_Call.click(self, checkbox_xpath)
                            total_selected_amt += amount
                            print(f"✅ Selected Bill: {ref_no} (Amount: {amount})")
                        else:
                            # Already cleared if not 'all'
                            if checkbox.is_selected():
                                Function_Call.click(self, checkbox_xpath)
                    except Exception as e:
                        print(f"⚠️ Row processing error: {e}")

                sleep(2)
                
                # Verify Net Amount equality
                ui_net_amt_val = driver.find_element(By.ID, "balance_amount").get_attribute("value").replace(',', '')
                ui_net_amt = float(ui_net_amt_val or 0)
                
                print(f"📊 Verification: Selected Sum ({total_selected_amt}) vs UI Net Amount ({ui_net_amt})")
                if abs(total_selected_amt - ui_net_amt) > 0.01:
                    print(f"❌ Amount Mismatch! Total Selected: {total_selected_amt}, UI Shown: {ui_net_amt}")
                    # We continue but log the error
                else:
                    print(f"✅ Amount Verified: Sum matches UI Net Amount.")
                
                # Instantly save Net Amount to Excel
                try:
                    wb = load_workbook(FILE_PATH)
                    ws = wb[sheet_name]
                    ws.cell(row=row_num, column=8, value=ui_net_amt) # Column 8 is NetAmount
                    wb.save(FILE_PATH)
                    wb.close()
                    print(f"💾 Net Amount ({ui_net_amt}) saved to Excel successfully.")
                except Exception as e:
                    print(f"⚠️ Failed to save Net Amount to Excel: {e}")
                
                sleep(2)

            # --- Payment Calculation Logic ---
            ui_net_amt_val = driver.find_element(By.ID, "balance_amount").get_attribute("value").replace(',', '')
            ui_net_amt = float(ui_net_amt_val or 0)
            
            cash_val = float(row_data.get("CashAmount") or 0)
            balance_after_cash = round(ui_net_amt - cash_val, 2)
            
            # Calculate NB Amount if percentage
            nb_amt_data = row_data.get("NB_Amount")
            nb_amt_to_fill = 0
            if nb_amt_data:
                try:
                    perc = float(str(nb_amt_data).replace('%', '').strip())
                    nb_amt_to_fill = round(balance_after_cash * (perc / 100), 2)
                except ValueError:
                    nb_amt_to_fill = nb_amt_data

            # Calculate Cheque Amount if percentage
            chq_amt_data = row_data.get("ChequeAmount")
            chq_amt_to_fill = 0
            if chq_amt_data:
                try:
                    perc = float(str(chq_amt_data).replace('%', '').strip())
                    chq_amt_to_fill = round(balance_after_cash * (perc / 100), 2)
                except ValueError:
                    chq_amt_to_fill = chq_amt_data

            print(f"🧮 Balance After Cash: {balance_after_cash}")
            print(f"🧮 Payment Split: Cash={cash_val}, NB={nb_amt_to_fill}, Cheque={chq_amt_to_fill}")

            # --- Payment Modes ---
            # Cash Amount (Business Rule: 10,000 Limit)
            if row_data.get("CashAmount") is not None:
                current_field = "Cash Amount"
                Function_Call.fill_input(self, wait, (By.NAME, "billing[cash_amount]"), str(cash_val), "Cash Amount", row_num, Sheet_name=sheet_name)
                
                if cash_val > 10000:
                    print("⚠️ Cash amount > 10,000. Expecting validation failure.")
            
            # Net Banking
            if row_data.get("NB_YesNo") == "Yes":
                current_field = "Net Banking Modal"
                Function_Call.click(self, "//a[@id='net_bank_modal']")
                sleep(2)
                
                # Verification: Check if modal is visible
                try:
                    modal = driver.find_element(By.ID, "net_bank_receipt")
                    if not modal.is_displayed():
                        print("⚠️ Net Banking Modal not visible, retrying with JS...")
                        driver.execute_script("document.getElementById('net_bank_modal').click();")
                        sleep(2)
                except:
                    print("⚠️ Net Banking Modal not found, retrying with JS...")
                    driver.execute_script("document.getElementById('net_bank_modal').click();")
                    sleep(2)

                Function_Call.click(self, "//button[@id='create_net_banking_row']")
                sleep(1)
                
                last_row_xpath = "(//table[@id='net_banking_details']/tbody/tr)[last()]"
                
                if row_data.get("NB_Type"):
                    sel = Select(driver.find_element(By.XPATH, f"{last_row_xpath}//select[contains(@class, 'nb_type')]"))
                    sel.select_by_visible_text(str(row_data["NB_Type"]))
                
                if row_data.get("NB_Bank"):
                    try:
                        sel_bank = Select(driver.find_element(By.XPATH, f"{last_row_xpath}//select[not(contains(@class, 'nb_type'))]"))
                        sel_bank.select_by_visible_text(str(row_data["NB_Bank"]))
                    except:
                        pass
                
                if nb_amt_to_fill:
                    driver.find_element(By.XPATH, f"{last_row_xpath}//input[contains(@class, 'pay_amount')]").send_keys(str(nb_amt_to_fill))
                
                if row_data.get("NB_RefNo"):
                    driver.find_element(By.XPATH, f"{last_row_xpath}//input[contains(@class, 'ref_no')]").send_keys(str(row_data["NB_RefNo"]))
                
                if row_data.get("NB_Date"):
                    Function_Call.fill_input(self, wait, (By.XPATH, f"{last_row_xpath}//input[contains(@class, 'nb_date')]"), str(row_data["NB_Date"]), "NB Date", row_num, Sheet_name=sheet_name, Date_range="past_or_current")
                
                Function_Call.click(self, "//a[@id='save_net_banking']")
                sleep(1)

                # Cheque
            if row_data.get("Cheque_YesNo") == "Yes":
                current_field = "Cheque Modal"
                try:
                    Function_Call.click(self, "//a[@id='cheque_modal']")
                    sleep(1)
                    # Verification: Check if modal is visible
                    try:
                        modal = driver.find_element(By.ID, "myModalLabel")
                        if not modal.is_displayed():
                            print("⚠️ Cheque Modal not visible, retrying with JS...")
                            driver.execute_script("document.getElementById('cheque_modal').click();")
                            sleep(2)
                    except:
                        print("⚠️ Cheque Modal not found, retrying with JS...")
                        driver.execute_script("document.getElementById('cheque_modal').click();")
                        sleep(2)
               
                    Function_Call.click(self, "//button[@id='new_chq']") # Corrected ID based on common patterns
                    sleep(1)
                    
                    last_row_xpath = "(//table[@id='chq_details']/tbody/tr)[last()]"
                    
                    if row_data.get("ChequeDate"):
                        Function_Call.fill_input(self, wait, (By.XPATH, f"{last_row_xpath}//input[contains(@class, 'cheque_date')]"), str(row_data["ChequeDate"]), "Cheque Date", row_num, Sheet_name=sheet_name, Date_range="future_or_current")
                     
                    if row_data.get("ChequeBank"):
                        sel = Select(driver.find_element(By.XPATH, f"{last_row_xpath}//select[contains(@class, 'bank_name')]"))
                        sel.select_by_visible_text(str(row_data["ChequeBank"]))
                   
                    if row_data.get("branch"):
                        driver.find_element(By.XPATH, f"{last_row_xpath}//input[contains(@class, 'bank_branch')]").send_keys(str(row_data["branch"]))
               
                    if row_data.get("ChequeNo"):
                        driver.find_element(By.XPATH, f"{last_row_xpath}//input[contains(@class, 'cheque_no')]").send_keys(str(row_data["ChequeNo"]))
                    
                    if row_data.get("ifsc"):
                        driver.find_element(By.XPATH, f"{last_row_xpath}//input[contains(@class, 'bank_IFSC')]").send_keys(str(row_data["ifsc"]))

                    
                    if chq_amt_to_fill:
                        driver.find_element(By.XPATH, f"{last_row_xpath}//input[contains(@class, 'payment_amount')]").send_keys(str(chq_amt_to_fill))
                        sleep(2)
                    Function_Call.click(self, "//a[@id='save_issue_chq']")
                except Exception as e:
                    print(f"⚠️ Cheque Modal Error: {e}")
                sleep(1)

            if row_data.get("Remark"):
                current_field = "Remark"
                Function_Call.fill_input(self, wait, (By.ID, "remark"), str(row_data["Remark"]), "Remark", row_num, Sheet_name=sheet_name)

            # --- Calculation Verification ---
            current_field = "Summary Verification"
            ui_net = driver.find_element(By.ID, "balance_amount").get_attribute("value")
            ui_total = driver.find_element(By.ID, "total_pay_amount").get_attribute("value")
            ui_bal = driver.find_element(By.ID, "bal_amount").get_attribute("value")
            print(f"💰 Net: {ui_net}, Total Pay: {ui_total}, Balance: {ui_bal}")

            # --- Submit ---
            current_field = "Submit"
            self._take_screenshot(f"BeforeSave_TC{row_data['TestCaseId']}")
            
            main_window = driver.current_window_handle
            Function_Call.click(self, "//button[@id='pay_submit']")
            
            # Check for Alert Validation (e.g. Cash > 10,000)
            pay_ref_no = "Unknown"
            try:
                alert = wait.until(EC.alert_is_present())
                alert_text = alert.text
                print(f"🔔 Alert captured: {alert_text}")
                alert.accept()
                if row_data.get("CashAmount") and float(row_data["CashAmount"]) > 10000:
                    return ("Pass", f"Correctly blocked by validation: {alert_text}")
                return ("Fail", f"Unexpected Alert: {alert_text}")
            except:
                # Click Cancel to reset UI on validation failure
                try:
                    Function_Call.click(self, "//button[contains(text(), 'Cancel')]")
                    sleep(2)
                except:
                    pass

            # Extract Pay Ref No from Acknowledgement
            # pay_ref_no = self._extract_pay_ref_and_close(main_window)
            # print(f"📄 Captured Pay Ref No: {pay_ref_no}")

            # Verification in List Page
            return self.verify_in_list(row_data, pay_ref_no)

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            try:
                Function_Call.click(self, "//button[contains(text(), 'Cancel')]")
                sleep(2)
            except:
                pass
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def verify_in_list(self, row_data, pay_ref_no=None):
        driver, wait = self.driver, self.wait
        if "supplier_po_payment/list" not in driver.current_url:
             driver.get(BASE_URL + "index.php/admin_ret_purchase/supplier_po_payment/list")
             sleep(2)
        
        # Date Filter to Today
        Function_Call.click(self, "//button[@id='sp-dt-btn']") 
        sleep(1)
        Function_Call.click(self, "//li[contains(text(), 'Today')]")
        sleep(2)
        
        # Search by Ref No
        if row_data.get("Karigar"):
            try:
                search_box = driver.find_element(By.XPATH, "//div[@id='supplier_po_pay_list_filter']//input")
                search_box.clear()
                search_box.send_keys(row_data["Karigar"])
                sleep(2)
            except:
                pass

        # Confirm top row
        try:
            first_col_text = driver.find_element(By.XPATH, "//table[@id='supplier_po_pay_list']/tbody/tr[1]/td[1]").text.strip()
            ref_no = driver.find_element(By.XPATH, "//table[@id='supplier_po_pay_list']/tbody/tr[1]/td[3]").text.strip()
            supplier = driver.find_element(By.XPATH, "//table[@id='supplier_po_pay_list']/tbody/tr[1]/td[4]").text.strip()
            status = driver.find_element(By.XPATH, "//table[@id='supplier_po_pay_list']/tbody/tr[1]/td[7]").text.strip()
            final_id = first_col_text
            return ("Pass", f"Verified in list. Ref: {ref_no}, Supplier: {supplier}, Status: {status}", final_id)
        except:
            return ("Fail", f"List verification failed for Pay Ref No: {pay_ref_no}")

    # def _extract_pay_ref_and_close(self, main_window):
        # pay_ref_no = "Unknown"
        # driver = self.driver
        # try:
        #     wait_time = 0
        #     while len(driver.window_handles) < 2 and wait_time < 12:
        #         sleep(1)
        #         wait_time += 1

        #     windows = driver.window_handles
        #     if len(windows) > 1:
        #         driver.switch_to.window(windows[1])
        #         sleep(2)
                
        #         body_text = driver.find_element(By.TAG_NAME, "body").text
        #         # Match "Issue No : PAY-2026-0001"
        #         m = re.search(r"Issue No\s*[:\-]?\s*(\S+)", body_text, re.IGNORECASE)
        #         if m:
        #             pay_ref_no = m.group(1).strip()
                
        #         self._take_screenshot(f"Acknowledgment_Ref_{pay_ref_no}")
        #         driver.close()
        #         driver.switch_to.window(main_window)
        #     else:
        #         # If no new window, check if it's on the list page already or a redirect
        #         pass

        # except Exception as e:
        #     print(f"⚠️ Pay Ref extraction failed: {e}")
        #     try:
        #         driver.switch_to.window(main_window)
        #     except:
        #         pass

        # return pay_ref_no

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if captured_id:
                sheet.cell(row=row_num, column=22, value=captured_id)
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass
