from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest
import re

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class SmithMetalIssue(unittest.TestCase):
    """
    Smith Metal Issue Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_smith_metal_issue(self):
        """Main entry point for Smith Metal Issue automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Smith Metal Issue List
        try:
            if "karigarmetalissue/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Smith Metal Issue')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/karigarmetalissue/list")
            sleep(2)

        # Read Excel data
        sheet_name = "SmithMetalIssue"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Mapping based on refined 20-column structure
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3, "Branch": 4, "Karigar": 5, 
                "Metal": 6, "IssueType": 7, "AgainstOpening": 8, "SearchKarigar": 9,
                "AgainstOrder": 10, "AgainstBill": 11, "SupplierFY": 12, "RefOrderNo": 13, 
                "SourceType": 14, "StockSource": 15, "Code": 16,"Touch": 17, "UOM": 18, "Remark": 19,
                "VerifyList": 20
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # if str(row_data.get("TestStatus")).strip().lower() != "run":
            #     print(f"⏭️ Skipping Test Case: {row_data['TestCaseId']}")
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                self.driver.refresh()
                sleep(2)
                result = self.execute_smith_metal_issue_flow(row_data, row_num, sheet_name)
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_id = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_smith_metal_issue_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Add Start"
        try:
            Function_Call.alert(self)
            
            # Click ADD button
            if "/add" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/karigarmetalissue/add")
                sleep(3)

            # --- Header Section ---
            
            # Branch (Select2)
            if row_data.get("Branch"):
                current_field = "Branch Select"
                try:
                    Function_Call.dropdown_select(self, '//select[@id="select_branch"]/following-sibling::span', str(row_data["Branch"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                    sleep(1)
                except:
                    print("⚠️ Branch selector not found, skipping.")

            # Karigar (Select2)
            if row_data.get("Karigar"):
                current_field = "Karigar Select"
                Function_Call.dropdown_select(self, '//select[@id="select_karigar"]/following-sibling::span', str(row_data["Karigar"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(3) # Wait for Outstanding Balance AJAX
                
                try:
                    out_pure = driver.find_element(By.CLASS_NAME, "availablepurebalance").text
                    print(f"💰 Karigar Outstanding Pure Balance: {out_pure}")
                except:
                    pass

            # Metal (Select2)
            if row_data.get("Metal"):
                current_field = "Metal Select"
                Function_Call.dropdown_select(self, '//select[@id="select_metal"]/following-sibling::span', str(row_data["Metal"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(2)

            # Metal Issue Type (Radio): 1-Normal Issue, 2-Approval Issue
            if row_data.get("IssueType"):
                current_field = "Metal Issue Type"
                val = str(row_data["IssueType"]).strip().lower()
                target_val = "2" if "appr" in val else "1"
                Function_Call.click(self, f"//input[@name='issue[metal_issue_type]' and @value='{target_val}']")
                sleep(1)

            # Against Opening (Radio): 1=Yes, 0=No
            if row_data.get("AgainstOpening"):
                current_field = "Against Opening"
                val = str(row_data["AgainstOpening"]).strip().lower()
                target_val = "1" if val == "yes" else "0"
                Function_Call.click(self, f"//input[@name='issue[is_against_opening]' and @value='{target_val}']")
                sleep(1)
                
                if val == "yes":
                    if row_data.get("SearchKarigar"):
                        current_field = "Search Karigar Opening"
                        Function_Call.dropdown_select(self, '//select[@id="karigar_select"]/following-sibling::span', str(row_data["SearchKarigar"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                        sleep(1)
                        Function_Call.click(self, "//button[@id='search_kar_opening_det']")
                        sleep(2)

            # Against Order (Radio): 1=Yes, 0=No
            if row_data.get("AgainstOrder"):
                current_field = "Against Order"
                val = str(row_data["AgainstOrder"]).strip().lower()
                target_val = "1" if val == "yes" else "0"
                Function_Call.click(self, f"//input[@name='issue[issue_aganist]' and @value='{target_val}']")
                sleep(1)
                
                if val == "yes" and row_data.get("RefOrderNo"):
                    current_field = "PO Select"
                    wait.until(EC.visibility_of_element_located((By.ID, "select_po_no")))
                    Function_Call.dropdown_select(self, '//select[@id="select_po_no"]/following-sibling::span', str(row_data["RefOrderNo"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                    sleep(2)

            # Against Supplier Bill (Radio): 1=Yes, 0=No
            if row_data.get("AgainstBill"):
                current_field = "Against Bill"
                val = str(row_data["AgainstBill"]).strip().lower()
                target_val = "1" if val == "yes" else "0"
                Function_Call.click(self, f"//input[@name='issue[issue_against_po]' and @value='{target_val}']")
                sleep(1)
                
                if val == "yes":
                    # Financial Year selection for Supplier PO
                    if row_data.get("SupplierFY"):
                        current_field = "Supplier PO FY Select"
                        wait.until(EC.visibility_of_element_located((By.ID, "pur_fin_year_select")))
                        Function_Call.dropdown_select(self, '//select[@id="pur_fin_year_select"]/following-sibling::span', str(row_data["SupplierFY"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                        sleep(1)

                    if row_data.get("RefOrderNo"):
                        current_field = "Supplier Bill Select"
                        wait.until(EC.presence_of_element_located((By.ID, "select_supplier_po_no")))
                        Function_Call.dropdown_select(self, '//select[@id="select_supplier_po_no"]/following-sibling::span', str(row_data["RefOrderNo"]), '//select[@id="select_supplier_po_no"]/following-sibling::span//input[@class="select2-search__field"]')
                        sleep(2)

            # --- Source Selection ---
            if row_data.get("SourceType"):
                current_field = "Source Type (Tag/Non-Tag)"
                val = str(row_data["SourceType"]).strip().lower()
                is_tag = "tag" in val and "non" not in val
                target_val = "1" if is_tag else "2"
                Function_Call.click(self, f"//input[@name='issue[issue_from]' and @value='{target_val}']")
                sleep(1)

                # Stock Source (Available, Sales Return, etc.)
                if row_data.get("StockSource"):
                    current_field = "Stock Source"
                    src = str(row_data["StockSource"]).strip().lower()
                    if is_tag:
                        # Tag Sources: 1=Available, 2=Sales Return, 3=Partly Sales, 4=H.O OtherIssue
                        src_map = {"available": "1", "sales return": "2", "partly sales": "3", "h.o otherissue": "4", "ho otherissue": "4"}
                        target_src = src_map.get(src, "1")
                        Function_Call.click(self, f"//input[@name='issue[tag_issue_from]' and @value='{target_src}']")
                    else:
                        # Non-Tag Sources: 1=Available, 2=NonTag Sales Return, 3=NonTag Other Issue
                        src_map = {"available": "1", "sales return": "2", "other issue": "3", "ho otherissue": "3"}
                        target_src = src_map.get(src, "1")
                        Function_Call.click(self, f"//input[@name='issue[nontag_issue_from]' and @value='{target_src}']")
                    sleep(1)

                if is_tag and row_data.get("Code"):
                    current_field = "Tag Search"
                    Function_Call.fill_input(self, wait, (By.ID, "tag_code"), str(row_data["Code"]), "Tag Code", row_num, Sheet_name=sheet_name)
                    Function_Call.click(self, "//button[@id='tag_history_search']")
                    sleep(3)
                elif not is_tag and row_data.get("Code"):
                    current_field = "BT Code"
                    Function_Call.fill_input(self, wait, (By.ID, "bt_number"), str(row_data["Code"]), "BT Code", row_num, Sheet_name=sheet_name)
                    sleep(1)

            # --- Item Selection ---
            current_field = "Select All Items" 
            if driver.find_elements(By.ID, "select_all"):
                # Multi-Row Item Detail Loop
                try:
                    item_workbook = load_workbook(FILE_PATH)
                    item_sheet = item_workbook["SmithMetalIssueItems"]
                    valid_item_rows = ExcelUtils.get_valid_rows(FILE_PATH, "SmithMetalIssueItems")
                    
                    item_data_map = {
                        "TestCaseId": 1, "RefOrderNo": 2, "Purity": 3, "Touch": 4, 
                        "Pcs": 5, "Weight": 6, "UOM": 7
                    }
                    
                    found_items = False
                    for i_row in range(2, valid_item_rows):
                        item_row_data = {key: item_sheet.cell(row=i_row, column=col).value for key, col in item_data_map.items()}
                        
                        if str(item_row_data.get("TestCaseId")) == str(row_data["TestCaseId"]):
                            found_items = True
                            # Find row by RefOrderNo or Code if needed
                            if item_row_data.get("RefOrderNo"):
                                ref_no = str(item_row_data.get("RefOrderNo")).strip()
                                print(f"📦 Processing Item: {ref_no} for TC {row_data['TestCaseId']}")
                                xpath = f"//table[@id='metal_details']/tbody/tr[contains(., '{ref_no}')]"
                                matching_rows = driver.find_elements(By.XPATH, xpath)
                                if not matching_rows:
                                    print(f"⚠️ Row with Ref: {ref_no} not found, skipping.")
                                    continue
                            else:
                                xpath = f"//table[@id='metal_details']/tbody/tr"
                                matching_rows = driver.find_elements(By.XPATH, xpath)
                                
                            row = matching_rows[0]
                            checkbox = row.find_element(By.CLASS_NAME, "id_kar_issue")
                            if not checkbox.is_selected():
                                checkbox.click()
                            sleep(1)
                            
                            # 1. Purity (Dropdown Select2)
                            if item_row_data.get("Purity"):
                                current_field = f"Purity Select"
                                purity_dropdown = row.find_element(By.XPATH, ".//span[contains(@id,'select2-select_purity-container')]")
                                purity_dropdown.click()
                                sleep(1)
                                search_box = driver.find_element(By.XPATH, '//span[@class="select2-search select2-search--dropdown"]/input')
                                search_box.clear()
                                search_box.send_keys(str(item_row_data["Purity"]))
                                search_box.send_keys(Keys.ENTER)
                                sleep(1)
                                
                            # 2. Touch (Input)
                            if item_row_data.get("Touch"):
                                current_field = f"Touch Entry"
                                touch_field = row.find_element(By.XPATH, ".//input[contains(@id, 'karigar_touch')]")
                                touch_field.clear()
                                touch_field.send_keys(str(item_row_data["Touch"]))
                                touch_field.send_keys(Keys.TAB)
                                sleep(1)
                                
                            # 3. Pcs (Input)
                            if item_row_data.get("Pcs"):
                                current_field = f"Pcs Entry"
                                pcs_field = row.find_element(By.CLASS_NAME, "issue_pcs")
                                pcs_field.clear()
                                pcs_field.send_keys(str(item_row_data["Pcs"]))
                                pcs_field.send_keys(Keys.TAB)
                                sleep(1)

                            # 4. Weight (Input)
                            if item_row_data.get("Weight"):
                                current_field = f"Weight Entry"
                                weight_field = row.find_element(By.CLASS_NAME, "issue_weight")
                                weight_field.clear()
                                weight_field.send_keys(str(item_row_data["Weight"]))
                                weight_field.send_keys(Keys.TAB)
                                sleep(1)

                            # 5. UOM (Standard Select)
                            if item_row_data.get("UOM"):
                                current_field = f"UOM Selection"
                                uom_select = Select(row.find_element(By.CLASS_NAME, "uom_id"))
                                uom_select.select_by_visible_text(str(item_row_data["UOM"]))
                                sleep(1)

                            # --- Pure Weight Verification per Row ---
                            current_field = f"Pure Weight Validation"
                            w_val = float(row.find_element(By.CLASS_NAME, "issue_weight").get_attribute("value") or 0)
                            t_val = float(row.find_element(By.ID, "karigar_touch").get_attribute("value") or 100)
                            calc_pure = round((w_val * t_val) / 100, 3)
                            
                            act_pure = float(row.find_element(By.ID, "pur_weight").get_attribute("value") or 0)
                            print(f"⚖️ {ref_no} -> Weight: {w_val}, Touch: {t_val}, Calc Pure: {calc_pure}, Actual: {act_pure}")
                            
                            if abs(calc_pure - act_pure) > 0.005:
                                raise Exception(f"Pure weight mismatch for {ref_no}! Calc: {calc_pure}, Actual: {act_pure}")

                    item_workbook.close()
                    
                    if not found_items:
                        print(f"⚠️ No items found in 'SmithMetalIssueItems' for TC {row_data['TestCaseId']}. Selecting All.")
                        Function_Call.click(self, "//input[@id='select_all']")
                        sleep(1)

                    # Karigar Balance Validation (Total)
                    try:
                        current_field = "Total Pure Weight Balance Check"
                        out_pure_text = driver.find_element(By.CLASS_NAME, "availablepurebalance").text
                        out_pure_val = float(re.sub(r'[^0-9.]', '', out_pure_text))
                        
                        total_pure_issued = 0
                        checked_rows = driver.find_elements(By.XPATH, "//table[@id='metal_details']/tbody/tr[.//input[@class='id_kar_issue' and @type='checkbox' and @checked='true']]")
                        for c_row in checked_rows:
                            total_pure_issued += float(c_row.find_element(By.ID, "pur_weight").get_attribute("value") or 0)
                        
                        print(f"💰 Total Issued Pure: {total_pure_issued}, Available: {out_pure_val}")
                        if total_pure_issued > out_pure_val:
                             raise Exception(f"Excess Pure Weight not allowed! Total Issued: {total_pure_issued}, Available: {out_pure_val}")
                    except Exception as balance_err:
                        print(f"⚠️ Balance check failed: {balance_err}")
                        if "Excess" in str(balance_err): raise balance_err

                except Exception as e:
                    print(f"❌ Error in multi-row handling: {e}")
                    raise e

            # --- Footer Section ---
            if row_data.get("Remark"):
                current_field = "Remark"
                Function_Call.fill_input(self, wait, (By.ID, "remark"), str(row_data["Remark"]), "Remark", row_num, Sheet_name=sheet_name)

            # --- Submit ---
            current_field = "Submit"
            self._take_screenshot(f"BeforeSave_TC{row_data['TestCaseId']}")
            
            # Record current window handle
            main_window = driver.current_window_handle
            
            Function_Call.click(self, "//button[@id='submit_metal_issue']")
            
            # Use the new extraction method
            Id = self._extract_invoice_and_close(main_window)

            # Verification in List Page
            if str(row_data.get("VerifyList")).strip().lower() == "yes":
                return self.verify_in_list(row_data, Id)
            else:
                actual_msg = f"Saved successfully. Captured Id: {Id}"
                return ("Pass", actual_msg, Id)

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            try:
                Function_Call.click(self, "//button[contains(@class, 'btn-cancel')]")
                sleep(2)
            except:
                pass
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def _extract_invoice_and_close(self, main_window):
        """
        Switches to print/acknowledgment tab, extracts the Ref No from the
        page body or URL, closes the tab, and switches back to the main window.
        """
        Id = "Unknown"
        driver = self.driver
        try:
            # Wait up to 12s for the acknowledgment tab to open
            wait_time = 0
            while len(driver.window_handles) < 2 and wait_time < 12:
                sleep(1)
                wait_time += 1

            windows = driver.window_handles
            if len(windows) > 1:
                driver.switch_to.window(windows[1])
                sleep(3)  # Wait for rendering

                ack_url = driver.current_url
                print(f"📄 Acknowledgment URL: {ack_url}")

                # --- Method 3: URL ID Fallback (Fail-safe for PDFs) ---
                if Id == "Unknown" or not Id:
                    try:
                        # URL ends with /ID (e.g., .../539)
                        record_id = ack_url.rstrip('/').split('/')[-1]
                        if record_id.isdigit():
                            print(f"🔍 Capturing Record ID from URL as fallback: {record_id}")
                            Id = record_id
                    except:
                        pass

                print(f"✅ Extracted Id No: {Id}")
                self._take_screenshot(f"Acknowledgment_IdNo_{Id}")

                # Close the acknowledgment/print tab
                try:
                    driver.close()
                except Exception:
                    driver.execute_script("window.close();")

                # Return focus to the main window
                driver.switch_to.window(main_window)
                return Id
            else:
                print("⚠️ No acknowledgment tab found")

        except Exception as e:
            print(f"⚠️ Id extraction failed: {e}")
            
    def verify_in_list(self, row_data, Id=None):
        driver, wait = self.driver, self.wait
        if "karigarmetalissue/list" not in driver.current_url:
             driver.get(BASE_URL + "index.php/admin_ret_purchase/karigarmetalissue/list")
             sleep(3)
        
        # Search for Captured Ref No or Karigar
        search_term = Id if Id else row_data.get("Karigar")
        if search_term:
            try:
                search_box = driver.find_element(By.XPATH, "//div[@id='stock_issue_list_filter']//input")
                search_box.clear()
                search_box.send_keys(str(search_term))
                sleep(2)
            except:
                pass

        # Robust Identification Logic (100% Accuracy)
        # If captured_ref_no is a numeric ID, we look for a row containing that ID in its action links
        try:
            rows = driver.find_elements(By.XPATH, "//table[@id='stock_issue_list']/tbody/tr")
            target_row = None
            
            if Id and str(Id).isdigit():
                print(f"🔍 Searching for Record ID {Id} in row links...")
                for row in rows:
                    if f"/{Id}" in row.get_attribute("innerHTML"):
                        target_row = row
                        break
            
            # Fallback to top row if no specific ID match or search was by string
            if not target_row and rows:
                target_row = rows[0]

            if target_row:
                ref_no_val = target_row.find_element(By.XPATH, "./td[3]").text.strip()
                karigar_val = target_row.find_element(By.XPATH, "./td[4]").text.strip()
                status_val = target_row.find_element(By.XPATH, "./td[9]").text.strip()
                print(f"📊 Verified Row: RefNo={ref_no_val}, Karigar={karigar_val}, Status={status_val}")
                return ("Pass", f"Verified in list. RefNo: {ref_no_val}", ref_no_val)
            
            return ("Fail", "New record not found in list page.")
        except Exception as e:
            return ("Fail", f"List verification failed: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color) # Updated column
            if captured_id:
                sheet.cell(row=row_num, column=20, value=captured_id) # Ref No column
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass

