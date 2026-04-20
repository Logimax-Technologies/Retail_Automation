from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class ApprovalToInvoice(unittest.TestCase):
    """
    Approval To Invoice Conversion (Supplier Rate Cut) Module Automation
    Handles converting Pure Weight or Amount balance into an Invoice.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_approval_to_invoice(self):
        """Main entry point for Approval To Invoice automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Add Page
        try:
            if "admin_ret_purchase/supplier_rate_cut/add" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Purchase')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Approval To Invoice Conversion')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/supplier_rate_cut/add")
            sleep(2)

        # Read Excel data
        sheet_name = "ApprovalToInvoice"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to reach Excel or sheet: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Mapping based on seeding in SupplierBillEntry.py
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "OpeningBal": 4,
                "RateCutType": 5, "ConvertTo": 6, "ConvType": 7,
                "SupplierName": 8, "RefNo": 9, "Metal": 10,
                "Category": 11, "Product": 12, "PureWeight": 13,
                "Amount": 14, "Rate": 15, "Remark": 16, "ActualStatus": 17
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            sleep(2)
            Function_Call.click(self, "//a[@id='add_Order']")

            # if str(row_data.get("TestStatus")).strip().lower() == "skip":
            #     print(f"⏭️ Skipping Test Case: {row_data['TestCaseId']}")
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                self.driver.refresh()
                sleep(2)
                
                result = self.execute_flow(row_data, row_num, sheet_name)
                
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)

                if result[0] == "Pass":
                    self._seed_approval_rate_fixing(row_num, row_data)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Initial Setup"
        try:
            Function_Call.alert(self)
            
            # --- 1. Opening Balance ---
            if row_data.get("OpeningBal") is not None:
                current_field = "Opening Balance"
                # Excel stores Yes/No text; radio button @value uses 1/0
                op_bal_raw = str(row_data["OpeningBal"]).strip().lower()
                op_bal_val = "1" if op_bal_raw == "yes" else "0"
                Function_Call.click(self, f"//input[@name='supplier_rate_cut[is_opening_blc]' and @value='{op_bal_val}']")
                sleep(1)

            # --- 2. Rate Cut Type ---
            if row_data.get("RateCutType"):
                current_field = "Rate Cut Type"
                # Excel stores text; radio button @value uses 1/2/3
                _rct_map = {"amount": "1", "pure to amount": "2", "amount to pure": "3"}
                rate_cut_val = _rct_map.get(str(row_data["RateCutType"]).strip().lower(), str(row_data["RateCutType"]))
                Function_Call.click(self, f"//input[@name='supplier_rate_cut[rate_cut_type]' and @value='{rate_cut_val}']")
                sleep(1)

            # --- 3. Convert Bill To ---
            if row_data.get("ConvertTo"):
                current_field = "Convert Bill To"
                # Excel stores text; radio button @value uses numeric codes
                _ct_map = {
                    "supplier"         : "1",
                    "manufacturer"     : "2",
                    "stone supplier"   : "3",
                    "diamond supplier" : "4",
                    "approval"         : "5",
                }
                convert_to_val = _ct_map.get(str(row_data["ConvertTo"]).strip().lower(), str(row_data["ConvertTo"]))
                Function_Call.click(self, f"//input[@name='supplier_rate_cut[convert_to]' and @value='{convert_to_val}']")
                sleep(1)

            # --- 4. Conversion Type ---
            if row_data.get("ConvType"):
                current_field = "Conversion Type"
                # Excel stores text; radio button @value uses 1/2
                _conv_map = {"fix": "1", "unfix": "2"}
                conv_type_val = _conv_map.get(str(row_data["ConvType"]).strip().lower(), str(row_data["ConvType"]))
                Function_Call.click(self, f"//input[@name='supplier_rate_cut[conversion_type]' and @value='{conv_type_val}']")
                sleep(1)

            # --- 5. Supplier (Select2) ---
            if row_data.get("SupplierName"):
                current_field = "Supplier"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_karigar-container']", 
                                             str(row_data["SupplierName"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # --- 6. PO No / Opening Ref No (Select2) ---
            if row_data.get("RefNo"):
                current_field = "Reference Number (PO/Op)"
                if str(row_data.get("OpeningBal", "")).strip().lower() == "yes":
                    Function_Call.dropdown_select(self, "//span[@id='select2-opening_ref_no-container']", 
                                                 str(row_data["RefNo"]), 
                                                 '//span[@class="select2-search select2-search--dropdown"]/input')
                else:
                    Function_Call.dropdown_select(self, "//span[@id='select2-select_po_ref_no-container']", 
                                                 str(row_data["RefNo"]), 
                                                 '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # --- 7. Metal/Category/Product (Select2) ---
            if row_data.get("Metal"):
                current_field = "Metal"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_metal-container']", 
                                             str(row_data["Metal"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            if row_data.get("Category"):
                current_field = "Category"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_category-container']", 
                                             str(row_data["Category"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            if row_data.get("Product"):
                current_field = "Product"
                Function_Call.dropdown_select(self, "//span[@id='select2-select_product-container']", 
                                             str(row_data["Product"]), 
                                             '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # --- 8. Weights & Rates ---
            # Normalise RateCutType from Excel text → numeric code (same map as radio button)
            _rct_map = {"amount": "1", "pure to amount": "2", "amount to pure": "3"}
            _rct_raw = str(row_data.get("RateCutType") or "2").strip()
            rate_cut_type = _rct_map.get(_rct_raw.lower(), _rct_raw)
            print("rate_cut_type", rate_cut_type)
            
            if rate_cut_type == "1" or rate_cut_type == "3": # Amount
                if row_data.get("Amount"):
                    current_field = "To Pay Amount"
                    Function_Call.fill_input2(self, "//input[@name='supplier_rate_cut[charges_amount]']", str(row_data["Amount"]))
            
            elif rate_cut_type == "2": # Pure to Amount
                if row_data.get("PureWeight"):
                    current_field = "Pure Weight"
                    Function_Call.fill_input2(self, "//input[@name='supplier_rate_cut[type2_wt]']", str(row_data["PureWeight"]))
            
            # Rate is common for types 2 and 3
            if row_data.get("Rate"):
                current_field = "Rate"
                Function_Call.fill_input2(self, "//input[@name='supplier_rate_cut[src_rate]']", str(row_data["Rate"]))
                sleep(1)

            # --- 9. Remark & Save ---
            current_field = "Remark and Save"
            if row_data.get("Remark"):
                Function_Call.fill_input2(self, "//textarea[@name='supplier_rate_cut[src_remark]']", str(row_data["Remark"]))
            
            # Save Button
            Function_Call.click(self, "//button[@id='supplier_rate_cut_submit']")
            sleep(3)

            # Verify response
            try:
                alert_text = driver.switch_to.alert.text
                driver.switch_to.alert.accept()
                if "Successfully" in alert_text:
                    return ("Pass", alert_text)
                else:
                    return ("Fail", f"Alert: {alert_text}")
            except:
                if "list" in driver.current_url:
                    return ("Pass", "Record saved and redirected to list.")
                else:
                    return ("Fail", "Save clicked but no confirmation received.")

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=16, value=actual_status).font = Font(bold=True, color=color)
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass

    def _seed_approval_rate_fixing(self, row_num, row_data):
        """Seed data into ApprovalRateFixing sheet for the next step"""
        try:
            workbook = load_workbook(FILE_PATH)
            if "ApprovalRateFixing" in workbook.sheetnames:
                rf_sheet = workbook["ApprovalRateFixing"]
                
                # Determine next row and numeric part of TC ID
                last_row = rf_sheet.max_row
                next_num = 1
                
                if last_row > 1:
                    last_tc_val = str(rf_sheet.cell(row=last_row, column=1).value or "")
                    # Extract numeric part from strings like 'TC003'
                    import re
                    match = re.search(r'\d+', last_tc_val)
                    if match:
                        next_num = int(match.group()) + 1
                    else:
                        next_num = last_row # Fallback
                
                new_tc_id = f"TC{next_num:03d}"
                next_row = last_row + 1
                
                # Col 1: TestCaseId, Col 4: Karigar, Col 6: TotalPureWt, Col 9: GST (Always 3%)
                rf_sheet.cell(row=next_row, column=1, value=new_tc_id)
                rf_sheet.cell(row=next_row, column=4, value=row_data.get("SupplierName"))
                rf_sheet.cell(row=next_row, column=6, value=row_data.get("PureWeight"))
                rf_sheet.cell(row=next_row, column=9, value=3)
                
                workbook.save(FILE_PATH)
                print(f"✅ Seeded ApprovalRateFixing sheet at Row {next_row} based on sequence: {new_tc_id}")
            else:
                print(f"⚠️ Sheet 'ApprovalRateFixing' not found in {FILE_PATH}")
            workbook.close()
        except Exception as e:
            print(f"⚠️ Failed to seed ApprovalRateFixing sheet: {e}")
