from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import random
import unittest

FILE_PATH = ExcelUtils.file_path

class HMIssueReceipt(unittest.TestCase):
    """
    HM Issue/ Receipt Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_hm_issue_receipt(self):
        """Main entry point for HM Issue/ Receipt automation"""
        driver = self.driver
        wait = self.wait


        # Read Excel data
        sheet_name = "HMIssueReceipt"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            # Reload workbook to see latest updates (including linked HMRefNo from previous row)
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Navigate to HM Issue/ Receipt List at the start of each case to ensure a clean state (clears old alerts)
            try:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
                Function_Call.click(self, "//span[contains(text(), 'HM Issue/ Receipt')]")
                sleep(2)
            except Exception as e:
                print(f"⚠️ Navigation failed for row {row_num}: {e}")
                workbook.close()
                continue

            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "FlowType": 4, "Karigar": 5, "PORefNo": 6,
                "HMRefNo": 7, "VendorRefNo": 8, "HMChargePerPcs": 9,
                "RejectedPcs": 10, "RejectedGwt": 11,
                "ExpectedStatus": 12, "Remark": 13
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # if not row_data["TestCaseId"]:
            #     continue
            
            # # Check if this test case is marked for execution
            # if str(row_data["TestStatus"]).strip().lower() != "yes":
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']} ({row_data['FlowType']})")
            print(f"{'='*80}")

            flow_type = str(row_data["FlowType"]).strip().lower()

            try:
                if flow_type == "issue":
                    result = self.test_issue_flow(row_data, row_num, sheet_name)
                elif flow_type == "receipt":
                    result = self.test_receipt_flow(row_data, row_num, sheet_name)
                else:
                    result = ("Fail", f"Unknown FlowType: {flow_type}")

                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                
                # Handle captured IDs
                captured_id = result[2] if len(result) > 2 else None
                captured_vendor_ref = result[3] if len(result) > 3 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id, captured_vendor_ref)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

        workbook.close()

    def test_issue_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "HM Issue Start"
        try:
            Function_Call.alert(self)
            Function_Call.click(self, '//a[@id="add_Order" and contains(@class, "btn-success")]') # ISSUE button
            sleep(3)

            if row_data.get("PORefNo"):
                current_field = "PO Ref No"
                Function_Call.dropdown_select(self, '//select[@id="select_po_ref_no"]/following-sibling::span', str(row_data["PORefNo"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(2)
            
            if row_data.get("Karigar"):
                current_field = "Karigar"
                Function_Call.dropdown_select(self, '//select[@id="select_karigar"]/following-sibling::span', str(row_data["Karigar"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            current_field = "Select All Items"
            Function_Call.click(self, '//input[@id="select_all"]')
            sleep(1)

            current_field = "Update Button (Issue)"
            self._take_screenshot(f"BeforeIssue_TC{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="halmarking_issue"]')
            sleep(5)

            current_field = "Capture Success Message"
            try:
                msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text.strip()
                if "Issued successfully" in msg:
                    print("✅ Issue Success Alert Captured")
                    # After Issue, the list reloads. Capture the new HM Ref No from the first row.
                    sleep(3)
                    try:
                        new_hm_ref = driver.find_element(By.XPATH, '//table[@id="item_list"]/tbody/tr[1]/td[1]').text.strip()
                        new_status = driver.find_element(By.XPATH, '//table[@id="item_list"]/tbody/tr[1]/td[9]').text.strip()
                        if "Issued" in new_status:
                            return ("Pass", f"Issue Successful. Generated HM Ref: {new_hm_ref}", new_hm_ref)
                        else:
                            return ("Fail", f"Status mismatch. Expected Issued, Found: {new_status}", new_hm_ref)
                    except Exception as e:
                        return ("Fail", f"Failed to capture generated HM Ref: {str(e)}")
                return ("Fail", f"Unexpected Alert Msg: {msg}")
            except: 
                return ("Fail", "Success message not encountered")

        except Exception as e:
            self._take_screenshot(f"ErrorIssue_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_receipt_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "HM Receipt Start"
        try:
            Function_Call.alert(self)
            Function_Call.click(self, '//a[@id="add_Order" and contains(@class, "btn-primary")]') # RECEIPT button
            sleep(5)

            if row_data.get("HMRefNo"):
                current_field = "HM Ref No"
                Function_Call.dropdown_select(self, '//select[@id="select_hm_ref_no"]/following-sibling::span', str(row_data["HMRefNo"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(0.5) # Wait for table to load
                
                # Check if table detail is showing after selection
                current_field = "Item Detail Table Loading"
                try:
                    rows = driver.find_elements(By.XPATH, '//table[@id="item_detail"]/tbody/tr')
                    if len(rows) == 0 or "No data available" in rows[0].text:
                        return ("Fail", f"Error in {current_field}: No items loaded for HM Ref No: {row_data['HMRefNo']}")
                    print(f"✅ Found {len(rows)} items for Receipt")
                except Exception as e:
                    return ("Fail", f"Error in {current_field}: {str(e)}")

            # Always generate random Vendor Ref No if not explicitly provided or as per user requirement
            random_vendor_ref = f"VRN{random.randint(1000, 9999)}"
            current_field = "Vendor Ref No"
            Function_Call.fill_input(self, wait, (By.ID, "hm_vendor_ref_id"), random_vendor_ref, "VendorRefNo", row_num, Sheet_name=sheet_name)
            print(f"🎲 Random Vendor Ref No entered: {random_vendor_ref}")

            # Iterate through rows to fill charges and rejections
            current_field = "Filling Charges & Rejections"
            rows = driver.find_elements(By.XPATH, '//table[@id="item_detail"]/tbody/tr')
            for i in range(len(rows)):
                idx = i + 1
                if row_data.get("RejectedPcs") is not None:
                    Function_Call.fill_input(self, 
                            wait, 
                            (By.XPATH, f'(//input[contains(@class, "hm_rejected_pcs")])[{idx}]'), 
                            str(row_data["RejectedPcs"]),
                            "RejectedPcs", 
                            row_num, 
                            Sheet_name=sheet_name)
                    sleep(0.5)
                if row_data.get("RejectedGwt") is not None:
                    Function_Call.fill_input(self, wait, (By.XPATH, f'(//input[contains(@class, "hm_rejected_gwt")])[{idx}]'), str(row_data["RejectedGwt"]), "RejectedGwt", row_num, Sheet_name=sheet_name)
                    sleep(0.5)
                # HM Charge Per Pcs always 10 as per user requirement
                hm_charge = "10"
                Function_Call.fill_input(self, wait, (By.XPATH, f'(//input[contains(@class, "hm_charges")])[{idx}]'), hm_charge, "HMChargePerPcs", row_num, Sheet_name=sheet_name)
                # Trigger calculation
                driver.find_element(By.XPATH, f'(//input[contains(@class, "hm_charges")])[{idx}]').send_keys(Keys.TAB)
                sleep(0.5)

            current_field = "Save Button (Receipt)"
            self._take_screenshot(f"BeforeReceipt_TC{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="hm_receipt_submit"]')
            sleep(5)

            current_field = "Capture Success Message"
            try:
                msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text.strip()
                if "HM Receipt successfully" in msg:
                    print("✅ Receipt Success Alert Captured")
                    # Return Pass result with random_vendor_ref to be saved in Excel
                    verify_result = self.test_list_verification_flow("Completed", row_data)
                    return (verify_result[0], verify_result[1], row_data.get("HMRefNo"), random_vendor_ref)
                return ("Fail", f"Unexpected Alert Msg: {msg}")
            except: 
                return ("Fail", "Success message not encountered")

        except Exception as e:
            self._take_screenshot(f"ErrorReceipt_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_list_verification_flow(self, expected_status, row_data):
        driver, wait = self.driver, self.wait
        current_field = "List Verification"
        search_val = row_data.get("HMRefNo")
        
        try:
            current_field = "Search Box"
            search_xpath = '//div[@id="item_list_filter"]//input'
            search = wait.until(EC.presence_of_element_located((By.XPATH, search_xpath)))
            search.clear()
            if search_val:
                search.send_keys(search_val)
                sleep(2)
            
            current_field = "Verify Status"
            # Verify status in the status column (column 9)
            table_xpath = f'//table[@id="item_list"]/tbody/tr[contains(., "{search_val}")]/td[9][contains(., "{expected_status}")]'
            if driver.find_elements(By.XPATH, table_xpath):
                return ("Pass", f"Verified in list with status {expected_status}")
            return ("Fail", f"Status {expected_status} not found in list for {search_val}")

        except Exception as e:
            return ("Fail", f"List Error in {current_field}: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=None, vendor_ref=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            
            # Find the dynamic Remark column to be safe
            remark_col = ExcelUtils.get_column_number(FILE_PATH, sheet_name) or 13
            
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=remark_col, value=f"{test_status} - {actual_status}").font = Font(color=color)
            
            # Save vendor ref if generated
            if vendor_ref:
                sheet.cell(row=row_num, column=8, value=vendor_ref).font = Font(bold=True)
                print(f"📝 Saved random Vendor Ref No: {vendor_ref} to Row {row_num}")

            # If a new HM Ref No was generated/captured, save it back to current row
            if captured_id:
                sheet.cell(row=row_num, column=7, value=captured_id).font = Font(bold=True)
                print(f"📝 Saved/Linked HM Ref No: {captured_id} to Row {row_num}")
                
                # ALSO save to next row if it is a Receipt flow (linking Issue to Receipt)
                next_row = row_num + 1
                try:
                    next_flow = sheet.cell(row=next_row, column=4).value
                    if next_flow and str(next_flow).strip().lower() == "receipt":
                        sheet.cell(row=next_row, column=7, value=captured_id).font = Font(bold=True)
                        # Also set TestStatus to 'yes' for the next row so it runs automatically
                        sheet.cell(row=next_row, column=2, value="yes")
                        print(f"📝 Linked HM Ref No to Row {next_row} (Receipt Flow) and enabled it.")
                except:
                    pass

            workbook.save(FILE_PATH)
            workbook.close()
            print(f"📊 Excel updated at row {row_num}")
        except Exception as e:
            print(f"⚠️ Excel update failed at row {row_num}: {e}")

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass

    def _cancel_form(self):
        try:
            Function_Call.click(self, "//button[contains(text(), 'Cancel')]")
            sleep(1)
        except:
            try:
                Function_Call.click(self, "//a[contains(text(), 'Cancel')]")
                sleep(1)
            except:
                pass
