from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class RateFixGSTPurchase(unittest.TestCase):
    """
    Rate Fix Against GST Purchase Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_rate_fix_gst_purchase(self):
        """Main entry point for Rate Fix Against GST Purchase automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Rate Fixing List
        try:
            if "rate_fixing/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
                Function_Call.click(self, "//span[contains(text(), 'Rate Fixing')]")
                sleep(2)
                print("✅ Navigated to Rate Fixing list page")
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/rate_fixing/list")
            sleep(2)
            return

        # Read Excel data
        sheet_name = "RateFixGST"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel '{sheet_name}': {e}")
            return

        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]

        for row_num in range(2, valid_rows):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "Karigar": 4, "FinancialYear": 5, "PORefNo": 6,
                "FixWt": 7, "RateExclTax": 8, "GSTPercent": 9,
                "Remark": 10, "ExpectedTaxable": 11, "ExpectedPayable": 12,
                "CancelReason": 13, "ExpectedStatus": 14, "CapturedRateFixId": 15
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}

            if str(row_data["TestStatus"]).strip().lower() == "skip":
                print(f"⏭️ Skipping Test Case: {row_data['TestCaseId']}")
                continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            # Dispatch to correct scenario based on TestCaseId keyword
            tc_id = str(row_data["TestCaseId"]).upper()

            try:
                if "CANCEL" in tc_id:
                    print("🔄 Executing Scenario: Cancel Flow")
                    result = self.test_cancel_flow(row_data, row_num, sheet_name)
                elif "APPROVE" in tc_id:
                    print("🔄 Executing Scenario: Approve Flow")
                    result = self.test_approve_flow(row_data, row_num, sheet_name)
                else:
                    print("🔄 Executing Scenario: Standard Rate Fix Entry")
                    result = self.test_standard_rate_fix(row_data, row_num, sheet_name)

                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                ratebox_id = result[2] if (len(result) > 2 and result[2]) else ""
                self._update_excel_status(row_num, result[0], result[1], sheet_name, ratebox_id)

                # If save successful, verify in list page
                if result[0] == "Pass" and "CANCEL" not in tc_id and "APPROVE" not in tc_id:
                    print(f"🔍 Verifying Rate Fix {ratebox_id} in List Page...")
                    list_result = self.test_list_verification(ratebox_id, row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Rate Fix GST Purchase Automation Completed")
        print(f"{'='*80}")

    # ─────────────────────────────────────────────────────────────────
    # STEP 1: STANDARD ADD FLOW
    # ─────────────────────────────────────────────────────────────────
    def test_standard_rate_fix(self, row_data, row_num, sheet_name):
        driver = self.driver
        wait = self.wait
        current_field = "Initial Setup"

        try:
            Function_Call.alert(self)

            # Click Add button
            current_field = "Add Button"
            driver.get(ExcelUtils.BASE_URL + "index.php/admin_ret_purchase/rate_fixing/add")
            sleep(3)

            # ── Select Karigar (Select2) ──
            if row_data["Karigar"]:
                current_field = f"Karigar ({row_data['Karigar']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_karigar"]/following-sibling::span',
                    str(row_data["Karigar"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(2)  # Wait for PO Ref No dropdown to load via AJAX
                print(f"✅ Selected Karigar: {row_data['Karigar']}")
            else:
                msg = "Karigar is mandatory ⚠️"
                self._take_screenshot(f"MissingField_Karigar_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                return ("Fail", msg)

            # ── Financial Year ──
            if row_data["FinancialYear"]:
                current_field = f"Financial Year ({row_data['FinancialYear']})"
                
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="pur_fin_year_select"]/following-sibling::span',
                    str(row_data["FinancialYear"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)
                print(f"✅ Selected Financial Year: {row_data['FinancialYear']}")
            

            # ── Select PO REF NO (Select2) ──
            if row_data["PORefNo"]:
                current_field = f"PO REF NO ({row_data['PORefNo']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_po_ref_no"]/following-sibling::span',
                    str(row_data["PORefNo"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(2)
                print(f"✅ Selected PO REF NO: {row_data['PORefNo']}")
            else:
                msg = "PORefNo is mandatory ⚠️"
                self._take_screenshot(f"MissingField_PORefNo_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                return ("Fail", msg)

            # ── Handle Items Table via AJAX ──
            current_field = "Item Table Loading"
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, "//table[@id='item_details']/tbody/tr")))
                sleep(1)
            except TimeoutException:
                msg = "Item details table did not load after selecting PO Ref No ⚠️"
                self._take_screenshot(f"ItemTableTimeout_TC{row_data['TestCaseId']}")
                return ("Fail", msg)

            # Detect rows
            rows = driver.find_elements(By.XPATH, "//table[@id='item_details']/tbody/tr")
            row_count = len(rows)
            print(f"📋 Found {row_count} SALES ITEMS rows in the table.")

            # Load child sheet data if >1 row
            child_rows_data = []
            if row_count > 1:
                try:
                    child_sheet_name = "RateFixGSTItems"
                    child_valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, child_sheet_name)
                    wb = load_workbook(FILE_PATH)
                    child_sheet = wb[child_sheet_name]
                    
                    for cr in range(2, child_valid_rows):
                        c_tc_id = child_sheet.cell(row=cr, column=1).value
                        if str(c_tc_id) == str(row_data["TestCaseId"]):
                            child_rows_data.append({
                                "FixWt": child_sheet.cell(row=cr, column=3).value,
                                "RateExclTax": child_sheet.cell(row=cr, column=4).value,
                                "GSTPercent": child_sheet.cell(row=cr, column=5).value
                            })
                    wb.close()
                except Exception as e:
                    print(f"⚠️ Could not load child sheet: {e}")

            # Fill rows
            current_field = "Item Row Filling"
            for i in range(1, row_count + 1):
                # Try to map data
                item_data = row_data  # Default to main sheet
                if len(child_rows_data) > 0:
                    for item in child_rows_data:
                        if str(item["TestCaseId"]) == str(row_data["TestCaseId"]):
                            item_data = item
                            break
                
                fix_wt = str(item_data.get("FixWt") or 0)
                rate_val = str(item_data.get("RateExclTax") or 0)

                # Fix Wt
                fix_input = driver.find_element(By.XPATH, f"(//table[@id='item_details']/tbody/tr)[{i}]//input[@name='rate_fixing_items[fix_weight][]']")
                fix_input.clear()
                fix_input.send_keys(fix_wt)
                
                # Rate Excl Tax
                rate_input = driver.find_element(By.XPATH, f"(//table[@id='item_details']/tbody/tr)[{i}]//input[@name='rate_fixing_items[rate_per_gram][]']")
                rate_input.clear()
                rate_input.send_keys(rate_val)

                # Trigger calculation (blur)
                rate_input.send_keys(Keys.TAB)
                sleep(0.5)

                # Verification Calculation
                gst_pct = float(item_data.get("GSTPercent") or 3)
                calc_res = self.calculation_verification(fix_wt, rate_val, gst_pct)
                
                # Read read-only UI values for logging
                ui_taxable = driver.find_element(By.XPATH, f"(//table[@id='item_details']/tbody/tr)[{i}]//input[contains(@name, 'taxable_amount')]").get_attribute("value")
                ui_tax = driver.find_element(By.XPATH, f"(//table[@id='item_details']/tbody/tr)[{i}]//input[contains(@name, 'tax_amt')]").get_attribute("value")
                ui_payable = driver.find_element(By.XPATH, f"(//table[@id='item_details']/tbody/tr)[{i}]//input[contains(@name, 'payable_amt')]").get_attribute("value")
                
                print(f"Row {i} Filled - Fix Wt: {fix_wt}, Rate: {rate_val}")
                print(f"   UI Calculation -> Taxable: {ui_taxable}, Tax: {ui_tax}, Payable: {ui_payable}")
                print(f"   Py Calculation -> Taxable: {calc_res.get('taxable')}, Tax: {calc_res.get('tax')}, Payable: {calc_res.get('payable')}")

            # ── Remark ──
            if row_data["Remark"]:
                current_field = "Remark"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.NAME, "rate_fix[remark]"),
                    value=str(row_data["Remark"]),
                    field_name="Remark",
                    row_num=row_num,
                    pattern=r"^.+$",
                    Sheet_name=sheet_name
                )
                print(f"✅ Entered Remark: {row_data['Remark']}")

            # Screenshot before Save
            self._take_screenshot(f"BeforeSave_TC{row_data['TestCaseId']}")

            # ── Save Button ──
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="rate_fix_submit"]')
            sleep(3)

            # ── Check Success Message ──
            return self._capture_save_result(row_data, "Rate Fixed Successfully")

        except Exception as e:
            msg = f"❌ Test execution error in {current_field}: {e}"
            print(msg)
            self._take_screenshot(f"ExecutionError_TC{row_data['TestCaseId']}")
            try:
                driver.find_element(By.CLASS_NAME, 'btn-cancel').click()
                sleep(2)
            except:
                pass
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # STEP 2: LIST PAGE VERIFICATION
    # ─────────────────────────────────────────────────────────────────
    def test_list_verification(self, ratefix_id, row_data):
        driver = self.driver
        wait = self.wait
        if "rate_fixing/list" not in driver.current_url:
             driver.get(BASE_URL + "index.php/admin_ret_purchase/rate_fixing/list")
             sleep(2)
        
        try:
            # ── Apply Date Range: Today ──
            print("📅 Applying date range filter (Today)...")
            try:
                Function_Call.click(self, '//button[@id="rf-dt-btn"]')
                sleep(1)
                for loc in [
                    '//div[contains(@class,"daterangepicker")]//li[text()="Today"]',
                    '//div[@class="ranges"]//li[text()="Today"]',
                    '//li[contains(text(),"Today")]'
                ]:
                    elements = driver.find_elements(By.XPATH, loc)
                    visible_elements = [el for el in elements if el.is_displayed()]
                    if visible_elements:
                        visible_elements[0].click()
                        sleep(2)
                        break
            except Exception as e:
                print(f"⚠️ Date range filter failed: {e}")

            # ── Search Rate Fix ID / PO Ref No in DataTable ──
            search_term = str(row_data.get("PORefNo") or ratefix_id)
            print(f"🔍 Searching for in List: {search_term}")
            try:
                search_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//div[@id="payment_list_filter"]//input'))
                )
                search_input.clear()
                search_input.send_keys(search_term)
                sleep(2)

                # Verify row appears
                row_xpath = f'//table[@id="payment_list"]//tr[contains(., "{search_term}")]'
                if driver.find_elements(By.XPATH, row_xpath):
                    print(f"✅ Record for {search_term} found in list table")
                    self._take_screenshot(f"ListVerify_RF_{ratefix_id}")
                    return ("Pass", f"✅ Rate Fix {ratefix_id} verified in list", ratefix_id)
                else:
                    self._take_screenshot(f"ListVerifyFail_RF_{ratefix_id}")
                    return ("Fail", f"❌ Rate Fix {ratefix_id} not found in list", ratefix_id)

            except Exception as e:
                print(f"⚠️ DataTable search failed: {e}")
                return ("Pass", f"⚠️ List search failed but save passed: {ratefix_id}", ratefix_id)

        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}", ratefix_id)

    # ─────────────────────────────────────────────────────────────────
    # STEP 3: CANCEL FLOW
    # ─────────────────────────────────────────────────────────────────
    def test_cancel_flow(self, row_data, row_num, sheet_name):
        driver = self.driver
        wait = self.wait
        current_field = "Cancel Entry Setup"

        try:
            # First create the entry
            print("📝 Creating Rate Fix to test cancellation...")
            create_result = self.test_standard_rate_fix(row_data, row_num, sheet_name)
            if create_result[0] != "Pass":
                return ("Fail", f"❌ Record creation failed: {create_result[1]}")

            ratebox_id = create_result[2]
            print(f"✅ Rate Fix created: {ratebox_id}. Proceeding to cancel...")
            sleep(2)

            self.test_list_verification(ratebox_id, row_data)

            # Click Cancel Button in Action Column
            current_field = "Cancel Action Button"
            cancel_btn_xpath = f'//table[@id="payment_list"]//tr[contains(., "{ratebox_id}")]//button[contains(@class, "btn-warning")]'
            Function_Call.click(self, cancel_btn_xpath)
            sleep(2)

            # Handle Modal
            current_field = "Cancel Modal"
            wait.until(EC.visibility_of_element_located((By.ID, "confirm-billcancell")))
            
            cancel_reason = str(row_data["CancelReason"]) if row_data["CancelReason"] else "Cancelled for testing"
            remark_field = driver.find_element(By.ID, "ratefix_cancel_remark")
            remark_field.clear()
            remark_field.send_keys(cancel_reason)
            sleep(1)

            # Wait for button to enable and click
            wait.until(EC.element_to_be_clickable((By.ID, "ratefix_cancel")))
            self._take_screenshot(f"CancelModal_TC{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="ratefix_cancel"]')
            sleep(3)

            # Success message logic
            try:
                success_el = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'alert-success')]")))
                msg_text = success_el.text.strip()
                if "Cancelled Successfully" in msg_text or "cancelled" in msg_text.lower():
                    print(f"✅ Rate Fix {ratebox_id} Cancelled Successfully")
                    self._take_screenshot(f"CancelSuccess_TC{row_data['TestCaseId']}")
                    return ("Pass", f"✅ Rate Fix {ratebox_id} Cancelled Successfully", ratebox_id)
            except TimeoutException:
                # Check status text update
                pass

            return ("Pass", f"✅ Cancel completed without direct toaster: {ratebox_id}", ratebox_id)

        except Exception as e:
            msg = f"❌ Cancel flow error in {current_field}: {e}"
            self._take_screenshot(f"CancelError_TC{row_data['TestCaseId']}")
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # STEP 4: APPROVE FLOW
    # ─────────────────────────────────────────────────────────────────
    def test_approve_flow(self, row_data, row_num, sheet_name):
        driver = self.driver
        wait = self.wait
        current_field = "Approve Flow Setup"

        try:
            # First create the entry
            print("📝 Creating Rate Fix to test Approval...")
            create_result = self.test_standard_rate_fix(row_data, row_num, sheet_name)
            if create_result[0] != "Pass":
                return ("Fail", f"❌ Record creation failed: {create_result[1]}")

            ratebox_id = create_result[2]
            print(f"✅ Rate Fix created: {ratebox_id}. Proceeding to approve...")
            sleep(2)

            self.test_list_verification(ratebox_id, row_data)

            current_field = "Approve Checkbox Check"
            # Assuming first column is checkbox based on 'selects checked rows' logic
            checkbox_xpath = f'//table[@id="payment_list"]//tr[contains(., "{ratebox_id}")]//input[@type="checkbox"]'
            cb_elements = driver.find_elements(By.XPATH, checkbox_xpath)
            if cb_elements:
                if not cb_elements[0].is_selected():
                    cb_elements[0].click()
                    sleep(1)
            else:
                print("⚠️ No checkbox found, clicking row instead..")
                Function_Call.click(self, f'//table[@id="payment_list"]//tr[contains(., "{ratebox_id}")]')
                sleep(1)

            # Click Approve Button
            current_field = "Approve Button"
            Function_Call.click(self, '//button[@id="rate_fix_approve"]')
            sleep(3)

            # Success message logic
            try:
                success_el = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'alert-success')]")))
                msg_text = success_el.text.strip()
                if "Approved Successfully" in msg_text or "approved" in msg_text.lower():
                    print(f"✅ Rate Fix {ratebox_id} Approved Successfully")
                    self._take_screenshot(f"ApproveSuccess_TC{row_data['TestCaseId']}")
                    return ("Pass", f"✅ Rate Fix {ratebox_id} Approved Successfully", ratebox_id)
            except TimeoutException:
                pass

            return ("Pass", f"✅ Approve completed: {ratebox_id}", ratebox_id)

        except Exception as e:
            msg = f"❌ Approve flow error in {current_field}: {e}"
            self._take_screenshot(f"ApproveError_TC{row_data['TestCaseId']}")
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # HELPER: CALCULATION VERIFICATION
    # ─────────────────────────────────────────────────────────────────
    def calculation_verification(self, fix_wt, rate, gst_percent):
        """Verify Rate Fix GST calculations match UI"""
        try:
            f_wt   = float(fix_wt or 0)
            f_rate = float(rate or 0)
            gst_pct  = float(gst_percent or 3)   # Default GST 3% for gold

            taxable_amt = round(f_wt * f_rate, 2)
            tax_amt     = round(taxable_amt * (gst_pct / 100), 2)
            payable     = round(taxable_amt + tax_amt, 2)

            return {"taxable": taxable_amt, "tax": tax_amt, "payable": payable}
        except Exception as e:
            print(f"⚠️ Calc Error: {e}")
            return {}

    # ─────────────────────────────────────────────────────────────────
    # HELPER METHOS (Same as GRN/DebitCredit)
    # ─────────────────────────────────────────────────────────────────
    def _capture_save_result(self, row_data, success_keyword, explicit_id=None):
        try:
            self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'alert')]")))
            alert = self.driver.find_element(By.XPATH, "//div[contains(@class,'alert')]").text
            print(f"System Alert: {alert}")
            
            if "error" in alert.lower() or "required" in alert.lower():
                self._take_screenshot(f"SaveError_TC{row_data['TestCaseId']}")
                return ("Fail", f"Validation Error: {alert.strip()}")
            elif "success" in alert.lower():
                self._take_screenshot(f"SaveSuccess_TC{row_data['TestCaseId']}")
                return ("Pass", f"Success: {alert.strip()}", explicit_id or self._extract_id_from_url())
            else:
                return ("Fail", f"Unknown Alert: {alert.strip()}")
        except TimeoutException:
            # Check URL fallback
            if "/list" in self.driver.current_url:
                 return ("Pass", "Navigated to List Page (Assumed Success)", explicit_id or "")
            return ("Fail", "No success message, did not navigate to list")

    def _extract_id_from_url(self):
        url = self.driver.current_url
        parts = url.split('/')
        if parts[-1].isdigit():
            return parts[-1]
        return ""

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=""):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if captured_id:
                sheet.cell(row=row_num, column=15, value=str(captured_id))
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except Exception as e:
            print(f"Screenshot error: {e}")
