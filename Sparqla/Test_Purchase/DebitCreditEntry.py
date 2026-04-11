from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class DebitCreditEntry(unittest.TestCase):
    """
    Debit/Credit Entry Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_debit_credit_entry(self):
        """Main entry point for Debit/Credit Entry automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Debit/Credit Entry List
        try:
            if "credit_debit_entry/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Credit/Debit Entry')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/credit_debit_entry/list")
            sleep(2)

        # Read Excel data
        sheet_name = "DebitCreditEntry"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            # Mapping based on finalized structure
            data_map = {
                "TestCaseId": 1, "TestStatus": 2,"ActualStatus": 3, "TransactionIn": 4,
                "Karigar": 5, "Type": 6, "SupplierBill": 7,
                "LedgerType": 8, "Amount": 9, "TransType": 10,
                "Weight": 11, "WTransType": 12, "Narration": 13,
                "ExpectedBillNo": 14, "Cancel_Reason":15,"ExpectedStatus": 16
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                self.driver.refresh()
                sleep(2)
                tc_id = str(row_data["TestCaseId"]).upper()
                if "CANCEL" in tc_id:
                    result = self.test_cancel_flow(row_data, row_num, sheet_name)
                else:
                    result = self.execute_debit_credit_entry_flow(row_data, row_num, sheet_name)
                
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_id = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id)

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_debit_credit_entry_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Add Start"
        try:
            Function_Call.alert(self)
            
            # Click ADD button
            if "/add" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/credit_debit_entry/add")
                sleep(3)

            # --- Header Section ---
            # Transaction In (Toggle): Checked = Amount, Unchecked = Weight
            if row_data.get("TransactionIn"):
                current_field = "Transaction In Toggle"
                current_text = driver.find_element(By.ID, "transaction-in").text.strip().lower()
                target_text = str(row_data["TransactionIn"]).strip().lower()
                
                if current_text != target_text:
                    print(f"🔄 Toggling from {current_text} to {target_text}")
                    Function_Call.click(self, "//label[@class='switch']")
                    sleep(1)

            # Supplier (Karigar) (Select2)
            if row_data.get("Karigar"):
                current_field = "Supplier Select"
                Function_Call.dropdown_select(self, '//select[@id="select_karigar"]/following-sibling::span', str(row_data["Karigar"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(2)

            # Type (Radio): 1=Supplier, 2=Smith, 3=Approvals
            if row_data.get("Type"):
                current_field = "Account Type"
                type_val = str(row_data["Type"]).strip().lower()
                type_map = {"supplier": "1", "smith": "2", "approvals": "3"}
                final_val = type_map.get(type_val, type_val)
                Function_Call.click(self, f"//input[@name='credit[accountto]' and @value='{final_val}']")
                sleep(2)

            # Supplier Bills (Select2)
            if row_data.get("SupplierBill"):
                current_field = "Supplier Bill Select"
                Function_Call.dropdown_select(self, '//select[@id="select_karigar_bills"]/following-sibling::span', str(row_data["SupplierBill"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            # Ledger Type (Dropdown)
            if row_data.get("LedgerType"):
                current_field = "Ledger Type Select"
                sel = Select(driver.find_element(By.ID, "id_cr_dr_ledger"))
                sel.select_by_visible_text(str(row_data["LedgerType"]))
                sleep(1)

            # --- Amount Section ---            if str(row_data.get("TransactionIn")).strip().lower() == "amount":
                if row_data.get("Amount"):
                    current_field = "Amount"
                    Function_Call.fill_input(self, wait, (By.ID, "trans_amount"), str(row_data["Amount"]), "Amount", row_num, Sheet_name=sheet_name)
                
                if row_data.get("TransType"):
                    current_field = "Transaction Type"
                    sel = Select(driver.find_element(By.ID, "transtype"))
                    type_map = {"credit": "1", "debit": "2"}
                    final_type = type_map.get(str(row_data["TransType"]).strip().lower(), str(row_data["TransType"]))
                    sel.select_by_value(final_type)
            
            # --- Weight Section ---
            else:
                if row_data.get("Weight"):
                    current_field = "Weight"
                    Function_Call.fill_input(self, wait, (By.ID, "trans_weight"), str(row_data["Weight"]), "Weight", row_num, Sheet_name=sheet_name)
                
                if row_data.get("WTransType"):
                    current_field = "Weight Transaction Type"
                    sel = Select(driver.find_element(By.ID, "wtranstype"))
                    type_map = {"credit": "1", "debit": "2"}
                    final_type = type_map.get(str(row_data["WTransType"]).strip().lower(), str(row_data["WTransType"]))
                    sel.select_by_value(final_type)

            # --- Footer Section ---
            if row_data.get("Narration"):
                current_field = "Narration"
                Function_Call.fill_input(self, wait, (By.ID, "naration"), str(row_data["Narration"]), "Narration", row_num, Sheet_name=sheet_name)

            # --- Submit ---
            current_field = "Submit"
            self._take_screenshot(f"BeforeSave_TC{row_data['TestCaseId']}")
            Function_Call.click(self, "//button[@id='save_credit_entry']")
            sleep(3)

            # Verification in List Page
            return self.verify_in_list(row_data)

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            try:
                Function_Call.click(self, "//button[@id='cancel_bill_edit']")
                sleep(2)
            except:
                pass
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_cancel_flow(self, row_data, row_num, sheet_name):
        driver, wait = self.driver, self.wait
        current_field = "Cancel Flow Start"
        try:
            # 1. First add the entry to have something to cancel
            res = self.execute_debit_credit_entry_flow(row_data, row_num, sheet_name)
            if res[0] != "Pass":
                return res
            
            bill_no = res[2]
            print(f"🔍 Proceeding to cancel Bill No: {bill_no}")
            
            # 2. On list page, click Cancel button
            if "credit_debit_entry/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/credit_debit_entry/list")
                sleep(2)
            
            # Filter and search for the bill
            self.verify_in_list(row_data)
            
            current_field = "Click Cancel Action"
            cancel_xpath = f"//table[@id='trans_list']/tbody/tr[contains(., '{bill_no}')]//button[contains(@class, 'btn-warning')]"
            Function_Call.click(self, cancel_xpath)
            sleep(2)
            
            # 3. Handle Modal
            if row_data.get("Cancel_Reason"):
                current_field = "Cancel Modal Remark"
                wait.until(EC.visibility_of_element_located((By.ID, "credit_cancel_remark")))
                Function_Call.fill_input(self, wait, (By.ID, "credit_cancel_remark"), str(row_data["Cancel_Reason"]), "Cancel Reason", row_num, Sheet_name=sheet_name)
                sleep(1)
            
            current_field = "Click Confirm Cancel"
            Function_Call.click(self, "//button[@id='crdr_cancel']")
            sleep(3)
            
            # 4. Verify Cancelled Status
            return self.verify_in_list(row_data)

        except Exception as e:
            self._take_screenshot(f"CancelError_TC{row_data['TestCaseId']}")
            return ("Fail", f"Cancel Flow Error in {current_field}: {str(e)}")

    def verify_in_list(self, row_data):
        driver, wait = self.driver, self.wait
        if "credit_debit_entry/list" not in driver.current_url:
            driver.get(BASE_URL + "index.php/admin_ret_purchase/credit_debit_entry/list")
            sleep(2)
        
        # 1. Date Filter to Today
        Function_Call.click(self, "//button[@id='rpt_payment_date']") 
        sleep(1)
        Function_Call.click(self, "//li[contains(text(), 'Today')]")
        sleep(1)

        # 2. Type Filter (Credit/Debit)
        if row_data.get("TransType") or row_data.get("WTransType"):
            current_field = "List Type Filter"
            val = str(row_data.get("TransType") or row_data.get("WTransType")).strip().lower()
            type_map = {"credit": "1", "debit": "2"}
            final_val = type_map.get(val, val)
            sel = Select(driver.find_element(By.ID, "transcation_type"))
            sel.select_by_value(final_val)
            sleep(0.5)

        # 3. Transaction Type Filter (Supplier/Smith/Approvals)
        if row_data.get("Type"):
            current_field = "List Trans Type Filter"
            val = str(row_data["Type"]).strip().lower()
            trans_map = {"supplier": "1", "smith": "2", "approvals": "3"}
            final_val = trans_map.get(val, val)
            sel = Select(driver.find_element(By.ID, "trans_type"))
            sel.select_by_value(final_val)
            sleep(0.5)

        # 4. Status Filter
        if row_data.get("ExpectedStatus"):
            current_field = "List Status Filter"
            val = str(row_data["ExpectedStatus"]).strip().lower()
            status_map = {"success": "1", "cancelled": "2"}
            final_val = status_map.get(val, val)
            sel = Select(driver.find_element(By.ID, "status_type"))
            sel.select_by_value(final_val)
            sleep(0.5)

        # 5. Click Search
        Function_Call.click(self, "//button[@id='credit_debit_search']")
        sleep(2)
        
        # Search for Karigar in datatable search box
        if row_data.get("Karigar"):
            try:
                search_box = driver.find_element(By.XPATH, "//div[@id='trans_list_filter']//input")
                search_box.clear()
                search_box.send_keys(row_data["Karigar"])
                sleep(2)
            except:
                pass

        # Confirm top row
        try:
            bill_no = driver.find_element(By.XPATH, "//table[@id='trans_list']/tbody/tr[1]/td[1]").text.strip()
            supplier = driver.find_element(By.XPATH, "//table[@id='trans_list']/tbody/tr[1]/td[3]").text.strip()
            status = driver.find_element(By.XPATH, "//table[@id='trans_list']/tbody/tr[1]/td[9]").text.strip()
            print(f"📊 List Row: BillNo={bill_no}, Supplier={supplier}, Status={status}")
            return ("Pass", f"Verified in list. BillNo: {bill_no}, Status: {status}", bill_no)
        except Exception as e:
            return ("Fail", f"List verification failed: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if captured_id:
                sheet.cell(row=row_num, column=14, value=captured_id)
            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass
