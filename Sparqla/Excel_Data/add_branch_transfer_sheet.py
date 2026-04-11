"""
Script: add_branch_transfer_sheet.py
Purpose: Add 'BranchTransfer' sheet + 'Master' row to the Sqarqla_Retail_data2.xlsx
Run once from terminal: python add_branch_transfer_sheet.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_PATH = r"C:\Users\Dell\Desktop\sqrqlas\Sqarqla_Retail_data2.xlsx"
SHEET_NAME = "BranchTransfer"

# ─── Column headers (must match BranchTransfer.py data_map) ───────────────────
HEADERS = [
    "TestCaseId", "TestStatus", "ActualStatus",
    "TransferType", "FromBranch", "ToBranch", "OtherIssue",
    "LotNo", "Section", "Product", "Design",
    "TagCode", "OldTagCode", "EstimationNo",
    "NT_Receipt",
    "PR_FromDate", "PR_ToDate",
    "Pack_Item", "Pack_Pcs",
    "OrderNo",
    "Remark",
]

# ─── Sample test data rows ─────────────────────────────────────────────────────
# Format: (TestCaseId, TestStatus, ActualStatus, TransferType, FromBranch, ToBranch,
#           OtherIssue, LotNo, Section, Product, Design, TagCode, OldTagCode, EstimationNo,
#           NT_Receipt, PR_FromDate, PR_ToDate, Pack_Item, Pack_Pcs, OrderNo, Remark)
ROWS = [
    # ── TC_BT_001: Tagged — Single Tag Code ───────────────────────────────────
    ("TC_BT_001", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "GBT-01475", "", "",
     "", "", "", "", "", "",
     "Tagged - Single Tag Code"),

    # ── TC_BT_002: Tagged — Multiple Tag Codes (pipe-separated) ───────────────
    ("TC_BT_002", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "GBT-01474|GBT-01476", "", "",
     "", "", "", "", "", "",
     "Tagged - Multi Tag Code: GBT-01474 then GBT-01476"),

    # ── TC_BT_003: Tagged — Multiple Old Tag Codes ────────────────────────────
    ("TC_BT_003", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "OT-001|OT-002", "",
     "", "", "", "", "", "",
     "Tagged - Multi Old Tag Code"),

    # ── TC_BT_004: Tagged — Single Lot No ────────────────────────────────────
    ("TC_BT_004", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "9580", "", "", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Single Lot No 9580"),

    # ── TC_BT_005: Tagged — Multiple Lot Nos (pipe-separated) ─────────────────
    ("TC_BT_005", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "9580|9581", "", "", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Multi Lot No: 9580 then 9581"),

    # ── TC_BT_006: Tagged — Single Section filter ─────────────────────────────
    ("TC_BT_006", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "HARAM GB", "", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Single Section HARAM GB"),

    # ── TC_BT_007: Tagged — Multiple Sections ────────────────────────────────
    ("TC_BT_007", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "HARAM GB|HARAM SECTION", "", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Multi Section: HARAM GB then HARAM SECTION"),

    # ── TC_BT_008: Tagged — Single Product ───────────────────────────────────
    ("TC_BT_008", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "GOLD MALAI", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Single Product GOLD MALAI"),

    # ── TC_BT_009: Tagged — Multiple Products ────────────────────────────────
    ("TC_BT_009", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "GOLD MALAI|SILVER CHAIN", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Multi Product: GOLD MALAI then SILVER CHAIN"),

    # ── TC_BT_010: Tagged — Section + Product (combined filter) ──────────────
    ("TC_BT_010", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "HARAM GB", "GOLD MALAI", "",
     "", "", "",
     "", "", "", "", "", "",
     "Tagged - Section+Product combined (single)"),

    # ── TC_BT_011: Tagged — Single Estimation No ─────────────────────────────
    ("TC_BT_011", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "1",
     "", "", "", "", "", "",
     "Tagged - Single Estimation No 1"),

    # ── TC_BT_012: Tagged — Multiple Estimation Nos ───────────────────────────
    ("TC_BT_012", "Run", "", "Tagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "1|2",
     "", "", "", "", "", "",
     "Tagged - Multi Estimation: 1 then 2"),

    # ── TC_BT_013: Non-Tagged via NT Receipt ─────────────────────────────────
    ("TC_BT_013", "Run", "", "NonTagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "",
     "NT-RCT-001", "", "", "", "", "",
     "NonTagged - NT Receipt"),

    # ── TC_BT_014: Non-Tagged via Product Entry ───────────────────────────────
    ("TC_BT_014", "Run", "", "NonTagged",
     "HEAD OFFICE", "CALICUT", "N",
     "", "SILVER", "KODI", "",
     "", "", "",
     "", "", "", "", "", "",
     "NonTagged - Section + Product"),

    # ── TC_BT_015: Purchase Return ────────────────────────────────────────────
    ("TC_BT_015", "Run", "", "PurchaseReturn",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "",
     "", "01-03-2026", "31-03-2026", "", "", "",
     "Purchase Return Transfer"),

    # ── TC_BT_016: Packing Items ─────────────────────────────────────────────
    ("TC_BT_016", "Run", "", "PackingItems",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "",
     "", "", "", "Box Small", 10, "",
     "Packing Items Transfer"),

    # ── TC_BT_017: Multi-Packing Items (pipe-separated) ──────────────────────
    ("TC_BT_017", "Run", "", "PackingItems",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "",
     "", "", "", "PEN|RING BOX", "2|2", "",
     "Packing Items - Multi: PEN then RING BOX"),

    # ── TC_BT_018: Repair Order ───────────────────────────────────────────────
    ("TC_BT_018", "Run", "", "RepairOrder",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "",
     "", "", "", "", "", "RO-2500",
     "Repair Order Transfer"),

    # ── TC_BT_019: Multi-Repair Order (pipe-separated) ───────────────────────
    ("TC_BT_019", "Run", "", "RepairOrder",
     "HEAD OFFICE", "CALICUT", "N",
     "", "", "", "",
     "", "", "",
     "", "", "", "", "", "ATM25-RE-00019|ATM25-RE-00020",
     "Repair Order - Multi: 00019 then 00020"),

    # ── TC_BT_020: Other Issue (Tagged) ──────────────────────────────────────
    ("TC_BT_020", "Run", "", "Tagged",
     "HEAD OFFICE", "", "Y",
     "", "", "", "",
     "GBT-01477", "", "",
     "", "", "", "", "", "",
     "Other Issue - ToBranch hidden, default HO"),
]

# ─── Styles ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN      = Side(style="thin")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROW_FILLS = [
    PatternFill("solid", fgColor="FFFFFF"),
    PatternFill("solid", fgColor="EBF3FB"),
]

def col_widths():
    return {
        "TestCaseId": 14, "TestStatus": 11, "ActualStatus": 30,
        "TransferType": 16, "FromBranch": 16, "ToBranch": 16, "OtherIssue": 12,
        "LotNo": 14, "Section": 16, "Product": 18, "Design": 16,
        "TagCode": 14, "OldTagCode": 14, "EstimationNo": 16,
        "NT_Receipt": 16,
        "PR_FromDate": 14, "PR_ToDate": 14,
        "Pack_Item": 18, "Pack_Pcs": 11,
        "OrderNo": 14,
        "Remark": 30,
    }


def add_bt_sheet(file_path):
    wb = load_workbook(file_path)

    # ── Remove existing sheet if present ─────────────────────────────────────
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
        print(f"[WARN]  Removed existing '{SHEET_NAME}' sheet.")

    ws = wb.create_sheet(SHEET_NAME)

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = BORDER

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r_idx, row_vals in enumerate(ROWS, start=2):
        fill = ROW_FILLS[(r_idx) % 2]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
            cell.border    = BORDER

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = col_widths()
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 15)

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Update Master sheet ───────────────────────────────────────────────────
    _update_master(wb)

    wb.save(file_path)
    wb.close()
    print(f"[OK] '{SHEET_NAME}' sheet added/updated in {file_path}")
    print(f"[OK] Master sheet updated — BranchTransfer row set to Execution=yes")


def _update_master(wb):
    """Add or update BranchTransfer row in the Master sheet."""
    if "Master" not in wb.sheetnames:
        print("[WARN]  Master sheet not found — skipping master update.")
        return

    ms = wb["Master"]
    # Find existing row or append
    found = False
    for row in ms.iter_rows(min_row=2):
        if str(row[0].value).strip() == SHEET_NAME:
            row[1].value = "yes"   # Execution column
            found = True
            break

    if not found:
        # Find next empty row
        next_row = ms.max_row + 1
        ms.cell(row=next_row, column=1, value=SHEET_NAME)
        ms.cell(row=next_row, column=2, value="yes")
        # Column 3 (Status) left blank
        print(f"   Added new Master row at row {next_row}")
    else:
        print(f"   Updated existing Master row for '{SHEET_NAME}'")


if __name__ == "__main__":
    add_bt_sheet(FILE_PATH)
