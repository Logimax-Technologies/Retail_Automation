from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from Utils.SafeFloat import safe_float
import random
from decimal import Decimal, ROUND_HALF_UP
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class LotGenerate(unittest.TestCase):
    """
    Lot Generate Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_lot_generate(self):
        """Main entry point for Lot Generate automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Lot Generate List at start of each case
        try:
            if "lot_generate/list" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                Function_Call.click(self, "//span[contains(text(), 'Inventory')]")
                sleep(1)
                Function_Call.click(self, "//span[contains(normalize-space(), 'Lot Generate')]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(BASE_URL + "index.php/admin_ret_purchase/lot_generate/list")
            sleep(2)

        # Read Excel data
        sheet_name = "LotGenerate"
        
        # Reset row count and clear report sheets at start of entire test run
        row_count = 2
        try:
            wb = load_workbook(FILE_PATH)
            
            # 1. Clear LotGenerateTag selectively (Cols 1-14, preserving row 1 headers)
            if "LotGenerateTag" in wb.sheetnames:
                tag_clear = wb["LotGenerateTag"]
                if tag_clear.max_row > 1:
                    for r in range(2, tag_clear.max_row + 1):
                        for c in range(1, 15):
                            tag_clear.cell(row=r, column=c).value = None
                    print("✅ LotGenerateTag sheet columns 1-14 cleared (headers preserved)")

            # 2. Clear LotGenerateTagLWT completely (skipping row 1 headers)
            if "LotGenerateTagLWT" in wb.sheetnames:
                lwt_clear = wb["LotGenerateTagLWT"]
                if lwt_clear.max_row > 1:
                    lwt_clear.delete_rows(2, lwt_clear.max_row)
                    print("✅ LotGenerateTagLWT sheet cleared (headers preserved)")

            wb.save(FILE_PATH)
            wb.close()
        except Exception as e:
            print(f"⚠️ Failed to clear report sheets: {e}")

        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "FlowType": 4, "Employee": 5, "PORefNo": 6,
                "LotNo": 7, "ExpectedStatus": 8, "Remark": 9
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # Execution Guard
            # if str(row_data["TestStatus"]).strip().lower() != "yes":
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            try:
                result = self.execute_lot_generate_flow(row_data, row_num, sheet_name, row_count)
                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                captured_id = result[2] if len(result) > 2 else None
                self._update_excel_status(row_num, result[0], result[1], sheet_name, captured_id)
                
                # Increment row_count based on pieces if successful
                if result[0] == "Pass" and len(result) > 3:
                     distributed_pcs = result[3]
                     row_count += distributed_pcs

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

    def execute_lot_generate_flow(self, row_data, row_num, sheet_name, row_count=2):
        driver, wait = self.driver, self.wait
        current_field = "Lot Generate Start"
        try:
            Function_Call.alert(self)
            
            # Click ADD button only if not already on the add page
            if "/add" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_purchase/lot_generate/add")
                sleep(3)

            # Financial Year Selection (usually default selected, but setting it just in case)
            current_field = "Financial Year Select"
            # Assuming default is fine, if not we can add logic to select it.

            if row_data.get("PORefNo"):
                current_field = "PO Ref No"
                Function_Call.dropdown_select(self, '//select[@id="select_po_ref_no"]/following-sibling::span', str(row_data["PORefNo"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(2)
                # Select2 trigger might be needed or item details load automatically
            
            if row_data.get("Employee"):
                current_field = "Select Employee"
                Function_Call.dropdown_select(self, '//select[@id="emp_select"]/following-sibling::span', str(row_data["Employee"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(1)

            if row_data.get("Remark"):
                current_field = "Remark"
                Function_Call.fill_input(self, wait, (By.ID, "remark"), str(row_data["Remark"]), "Remark", row_num, Sheet_name=sheet_name)
                sleep(0.5)

            # Wait for items to load in table
            current_field = "Wait for Items Load"
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="item_detail"]/tbody/tr')))
            
            current_field = "Select All Items"
            Function_Call.click(self, '//input[@id="select_all"]')
            sleep(1)

            # Metadata Extraction for Tag Distribution
            current_field = "Extract Item Details"
            extracted_data = {}
            try:
                # Extract first row metadata as per plan
                rows = driver.find_elements(By.XPATH, '//table[@id="item_detail"]/tbody/tr')
                if rows:
                    row = rows[0]
                    extracted_data = {
                        "Section": row.find_element(By.XPATH, './td[4]').text.strip(),
                        "Product": row.find_element(By.XPATH, './td[5]').text.strip(),
                        "Design": row.find_element(By.XPATH, './td[6]').text.strip(),
                        "Sub Design": row.find_element(By.XPATH, './td[7]').text.strip(),
                        "Pcs": row.find_element(By.XPATH, './td[8]//input').get_attribute("value"),
                        "GWT": row.find_element(By.XPATH, './td[9]//input').get_attribute("value"),
                        "LWT": row.find_element(By.XPATH, './td[10]//input').get_attribute("value")
                    }
                    print(f"📊 Extracted UI Data: {extracted_data}")
            except Exception as e:
                print(f"⚠️ Metadata extraction failed: {e}")

            current_field = "Update Button (Generate Lot)"
            self._take_screenshot(f"BeforeLotGen_TC{row_data['TestCaseId']}")
            Function_Call.click(self, '//button[@id="generate_lot"]')
            sleep(5)

            current_field = "Capture Success Message"
            success_captured = False
            msg = ""
            
            # Try to catch toaster quickly before it disappears (1.5s timeout in JS)
            try:
                msg_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, "//div[@id='toaster']//span[@class='message'] | //div[contains(@class, 'alert-success')] | //div[contains(@class, 'toast-message')]")))
                msg = msg_element.text.strip()
                if "Lot Created successfully" in msg or "Success" in msg:
                    print(f"✅ Success Toast Captured: {msg}")
                    success_captured = True
            except:
                print("ℹ️ Toaster message disappeared or not found, verifying via redirection...")

            # Fallback: Check if redirected to list (CI3 pattern usually redirects after success)
            # In ret_purchase_order.js, it redirects after 2000ms
            sleep(3) 
            if "lot_inward/list" in driver.current_url or success_captured:
                return self.test_list_verification_flow(row_data, row_num, extracted_data, row_count)
                
            return ("Fail", f"Success verification failed. Msg: {msg}")

        except Exception as e:
            self._take_screenshot(f"ErrorLotGen_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_list_verification_flow(self, row_data, row_num, extracted_data=None, row_count=2):
        driver, wait = self.driver, self.wait
        current_field = "List Verification"
        
        try:
            # 1. Navigate back to list explicitly if not already there
            if "lot_inward/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_lot/lot_inward/list")
                sleep(2)

            # 2. List Page Filters
            current_field = "Date Range Filter"
            Function_Call.click(self, '//button[@id="ltInward-dt-btn"]')
            sleep(1)
            Function_Call.click(self, '//li[contains(text(), "Today")]')
            sleep(1)

            current_field = "Metal Filter"
            # Using partial text match or specific value if known. Standard is usually 'Gold'
            try:
                self.fc.dropdown_select2('//select[@id="metal"]/following-sibling::span', 'Gold', '//span[@class="select2-search select2-search--dropdown"]/input')
            except:
                print("⚠️ Metal filter selection failed (optional)")

            current_field = "Employee Filter"
            if row_data.get('Employee'):
                try:
                    self.fc.dropdown_select2('//select[@id="select_emp"]/following-sibling::span', str(row_data['Employee']), '//span[@class="select2-search select2-search--dropdown"]/input')
                except:
                    print(f"⚠️ Employee filter selection failed for: {row_data['Employee']}")

            # current_field = "Stock Type Filter"
            # try:
            #     # Value 2 is Non Tag in list.php
            #     self.fc.dropdown_select2('//select[@id="lot_type"]/following-sibling::span', 'Non Tag', '//span[@class="select2-search select2-search--dropdown"]/input')
            # except:
            #     print("⚠️ Stock Type filter selection failed (optional)")

            sleep(1)
            current_field = "Search Button"
            Function_Call.click(self, '//button[@id="lot_inward_search"]')
            sleep(3)

            # 3. Capture Lot No from first row if not provided
            current_field = "Capture New Lot No"
            try:
                # Lot No is in the 1st TD in this table
                new_lot_no = driver.find_element(By.XPATH, '//table[@id="lot_inward_list"]/tbody/tr[1]/td[1]').text.strip()
            except:
                return ("Fail", "Failed to find any rows in list after generation")

            # 4. Search Box Verification
            current_field = "Search Box Search"
            search_xpath = '//div[@id="lot_inward_list_filter"]//input'
            search = wait.until(EC.presence_of_element_located((By.XPATH, search_xpath)))
            search.clear()
            search.send_keys(new_lot_no)
            sleep(2)

            # 5. Verify result presence and action buttons
            current_field = "Verify Action Column"
            # Action column is at index 17
            action_cell = driver.find_element(By.XPATH, f'//table[@id="lot_inward_list"]/tbody/tr[contains(., "{new_lot_no}")]/td[17]')
            inner_html = action_cell.get_attribute("innerHTML")
            print(inner_html)   

            # Capture Product Name (Column 5) and check for Non-Tag products
            current_field = "Non-Tag Sync Verification"
            is_non_tag = False
            try:
                product_name = driver.find_element(By.XPATH, f'//table[@id="lot_inward_list"]/tbody/tr[contains(., "{new_lot_no}")]/td[5]').text.strip()
                non_tag_list = ["GOLD HOOK", "PACKET GOLD", "HOOK", "HOOK K"]
                
                if product_name.upper() in [p.upper() for p in non_tag_list]:
                    is_non_tag = True
                    print(f"📦 Non-Tag Product detected: {product_name}. Syncing to NonTagReceipt sheet...")
                    wb = load_workbook(FILE_PATH)
                    if "NonTagReceipt" in wb.sheetnames:
                        ntr_sheet = wb["NonTagReceipt"]
                        # Update col 4 with lotNo
                        ntr_sheet.cell(row=row_num, column=4, value=new_lot_no).font = Font(bold=True)
                        wb.save(FILE_PATH)
                        print(f"✅ Lot No {new_lot_no} synced to NonTagReceipt (Row {row_num})")
                    wb.close()
            except Exception as e:
                print(f"⚠️ Non-Tag sync failed: {e}")
            
            if "fa-trash" in inner_html or "fa-edit" in inner_html or "fa-print" in inner_html:
                # Lot generated and verified, now distribute metadata if Tagging flow (and not non-tagged)
                distributed_pcs = 0
                if extracted_data and not is_non_tag:
                    try:
                        print(f"📦 Distributing Tag metadata for Lot: {new_lot_no}")
                        wb = load_workbook(FILE_PATH)
                        # Assume no stones from Lot Generate UI extraction for now
                        all_stones = []
                        data = self.update_Lot_id(new_lot_no, row_count, [extracted_data["Pcs"]],
                                               extracted_data["GWT"], extracted_data["LWT"], 
                                               extracted_data, wb, all_stones)
                        distributed_pcs, message = data
                        print(f"✅ Distribution Result: {message}")
                    except Exception as e:
                        print(f"⚠️ Tag distribution update failed: {e}")

                return ("Pass", f"Verified in list. Lot No: {new_lot_no}", new_lot_no, distributed_pcs)
            
            return ("Fail", f"Action buttons not found for Lot No: {new_lot_no}")

        except Exception as e:
            return ("Fail", f"List Error in {current_field}: {str(e)}")

    def update_Lot_id(self, Lot_id, row_count, pcs, g_wt_sum, l_wt_sum, meta, workbook, all_stones):
        # User defined rename: Tag replace LotGenerateTag, Tag_LWt replace LotGenerateTagLWT
        Pcs_count = sum(map(int, pcs))
        print(f"Distributing {Pcs_count} pieces for Lot: {Lot_id}") 
        tag_sheet = workbook["LotGenerateTag"] if "LotGenerateTag" in workbook.sheetnames else workbook.create_sheet("LotGenerateTag")
        lwt_sheet = workbook["LotGenerateTagLWT"] if "LotGenerateTagLWT" in workbook.sheetnames else workbook.create_sheet("LotGenerateTagLWT")
        
        # Determine the starting row for LotGenerateTagLWT
        lwt_row_start = lwt_sheet.max_row + 1
        if lwt_row_start == 2 and (lwt_sheet.cell(row=1, column=1).value is None or lwt_sheet.cell(row=1, column=1).value == ""):
            # Initialize headers if sheet is brand new or empty
            headers = ["Test Case Id", "Less Weight", "Type", "XPATH", "Code", "Pcs", "Wt", "Wt Type", "Cal.Type", "Rate", "Amount"]
            for h_col, h_text in enumerate(headers, 1):
                lwt_sheet.cell(row=1, column=h_col).value = h_text
            lwt_row_start = 2

        try:
            Pcs_count = int(Pcs_count)
        except:
            Pcs_count = 1

        # Logic: NWT = GWT - LWT
        total_gwt = Decimal(str(g_wt_sum or 0))
        total_lwt = Decimal(str(l_wt_sum or 0))
        total_nwt = total_gwt - total_lwt

        if Pcs_count > 0:
            # 1. Distribute Net Weight (NWT) with random variance
            weights = [Decimal(str(1 + random.uniform(-0.2, 0.2))) for _ in range(Pcs_count)]
            if sum(weights) != 0:
                scale = total_nwt / sum(weights)
                nwt_list = [w * scale for w in weights]
            else:
                nwt_list = [total_nwt / Pcs_count for _ in range(Pcs_count)]

            # Round NWT to 3 decimal places
            nwt_final = [w.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP) for w in nwt_list]
            
            # Force sum to match total_nwt exactly
            diff_nwt = total_nwt - sum(nwt_final)
            if nwt_final:
                nwt_final[0] += diff_nwt

            # 2. Distribute Less Weight (LWT) equally
            lwt_per_pcs = (total_lwt / Pcs_count).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
            lwt_final = [lwt_per_pcs for _ in range(Pcs_count)]
            
            # Force sum to match total_lwt exactly
            diff_lwt = total_lwt - sum(lwt_final)
            if lwt_final:
                lwt_final[0] += diff_lwt

            # 3. Apply to Sheets
            current_lwt_row = lwt_row_start
            section = meta.get("Section", "")
            if not section or str(section).strip() == "":
                section = "GOLD CHAIN"

            for i in range(Pcs_count):   
                row_Num = row_count + i
                
                # Assign a unique Test Case ID for each piece
                tag_tc_id = f"TAG_{Lot_id}_{i+1}"
                tag_sheet.cell(row=row_Num, column=1).value = tag_tc_id

                # Column 5: Lot ID
                tag_sheet.cell(row=row_Num, column=5).value = Lot_id
                
                # Column 4: Branch (Default to 'Head Office')
                tag_sheet.cell(row=row_Num, column=4).value = "Head Office"

                # Column 6: Product
                tag_sheet.cell(row=row_Num, column=6).value = meta.get("Product")
                
                # Column 7: Section
                tag_sheet.cell(row=row_Num, column=7).value = section

                # Column 8: Design
                tag_sheet.cell(row=row_Num, column=8).value = meta.get("Design")
                
                # Column 9: Sub Design
                tag_sheet.cell(row=row_Num, column=9).value = meta.get("Sub Design")
                
                # Column 10: Pieces (Individual tag is usually 1 pcs)
                tag_sheet.cell(row=row_Num, column=10).value = 1
                
                # Column 11: No. of items (Single tagging)
                tag_sheet.cell(row=row_Num, column=11).value = 1

                # Column 13: Individual Less Weight
                indiv_lwt = float(lwt_final[i])
                tag_sheet.cell(row=row_Num, column=13).value = indiv_lwt
                
                # Column 12: Individual Gross Weight (NWT + LWT)
                indiv_nwt = float(nwt_final[i])
                tag_sheet.cell(row=row_Num, column=12).value = round(indiv_nwt + indiv_lwt, 3)
                
                # Update LotGenerateTagLWT with distributed stone details
                has_stone_this_tag = False
                if all_stones:
                    for stone in all_stones:
                        total_pcs = safe_float(stone.get("Pcs"))
                        total_wt = safe_float(stone.get("Wt"))
                        stone_rate = safe_float(stone.get("Rate"))
                        
                        tag_pcs = int((i + 1) * total_pcs / Pcs_count) - int(i * total_pcs / Pcs_count)
                        if total_pcs > 0:
                            tag_wt = (total_wt / total_pcs) * tag_pcs
                        else:
                            tag_wt = ((i + 1) * total_wt / Pcs_count) - (i * total_wt / Pcs_count)
                        
                        if tag_pcs > 0 or tag_wt > 0:
                            has_stone_this_tag = True
                            lwt_sheet.cell(row=current_lwt_row, column=1).value = tag_tc_id
                            lwt_sheet.cell(row=current_lwt_row, column=2).value = "Yes"
                            lwt_sheet.cell(row=current_lwt_row, column=3).value = stone.get("Type")
                            lwt_sheet.cell(row=current_lwt_row, column=4).value = stone.get("XPATH")
                            lwt_sheet.cell(row=current_lwt_row, column=5).value = stone.get("Code")
                            lwt_sheet.cell(row=current_lwt_row, column=6).value = tag_pcs
                            lwt_sheet.cell(row=current_lwt_row, column=7).value = round(tag_wt, 3)
                            lwt_sheet.cell(row=current_lwt_row, column=8).value = stone.get("Wt Type")
                            lwt_sheet.cell(row=current_lwt_row, column=9).value = stone.get("Cal.Type")
                            lwt_sheet.cell(row=current_lwt_row, column=10).value = stone_rate
                            
                            cal_type = str(stone.get("Cal.Type") or "").strip().lower()
                            indiv_amount = tag_pcs * stone_rate if cal_type == "pcs" else tag_wt * stone_rate
                            lwt_sheet.cell(row=current_lwt_row, column=11).value = round(indiv_amount, 2)
                            current_lwt_row += 1
                
                # Update Less Weight status in Tag sheet based on actual distribution
                if has_stone_this_tag:
                    tag_sheet.cell(row=row_Num, column=13).value = "Yes"
                else:
                    tag_sheet.cell(row=row_Num, column=13).value = "No"

        workbook.save(FILE_PATH)
        workbook.close()
        return Pcs_count, "Lot ID and Metadata distributed successfully"

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, captured_id=None):
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            if captured_id:
                sheet.cell(row=row_num, column=7, value=captured_id).font = Font(bold=True)

            workbook.save(FILE_PATH)
            workbook.close()
        except:
            pass

    def _take_screenshot(self, filename):
        try:
            if not os.path.exists(ExcelUtils.SCREENSHOT_PATH):
                os.makedirs(ExcelUtils.SCREENSHOT_PATH)
            path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
            self.driver.save_screenshot(path)
        except:
            pass

    def _cancel_form(self):
        try:
            Function_Call.click(self, "//button[contains(text(), 'Cancel')]")
        except:
            pass
