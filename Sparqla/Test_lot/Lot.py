# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Test_lot.othermetal import Othermetal
from Test_lot.Stone import Stone
import win32com.client as win32
import re
from time import sleep
import os
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
import random
from decimal import Decimal, ROUND_HALF_UP
from Utils.SafeFloat import safe_float
from time import sleep

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL
class Lot(unittest.TestCase):  
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def _handle_print_tab(self, main_window):
        """Helper to close print tab and return focus to main window while capturing lot ID from URL"""
        sleep(5)  # Wait for print tab to open
        lot_id = None
        try:
            windows = self.driver.window_handles
            if len(windows) > 1:
                # Switch to the last opened tab (the print tab)
                self.driver.switch_to.window(windows[-1])
                current_url = self.driver.current_url
                print(f"DEBUG: Print Tab URL: {current_url}")
                
                # ID is at the end of URL: .../admin_ret_lot/lot_acknowledgement/1/9609
                match = re.search(r"/(\d+)$", current_url)
                if match:
                    lot_id = match.group(1)
                    print(f"🔍 Captured Lot ID from URL: {lot_id}")
                
                self.driver.close()
                self.driver.switch_to.window(main_window)
            else:
                print("⚠️ Print tab not opened found")
        except Exception as e:
            print(f"⚠️ Print tab handling failed: {e}")
            try:
                self.driver.switch_to.window(main_window)
            except: pass
        return lot_id

    def test_lot(self):
        driver = self.driver
        wait = self.wait
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass

            Function_Call.click(self, "//span[contains(text(), 'Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Lot Inward')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_lot/lot_inward/list")
            sleep(2)
        # Moved add_lot click to create() method for each new lot iteration
        function_name = "Lot"

        # Clear BranchTransferApproval sheet at start of entire test run
        try:
            wb = load_workbook(FILE_PATH)
            # Clear BranchTransfer sheet (except Row 1 and Column 6 as requested)
            if "BranchTransfer" in wb.sheetnames:
                bt_clear = wb["BranchTransfer"]
                if bt_clear.max_row > 1:
                    for r in range(2, bt_clear.max_row + 1):
                        for c in range(1, bt_clear.max_column + 1):
                            if c != 6:
                                bt_clear.cell(row=r, column=c).value = None
                wb.save(FILE_PATH)
            wb.close()
            print("✅ BranchTransfer sheet cleared (headers preserved)")
        except Exception as e:
            print(f"⚠️ Failed to clear BranchTransferApproval: {e}")

        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        # Call the function
        Lot=ExcelUtils.Lot_details(FILE_PATH, function_name)
        Window=1
        Products=[]
        beforelist=''
        
        # Always start from Row 2 to overwrite tags while keeping static columns (like Branch, Section)
        row_count = 2

        # Only clear Tag_LWt sheet before starting new Lot processing (only headers preserved)
        if "Tag_LWt" in workbook.sheetnames:
            lwt_clear = workbook["Tag_LWt"]
            if lwt_clear.max_row > 1:
                lwt_clear.delete_rows(2, lwt_clear.max_row)
        pcs=[0]
        gwt_total=[0.0]
        lwt_total=[0.0]
        all_stones = []
        meta = {"Product": "", "Design": "", "Sub Design": "", "Lot Received": "", "Section": ""}
        
        for row_num in range(2, valid_rows):   
            # Define columns and dynamically fetch their values   
                data = {
                        "Test Case Id": 1,
                        "Test Status": 2,
                        "Actual Status": 3,
                        "Lot": 4,
                        "Lot Received": 5,
                        "Smith": 6,
                        "StockType": 7,
                        "Category": 8,
                        "Purity": 9,
                        "Section": 10,
                        "Product": 11,
                        "Design": 12,
                        "Sub Design": 13,
                        "Pcs": 14,
                        "GWT": 15,
                        "LWT": 16,           # Adjust if actual LWT column is different
                        "Other metal": 17,   # shifted to match correct position
                        "Charge Name": 18,
                        "Type": 19,
                        "Charge": 20,
                        "Purchase MC": 21,
                        "Purchase MC Type": 22,
                        "Purchase Wastage": 23,
                        "Purchase Rate": 24,
                        "Purchase Rate Type": 25,
                        "Metal Type": 26,
                        "Employee": 27
                } 
                row_data = {key: sheet.cell(row=row_num, column=col).value 
                                for key, (col) in data.items()}
                print(row_data)

                print(f"\n{'='*80}")
                print(f"🧪 Running Test Case: {row_data['Test Case Id']}")
                print(f"{'='*80}")
                
                lot_no = str(row_data.get("Lot") or "")
                row_lot_data = self.Lotdetails(row_data["Test Case Id"]) 
                row_no=row_num+1
                Next_Lot = sheet.cell(row=row_no, column=4).value  # Column 4 = Lot
                
                Create_data=self.create(row_data,lot_no,Next_Lot,row_num,beforelist,Window,Products)
                print(Create_data)
                if Create_data and len(Create_data) == 5:
                    lot, pcs_count, Product, lwt_from_create, stone_list = Create_data
                    if lot == lot_no:
                        pcs[0] = pcs[0] + int(pcs_count)
                        gwt_total[0] = gwt_total[0] + safe_float(row_data.get("GWT"))
                        lwt_total[0] = lwt_total[0] + safe_float(lwt_from_create)
                        if stone_list:
                            all_stones.extend(stone_list)
                        meta["Product"] = str(row_data.get("Product") or "")
                        meta["Design"] = str(row_data.get("Design") or "")
                        meta["Sub Design"] = str(row_data.get("Sub Design") or "")
                        meta["Lot Received"] = str(row_data.get("Lot Received") or "")
                        meta["Section"] = str(row_data.get("Section") or "")

                        beforelist = lot_no
                        Products.append(Product)
                        print(beforelist)
                        continue
                elif Create_data and len(Create_data) >= 5:
                    if len(Create_data) == 6:
                        Test_Status, Actual_Status, Lot_id, pcs_count, lwt_from_create, stone_list = Create_data
                    else: # Fallback for loop continue cases
                        lot, pcs_count, product, lwt_from_create, stone_list = Create_data
                    
                    pcs[0] = pcs[0] + int(pcs_count)
                    lwt_total[0] = lwt_total[0] + safe_float(lwt_from_create)
                    gwt_total[0] = gwt_total[0] + safe_float(row_data.get("GWT"))
                    if stone_list:
                        all_stones.extend(stone_list)
                    
                    meta["Product"] = str(row_data.get("Product") or "")
                    meta["Design"] = str(row_data.get("Design") or "")
                    meta["Sub Design"] = str(row_data.get("Sub Design") or "")
                    meta["Lot Received"] = str(row_data.get("Lot Received") or "")
                    meta["Section"] = str(row_data.get("Section") or "")

                    
                    print(f"Lot ID: {Lot_id if 'Lot_id' in locals() else 'New'}")
                    sheet.cell(row=row_num, column=2).value = Test_Status
                    sheet.cell(row=row_num, column=3).value = Actual_Status
                    workbook.save(FILE_PATH)
                    
                    if Products:
                        Products.clear()
                    Status = ExcelUtils.get_Status(FILE_PATH, function_name)  
                    print(Status)   
                    if row_data["StockType"] == "Tagged":         
                        data = self.update_Lot_id(Lot_id, row_count, pcs, gwt_total[0], lwt_total[0], meta, workbook, all_stones, Test_Status, Actual_Status)

                        pcs_count_val, message = data
                        row_count = pcs_count_val + row_count 
                        print(row_count)  
                        print(message)  
                    elif str(row_data["StockType"]).strip().lower() in ["non-tagged", "non tagged"]:
                        self.update_BranchTransfer(row_data, workbook, row_num)
                        self.update_NonTag_Detail(row_data, workbook, row_num)

                    pcs[0] = 0
                    gwt_total[0] = 0.0
                    lwt_total[0] = 0.0
                    all_stones.clear()
                    print(pcs)
                    Update_master = ExcelUtils.update_master_status(FILE_PATH, Status, function_name) 
                    
                    # Re-open workbook for the next iteration since update_Lot_id or other calls might have closed it
                    workbook = load_workbook(FILE_PATH)
                    sheet = workbook[function_name]
                    continue
        
    def create(self,row_data,lot_no,Next_Lot,row_num,beforelist,Window,Products): 
        driver = self.driver
        wait = self.wait    
        if beforelist != lot_no:
            # Ensure no previous alerts are blocking
            try:
                alert = driver.switch_to.alert
                print(f"⚠️ Unexpected alert found: {alert.text}")
                alert.accept()
                sleep(1)
            except:
                pass

            # Open Add Lot form for new lot if not already there
            if "/index.php/admin_ret_lot/lot_inward/add" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.ID,"add_lot"))).click()
                sleep(2)
            print(f"🚀 Started Lot Processing: {lot_no}")

            Function_Call.dropdown_select(
                self,'//span[@id="select2-lt_rcvd_branch_sel-container"]', 
                row_data["Lot Received"],"//input[@type='search']"
                )
            print("yes1")
            wait.until(EC.visibility_of_element_located((By.XPATH,"//span[@id='select2-lt_gold_smith-container']/span"))).click()
            wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']"))).clear()
            wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']"))).send_keys(row_data["Smith"], Keys.ENTER)
            
            print(row_data["StockType"])
            if row_data["StockType"]=="Tagged" :
                wait.until(EC.element_to_be_clickable((By.XPATH,"//form[@id='lot_form']/div/div/div[3]/div/input[1]"))).click()
            else: 
                wait.until(EC.element_to_be_clickable((By.XPATH,"//form[@id='lot_form']/div/div/div[3]/div/input[2]"))).click()
            
            # Select Category (Select2)
            Function_Call.dropdown_select2(self, '//span[@id="select2-category-container"]/following-sibling::span', 
                                    str(row_data["Category"]), 
                                    '//span[@class="select2-search select2-search--dropdown"]/input')
            
            # Select Purity (Select2)
            Function_Call.dropdown_select2(self, '//span[@id="select2-purity-container"]/following-sibling::span', 
                                    str(row_data["Purity"]), 
                                    '//span[@class="select2-search select2-search--dropdown"]/input')
            sleep(1)
        else:
            print("Same to lot")    
        if row_data["StockType"]=="Non-Tagged" :
            print(f"Selecting Section: {row_data['Section']}")
            Function_Call.dropdown_select2(self, '//span[@id="select2-select_section-container"]/following-sibling::span', 
                                    str(row_data["Section"]), 
                                    '//span[@class="select2-search select2-search--dropdown"]/input')
        else: 
            print("Tagged Items (Skipping Section selection as per logic)")

        # Select Product
        print(f"Selecting Product: {row_data['Product']}")
        Function_Call.dropdown_select2(self, '//span[@id="select2-select_product-container"]/following-sibling::span', 
                                    str(row_data["Product"]), 
                                    '//span[@class="select2-search select2-search--dropdown"]/input')
        sleep(2)

        # Select Design
        print(f"Selecting Design: {row_data['Design']}")
        Function_Call.dropdown_select2(self, '//span[@id="select2-select_design-container"]/following-sibling::span', 
                                    str(row_data["Design"]), 
                                    '//span[@class="select2-search select2-search--dropdown"]/input')
        sleep(2)

        # Select Sub Design
        print(f"Selecting Sub Design: {row_data['Sub Design']}")
        Function_Call.dropdown_select2(self, '//span[@id="select2-select_sub_design-container"]/following-sibling::span', 
                                    str(row_data["Sub Design"]), 
                                    '//span[@class="select2-search select2-search--dropdown"]/input')
        sleep(2)

        wait.until(EC.visibility_of_element_located((By.ID,"lot_pcs"))).click()
        wait.until(EC.visibility_of_element_located((By.ID,"lot_pcs"))).clear()
        wait.until(EC.visibility_of_element_located((By.ID,"lot_pcs"))).send_keys(row_data["Pcs"])
        wait.until(EC.visibility_of_element_located((By.ID,"lot_gross_wt"))).click()
        wait.until(EC.visibility_of_element_located((By.ID,"lot_gross_wt"))).clear()
        wait.until(EC.visibility_of_element_located((By.ID,"lot_gross_wt"))).send_keys(row_data["GWT"])
        print(row_data["Type"])
        test_case_id =row_data["Test Case Id"]
        if row_data["LWT"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@class="input-group-addon input-sm add_tag_lwt"]'))).click()
            Sheet_name = "Lot_Lwt"
            LessWeight=Stone.test_tagStone(self,Sheet_name,test_case_id)
            print(LessWeight)
            Lwt,Wt_gram,TotalAmount,stone_list=LessWeight
            print(LessWeight)
            print(Lwt)
            print(Wt_gram)
            print(TotalAmount)
            print(f"Stones collected: {len(stone_list)}")
        else:
            Wt_gram = 0.0
            stone_list = []
            print("There is no Less Weight in this product")
        if row_data["Other metal"]=="Yes":
                wait.until(EC.element_to_be_clickable((By.ID,"other_metal_amount"))).click()
                Sheet_name = "Lot_othermetal"
                Data=Othermetal.test_othermetal(self,Sheet_name,test_case_id)
                OtherMetal,OtherMetalAmount =Data
                print(OtherMetal)
                print(OtherMetalAmount)
        else:
                print("There is no Other Metal in this product")    
        
        # Check for any validation alerts before clicking Charges (to handle "Please Fill The Required Details")
        try:
            alert = driver.switch_to.alert
            alert_text = alert.text
            print(f"⚠️ Validation alert before Charges: {alert_text}")
            alert.accept()
            sleep(1)
            # If we hit an alert here, it means some field above failed. 
            # We'll log it and try to proceed, but it might fail the next step.
        except:
            # No alert, which is good
            pass

        wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@id='item_details']/div[2]/div/div[6]/div/div/span"))).click() 
        wait.until(EC.element_to_be_clickable((By.XPATH,"//table[@id='table_charges']/tbody/tr/td[2]/select"))).click()
        sleep(2)
        print(row_data["Charge Name"])
        Select(wait.until(EC.element_to_be_clickable((By.XPATH,"//table[@id='table_charges']/tbody/tr/td[2]/select")))).select_by_visible_text(row_data["Charge Name"])
        wait.until(EC.element_to_be_clickable((By.XPATH,"//table[@id='table_charges']/tbody/tr/td[3]/select"))).click()
        Select( wait.until(EC.element_to_be_clickable((By.XPATH,"//table[@id='table_charges']/tbody/tr/td[3]/select")))).select_by_visible_text(row_data["Type"])
        wait.until(EC.element_to_be_clickable((By.ID,"update_charge_details"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"mc_value"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"mc_value"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"mc_value"))).send_keys(row_data["Purchase MC"])
        wait.until(EC.element_to_be_clickable((By.XPATH,"//form[@id='lot_form']/div/div[3]/div/div"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"mc_type"))).click()
        Select(wait.until(EC.element_to_be_clickable((By.ID,"mc_type")))).select_by_visible_text(row_data["Purchase MC Type"])
        wait.until(EC.element_to_be_clickable((By.ID,"lot_wastage"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"lot_wastage"))).send_keys(row_data["Purchase Wastage"])
        wait.until(EC.element_to_be_clickable((By.ID,"rate_per_gram"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"rate_per_gram"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"rate_per_gram"))).send_keys(row_data["Purchase Rate"])
        wait.until(EC.element_to_be_clickable((By.ID,"rate_calc_type"))).click()
        Select(wait.until(EC.element_to_be_clickable((By.ID,"rate_calc_type")))).select_by_visible_text(row_data["Purchase Rate Type"])
        wait.until(EC.element_to_be_clickable((By.ID,"add_lot_items"))).click()
        if Next_Lot == lot_no: 
            Product=row_data["Product"]
            pcs=row_data["Pcs"]
            return lot_no,pcs,Product,Wt_gram,stone_list
        else:    
            main_window = self.driver.current_window_handle
            wait.until(EC.element_to_be_clickable((By.ID,"save_all"))).click()
            
            # Capture Lot ID from the newly opened print tab
            captured_lot_id = self._handle_print_tab(main_window)
            
            sleep(2)
            try:
                if captured_lot_id:                   
                    message = "Add Lot! Lot added successfully"
                    driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, 'Lot.png.png'))
                    Test_Status="Pass"
                    Actual_Status= message
                else:
                    message = "Add Lot! Lot not added successfully"
                    Test_Status="Fail"
                    Actual_Status= message
            except:
                driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, 'Loterror.png.png'))
                Test_Status="Fail"
                Actual_Status="Lot Not Add Successfully"     
            
            # Use base_url for fast navigation back to the list page
            if "/admin_ret_lot/lot_inward/list" not in driver.current_url:
                driver.get(BASE_URL+"/index.php/admin_ret_lot/lot_inward/list")          
            wait.until(EC.element_to_be_clickable((By.ID,"ltInward-dt-btn"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Sa'])[2]/following::li[1]"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"select2-metal-container"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//span/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//span/input"))).send_keys(row_data["Metal Type"],Keys.ENTER)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@id='lot_inward_search']/i"))).click()
            sleep(3)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).clear()
            if Products:
                Entered_Product=Products[0]
            else:
                Entered_Product=row_data["Product"]
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).send_keys(Entered_Product)
            sleep(5) 
            
            # Use captured_lot_id if available, otherwise fallback to table scraper
            if captured_lot_id:
                Lot_id = captured_lot_id
            else:
                Lot_id = driver.find_element(By.XPATH,"//table[@id='lot_inward_list']/tbody/tr[1]/td[1]").text
            
            print(Lot_id)
            print(type(Lot_id))
            pcs=row_data["Pcs"]
            
            # windows = driver.window_handles
            # driver.switch_to.window(windows[1])   # switch to window 1 (second window)
            # driver.close()
            # driver.switch_to.window(windows[0])
            # driver.execute_script("window.scrollBy(0, -300);") 
            return Test_Status,Actual_Status,Lot_id,pcs,Wt_gram,stone_list
            
    def update_BranchTransfer(self, row_data, workbook, row_num):
        """
        Takes Non-Tagged Section and Product to BranchTransfer sheet.
        TransferType: NonTagged (col 4), Section (col 9), Product (col 10)
        """
        if "BranchTransfer" not in workbook.sheetnames:
            print("⚠️ BranchTransfer sheet not found")
            return

        
        bt_sheet = workbook["BranchTransfer"]
        
        # Find the first empty row in BranchTransfer sheet to ensure sequential updates (2, 3, 4...)
        target_row = 2
        while bt_sheet.cell(row=target_row, column=1).value is not None:
            target_row += 1
            
        # TC numbering starts at row 2 as 001
        test_case_id = f"TC_BT_{target_row - 1:03d}"
            
        print(f"📊 Updating data in BranchTransfer sheet at row {target_row} with {test_case_id}")

        
        # Set new values
        bt_sheet.cell(row=target_row, column=1).value = test_case_id
        bt_sheet.cell(row=target_row, column=4).value = "NonTagged"
        bt_sheet.cell(row=target_row, column=5).value = "HEAD OFFICE"
        bt_sheet.cell(row=target_row, column=7).value = "NO"
        bt_sheet.cell(row=target_row, column=9).value = row_data.get("Section")
        bt_sheet.cell(row=target_row, column=10).value = row_data.get("Product")
        
        workbook.save(FILE_PATH)
        print(f"✅ BranchTransfer Row {target_row} updated with: {test_case_id} / {row_data.get('Lot Received')} / {row_data.get('Product')}")

    def update_NonTag_Detail(self, row_data, workbook, row_num):
        """
        Updates NonTag_Detail sheet for non-tagged lot items.
        """
        if "NonTag_Detail" not in workbook.sheetnames:
            print("⚠️ NonTag_Detail sheet not found")
            return

        nt_sheet = workbook["NonTag_Detail"]
        
        # Find the first empty row
        target_row = nt_sheet.max_row + 1
        while nt_sheet.cell(row=target_row, column=1).value is not None:
            target_row += 1
            
        test_case_id = f"TC_NT_{target_row - 1:03d}"
        print(f"📊 Updating NonTag_Detail at row {target_row}: {test_case_id}")

        # Mapping based on screenshot structure:
        nt_sheet.cell(row=target_row, column=1).value = test_case_id
        nt_sheet.cell(row=target_row, column=2).value = row_data.get("Section")
        nt_sheet.cell(row=target_row, column=3).value = row_data.get("Product")
        nt_sheet.cell(row=target_row, column=4).value = row_data.get("Design")
        nt_sheet.cell(row=target_row, column=5).value = row_data.get("Sub Design")
        nt_sheet.cell(row=target_row, column=6).value = row_data.get("Purity")
        nt_sheet.cell(row=target_row, column=7).value = row_data.get("Pcs")
        nt_sheet.cell(row=target_row, column=8).value = row_data.get("GWT")
        nt_sheet.cell(row=target_row, column=9).value = row_data.get("Purchase MC Type")
        nt_sheet.cell(row=target_row, column=10).value = row_data.get("Purchase MC")
        nt_sheet.cell(row=target_row, column=11).value = row_data.get("Purchase Wastage")
        nt_sheet.cell(row=target_row, column=12).value = row_data.get("Charge Name")
        workbook.save(FILE_PATH)
        print(f"✅ NonTag_Detail Row {target_row} updated correctly.")

    def update_Lot_id(self, Lot_id, row_count, pcs, g_wt_sum, l_wt_sum, meta, workbook, all_stones, test_status=None, actual_status=None):
        """Distributes Lot metadata across Tag and Tag_LWt sheets. Returns (Pcs_count, message)."""
        try:
            Pcs_count = sum(map(int, pcs))
            print(f"Distributing {Pcs_count} pieces for Lot: {Lot_id}")
            tag_sheet = workbook["Tag"]
            lwt_sheet = workbook["Tag_LWt"] if "Tag_LWt" in workbook.sheetnames else workbook.create_sheet("Tag_LWt")

            # Determine the starting row for Tag_LWt
            lwt_row_start = lwt_sheet.max_row + 1
            if lwt_row_start == 2 and lwt_sheet.cell(row=1, column=1).value is None:
                # Initialize headers if sheet is brand new
                headers = ["Test Case Id", "Less Weight", "Type", "Name", "Code", "Pcs", "Wt", "Wt Type", "Cal.Type", "Rate", "Amount"]
                for h_col, h_text in enumerate(headers, 1):
                    lwt_sheet.cell(row=1, column=h_col).value = h_text
                lwt_row_start = 2

            try:
                Pcs_count = int(Pcs_count)
            except Exception:
                Pcs_count = 1

            # Logic: NWT = GWT - LWT
            total_gwt = Decimal(str(g_wt_sum or 0))
            total_lwt = Decimal(str(l_wt_sum or 0))
            total_nwt = total_gwt - total_lwt

            if Pcs_count > 0:
                # 1. Distribute Net Weight (NWT) with random variance
                weights = [Decimal(str(1 + random.uniform(-0.2, 0.2))) for _ in range(Pcs_count)]
                if sum(weights) != 0:
                    scale = total_nwt / sum(weights)
                    nwt_list = [w * scale for w in weights]
                else:
                    nwt_list = [total_nwt / Pcs_count for _ in range(Pcs_count)]

                # Round NWT to 3 decimal places
                nwt_final = [w.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP) for w in nwt_list]

                # Force sum to match total_nwt exactly
                diff_nwt = total_nwt - sum(nwt_final)
                if nwt_final:
                    nwt_final[0] += diff_nwt

                # 2. Distribute Less Weight (LWT) equally
                lwt_per_pcs = (total_lwt / Pcs_count).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
                lwt_final = [lwt_per_pcs for _ in range(Pcs_count)]

                # Force sum to match total_lwt exactly
                diff_lwt = total_lwt - sum(lwt_final)
                if lwt_final:
                    lwt_final[0] += diff_lwt

                # 3. Apply to Sheets
                current_lwt_row = lwt_row_start
                for i in range(Pcs_count):
                    row_Num = row_count + i

                    # Assign a unique Test Case ID for each piece
                    tag_tc_id = f"TAG_{Lot_id}_{i+1}"
                    tag_sheet.cell(row=row_Num, column=1).value = tag_tc_id

                    # Column 2 & 3: Status
                    if test_status:
                        tag_sheet.cell(row=row_Num, column=2).value = test_status
                    if actual_status:
                        tag_sheet.cell(row=row_Num, column=3).value = actual_status


                    # Column 4: Branch
                    tag_sheet.cell(row=row_Num, column=4).value = meta.get("Lot Received")

                    # Column 5: Lot ID
                    tag_sheet.cell(row=row_Num, column=5).value = Lot_id

                    # Column 6: Product
                    tag_sheet.cell(row=row_Num, column=6).value = meta.get("Product")

                    # Column 7: Section
                    tag_sheet.cell(row=row_Num, column=7).value = "GOLD CHAIN"



                    # Column 8: Design
                    tag_sheet.cell(row=row_Num, column=8).value = meta.get("Design")

                    # Column 9: Sub Design
                    tag_sheet.cell(row=row_Num, column=9).value = meta.get("Sub Design")

                    # Column 10: Pieces (Individual tag is usually 1 pcs)
                    tag_sheet.cell(row=row_Num, column=10).value = 1

                    # Column 11: No. of items (Single tagging)
                    tag_sheet.cell(row=row_Num, column=11).value = 1

                    # Column 13: Individual Less Weight
                    indiv_lwt = float(lwt_final[i])
                    tag_sheet.cell(row=row_Num, column=13).value = indiv_lwt

                    # Column 12: Individual Gross Weight (NWT + LWT)
                    indiv_nwt = float(nwt_final[i])
                    tag_sheet.cell(row=row_Num, column=12).value = round(indiv_nwt + indiv_lwt, 3)

                    # Update Tag_LWt with distributed stone details for this specific piece
                    has_stone_this_tag = False
                    if all_stones:
                        for stone in all_stones:
                            # Cumulative distribution to preserve integers and match Lot totals
                            total_pcs = safe_float(stone.get("Pcs"))
                            total_wt = safe_float(stone.get("Wt"))
                            stone_rate = safe_float(stone.get("Rate"))

                            # (i+1) vs i approach ensures 25/50 leads to 1 or 0 per tag proportionally
                            tag_pcs = int((i + 1) * total_pcs / Pcs_count) - int(i * total_pcs / Pcs_count)

                            # Weight Distribution logic:
                            if total_pcs > 0:
                                # Tied to piece distribution - weighted by pieces
                                tag_wt = (total_wt / total_pcs) * tag_pcs
                            else:
                                # Uniform distribution for weight-only stones
                                tag_wt = ((i + 1) * total_wt / Pcs_count) - (i * total_wt / Pcs_count)

                            # Only add to Tag_LWt if there's a stone or weight for this piece
                            if tag_pcs > 0 or tag_wt > 0:
                                has_stone_this_tag = True
                                lwt_sheet.cell(row=current_lwt_row, column=1).value = tag_tc_id
                                lwt_sheet.cell(row=current_lwt_row, column=2).value = "Yes"
                                lwt_sheet.cell(row=current_lwt_row, column=3).value = stone.get("Type")
                                lwt_sheet.cell(row=current_lwt_row, column=4).value = stone.get("Name")
                                lwt_sheet.cell(row=current_lwt_row, column=5).value = stone.get("Code")
                                lwt_sheet.cell(row=current_lwt_row, column=6).value = tag_pcs
                                lwt_sheet.cell(row=current_lwt_row, column=7).value = round(tag_wt, 3)
                                lwt_sheet.cell(row=current_lwt_row, column=8).value = stone.get("Wt Type")
                                lwt_sheet.cell(row=current_lwt_row, column=9).value = stone.get("Cal.Type")
                                lwt_sheet.cell(row=current_lwt_row, column=10).value = stone_rate

                                # Recalculate amount based on distributed values and Cal.Type
                                cal_type = str(stone.get("Cal.Type") or "").strip().lower()
                                if cal_type == "pcs":
                                    indiv_amount = tag_pcs * stone_rate
                                else:  # Default to "wt" calculation
                                    indiv_amount = tag_wt * stone_rate

                                lwt_sheet.cell(row=current_lwt_row, column=11).value = round(indiv_amount, 2)
                                current_lwt_row += 1

                    # Update Less Weight status in Tag sheet based on actual distribution
                    if has_stone_this_tag:
                        tag_sheet.cell(row=row_Num, column=13).value = "Yes"
                    else:
                        tag_sheet.cell(row=row_Num, column=13).value = "No"

                    # Note: User explicitly said "col 13 net no need"

            workbook.save(FILE_PATH)
            workbook.close()
            return Pcs_count, "Lot ID and Metadata distributed in Tag sheet successfully"
        except Exception as e:
            print(f"❌ update_Lot_id failed: {e}")
            return 0, f"update_Lot_id error: {str(e)}"

    def Lotdetails(self, TestCaseId):
        """Returns row data dict for given TestCaseId from the Lot sheet, or {} on failure."""
        print(TestCaseId)
        try:
            function_name = "Lot"
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[function_name]
            for row_num in range(2, valid_rows):
                current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
                if current_id == TestCaseId:
                    data = {
                        "Test Case Id": 1,
                        "Pcs": 14,
                        "GWT": 15,
                        "LWt": 16,
                    }
                    row_Lotdata = {key: sheet.cell(row=row_num, column=col).value
                                    for key, col in data.items()}
                    print(row_Lotdata)
                    return row_Lotdata
            print(f"⚠️ Lotdetails: TestCaseId '{TestCaseId}' not found in sheet")
            
        except Exception as e:
            print(f"⚠️ Lotdetails failed for '{TestCaseId}': {e}")
            return {}

    def is_element_present(self, how, what):
        try: self.wait.until(EC.element_to_be_clickable(by=how, value=what))
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True

if __name__ == "__main__":
    unittest.main()
