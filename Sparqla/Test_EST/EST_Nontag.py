from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import random
import math


FILE_PATH = ExcelUtils.file_path
class ESTIMATION_NonTag(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_estimation_Nontag(self,test_case_id,Board_Rate):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'NonTag_Est'
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name,test_case_id)
        print(value)
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        print(row)
        salevalue=0
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id":1,
                    "Test Status":2,
                    "Actual Status":3,				 	
                    "Section":4,
                    "Product":5,
                    "Design":6,
                    "Sub Design":7, 
                    "Employee":8,
                    "Purity":9,
                    "Size":10, 
                    "Pcs":11,
                    "G.Wt":12, 
                    "Rate":13, 
                    "MC Type":14, 
                    "MC Value":15,
                    "Wastage%":16, # Note: Sheet has 'Wastage% '
                    "Charge":17,
                    "Field_validation_status":18,
                }
                row_data = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, (col) in data.items()
                }
                print(row_data)               
                # [NEW] Prioritized Inventory Lookup
                sec = str(row_data["Section"] or "")
                prod = str(row_data["Product"] or "")
                des = str(row_data["Design"] or "")
                sub = str(row_data["Sub Design"] or "")
                needed_pcs = int(row_data["Pcs"] or 1)
                

                src_sheet, src_row, tot_pcs, tot_purity, tot_gwt, tot_Mctype, tot_McValue, tot_Wastage, othercharge = ESTIMATION_NonTag._find_nontag_source(sec, prod, des, sub)
                
                if src_sheet:
                    # Calculate proportional weight
                    weight_per_pc = tot_gwt / tot_pcs
                    Wastage = tot_Wastage +2
                    if tot_Mctype == 'Per Piece':
                        MC_Type = 'Piece'
                    else:
                        MC_Type = 'Gram'
                    print(Wastage)

                    taken_weight = round(weight_per_pc * needed_pcs, 3)
                    
                    print(f"🔍 Found inventory in {src_sheet} row {src_row}. Calculating weight for {needed_pcs} pcs: {taken_weight}")
                    
                    # Update NonTag_Est sheet immediately
                    sheet.cell(row=row_num, column=9, value=tot_purity)
                    sheet.cell(row=row_num, column=11, value=needed_pcs)
                    sheet.cell(row=row_num, column=12, value=taken_weight)
                    sheet.cell(row=row_num, column=14, value=MC_Type)
                    sheet.cell(row=row_num, column=15, value=tot_McValue)
                    sheet.cell(row=row_num, column=16, value=Wastage)
                    sheet.cell(row=row_num, column=17, value=othercharge)
                    workbook.save(FILE_PATH)
                    
                    # [FIX] Update row_data so 'create' method uses the latest values
                    row_data["Purity"] = tot_purity
                    row_data["Pcs"] = needed_pcs
                    row_data["G.Wt"] = taken_weight
                    row_data["MC Type"] = MC_Type
                    row_data["MC Value"] = tot_McValue
                    row_data["Wastage%"] = Wastage
                    row_data["Charge"] = othercharge
                    
                    # Store track of taken inventory
                    if not hasattr(self, 'nontag_found_rows'):
                        self.nontag_found_rows = []
                    self.nontag_found_rows.append((src_sheet, src_row, needed_pcs, taken_weight))
                    
                    # Use the calculated weight for estimation
                    row_data["G.Wt"] = taken_weight
                else:
                    msg = "❌ Faill stock i completed add stock"
                    print(msg)
                    ESTIMATION_NonTag.update_excel_status(self, row_num, "Fail", msg, Sheet_name)
                    break

                # Call your 'create' method
                Create_data = ESTIMATION_NonTag.create(self,row_data, row_num, Sheet_name, row,Board_Rate)
                print(Create_data)
                row = row+1
                
                if Create_data:
                    ceil_value,Test_Status,Actual_Status= Create_data
                    ESTIMATION_NonTag.update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name)
                    salevalue = salevalue + float(ceil_value)
        return salevalue, getattr(self, 'nontag_found_rows', [])
                # Remove processed customer from the list
                
    def create(self,row_data, row_num, Sheet_name, row,Board_Rate):
        driver=self.driver
        wait = self.wait
        Mandatory_field=[]
        Error_field_val=[]
        sleep(3)
        #Tag Check box selected
   
        if row > 1:
            sleep(4)
            Function_Call.click(self,'//button[@id="create_catalog_details"]')
        else:
            Function_Call.click2(self,'//input[@id="select_catalog_details"]')            
        # category 
        sleep(3)
        if row_data["Section"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_catalog') and contains(@id,'[id_section]')])[{row}]", 
                row_data["Section"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Category field is mandatory ⚠️"
            Mandatory_field.append("Category"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)           
              
        if row_data["Product"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_catalog') and contains(@id,'[product]')])[{row}]", 
                row_data["Product"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Product field is mandatory ⚠️"
            Mandatory_field.append("Product"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
               
        # Design 
        if row_data["Design"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_catalog') and contains(@id,'[design]')])[{row}]", 
                row_data["Design"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Design field is mandatory ⚠️"
            Mandatory_field.append("Design"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        #Sub Design 
        if row_data["Sub Design"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_catalog') and contains(@id,'[id_sub_design]')])[{row}]", 
                row_data["Sub Design"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Sub Design field is mandatory ⚠️"
            Mandatory_field.append("Sub Design"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        # Employee 
        if row_data["Employee"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_catalog') and contains(@id,'[item_emp_id]')])[{row}]", 
                row_data["Employee"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Employee field is mandatory ⚠️"
            Mandatory_field.append("Employee"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        # Purity
        if row_data["Purity"] is not None:
            sleep(2)
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_catalog') and contains(@id,'[purity]')]/span)[{row}]", 
                row_data["Purity"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Purity field is mandatory ⚠️"
            Mandatory_field.append("Purity"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)

        if row_data["Size"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, f'(//input[@name="est_catalog[size][]"])[1]'),
                value=row_data["Size"],
                pattern = r"\d{1,2}(\.\d{1,2})?$",
                field_name="Size",
                screenshot_prefix="Size",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass

        if row_data["Pcs"]:
            Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_catalog[pcs][]"])[{row}]'),
            value=row_data["Pcs"],
            pattern = r"^\d{1,3}$",
            field_name="Pcs",
            screenshot_prefix="Pcs",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            
        else:
            msg = f"'{None}' → Pcs field is mandatory ⚠️"
            Mandatory_field.append("Pcs"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)       
        
        if row_data["G.Wt"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_catalog[gwt][]"])[{row}]'),
            value=row_data["G.Wt"], 
            pattern = r"^\d{1,3}(\.\d{1,3})?$",
            field_name="G.Wt",
            screenshot_prefix="G.Wt",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            
        else:
            msg = f"'{None}' → G.Wt field is mandatory ⚠️"
            Mandatory_field.append("G.Wt"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["MC Type"]:
            Function_Call.select_visible_text(self,f'//select[@class="form-control mc_type"]', row_data["MC Type"])
            
        if row_data["MC Value"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_catalog[mc][]"])[{row}]'),
            value=row_data["MC Value"],
            pattern = r"^\d{1,5}(\.\d{1,2})?$",
            field_name="MC Value",
            screenshot_prefix="MC Value",
            row_num=row_num,
            Sheet_name=Sheet_name)
            
        else:
            msg = f"'{None}' → MC Value field is mandatory ⚠️"
            Mandatory_field.append("MC Value"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)        
        
        if row_data["Wastage%"]:
            errors=Function_Call.fill_input(    
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_catalog[wastage][]"])[{row}]'),
            value=row_data["Wastage%"],
            pattern = r"^\d{1,2}(\.\d{1,2})?$",
            field_name="Wastage%",
            screenshot_prefix="Wastage%",
            range_check = lambda v: 0 <= float(v) <= 99,
            row_num=row_num,
            Sheet_name=Sheet_name)
            
        else:
            msg = f"'{None}' → Wastage% Value field is mandatory ⚠️"
            Mandatory_field.append("Wastage%"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)        
        
        # Open Other Charge section
        Function_Call.click(self, f"(//table[@id='estimation_catalog_details']//td[21]/a)[{row}]")

        charges_raw = row_data["Charge"]
        if not charges_raw:
            msg = "⚠️ OtherCharge flag is Yes but no ChargeName provided"
            print(msg)
            return "0.00", "Fail", msg
        charges_list = [s.strip() for s in charges_raw.split(",")]

        for idx, charge in enumerate(charges_list):
            # For the 2nd, 3rd, ... charges → click +Add
            if idx > 0:
                Function_Call.click(self, '//button[@id="add_new_charge"]')
        
            # Select charge type
            sleep(3)
            Function_Call.select(self,f'(//select[@name="est_stones_item[id_charge][]"])[{idx+1}]',charge)
         
            # Locate corresponding value field (same row as idx+1)
            value_input = wait.until(EC.presence_of_element_located(
                (By.XPATH, f"(//input[@name='est_stones_item[value_charge][]'])[{idx+1}]")
            ))
            current_value = value_input.get_attribute("value").strip()

            # If empty or "0.00" → auto-fill random multiple of 100
            if current_value == "0.00":
                random_value = random.randint(1, 10) * 100
                sleep(1)
                value_input.clear()
                value_input.send_keys(str(random_value))
                print(f"⚡ Added random value {random_value} for {charge}")
            else:
                print(f"✅ Auto value {current_value} kept for {charge}")

        # Save button
        wait.until(EC.element_to_be_clickable((By.ID, "update_charge_details"))).click()
        print("Field✅ OtherCharges added:", charges_list)        
        
        # Fetch values with one-liners
        Gwt   = ESTIMATION_NonTag.get_val(self, f'(//input[@name="est_catalog[gwt][]"])[{row}]')
        Nwt   = ESTIMATION_NonTag.get_val(self, f'(//input[@name="est_catalog[nwt][]"])[{row}]')
        PCS   = ESTIMATION_NonTag.get_val(self, f'(//input[@name="est_catalog[pcs][]"])[{row}]', cast=float)
        Wast_per = ESTIMATION_NonTag.get_val(self, f'(//input[@name="est_catalog[wastage][]"])[{row}]')
        Wast  = ESTIMATION_NonTag.get_val(self, f'(//input[@class="form-control cat_wastage_wt"])[{row}]')
        Mc    = ESTIMATION_NonTag.get_val(self, f'(//input[@name="est_catalog[mc][]"])[{row}]')
        Taxvalue=ESTIMATION_NonTag.get_val(self,f'(//input[@name="est_catalog[tax_per][]"])[{row}]')
        other_Amt=ESTIMATION_NonTag.get_val(self,f'(//input[@name="est_catalog[value_charge][]"])[{row}]')
        print(Taxvalue)
        print(type(Taxvalue))
        LWT_Tot_Amt=0
        Stone = LWT_Tot_Amt                    

        # Taxable amount kept as string (not converted to float)
        Taxable_Amt = Function_Call.get_value(self,f'(//input[@name="est_catalog[amount][]"])[{row}]')
        print(Taxable_Amt) 
        # MC type  
        mc_type_dropdown = wait.until(EC.presence_of_element_located((By.XPATH, '//select[@class="form-control mc_type"]')))
        sleep(3)
        select = Select(mc_type_dropdown)
        selected_text = select.first_selected_option.text
        print("Selected MC Type:", selected_text)
        Mc_type =selected_text 
        #purity
        purity_dropdown=wait.until(EC.presence_of_element_located((By.XPATH, "//span[starts-with(@id,'select2-est_catalog[purity]') and contains(@id,'-container')]")))
        # purity_dropdown = wait.until(EC.presence_of_element_located((By.XPATH, '//span[@id="select2-est_catalog[purity][]-az-container"]')))
        sleep(3)
        purity = purity_dropdown.get_attribute("title")
        print(purity)
        
        
        gold_rate = 0
        try:
            purity_val = float(purity)
            if purity_val == 91.60:
                gold_rate = Board_Rate[0]
            elif purity_val == 75.0:
                gold_rate = Board_Rate[1]
            elif purity_val == 92.5:
                gold_rate = Board_Rate[2]
            else:
                print(f"⚠️ Unknown purity value: {purity_val}. Using rate 0.")
        except (ValueError, TypeError):
            print(f"❌ Could not convert purity '{purity}' to float.")

        
        
        
        Cal_current_value=Function_Call.get_value(self, f'(//input[@class="cat_calculation_based_on"])[{row}]') 
        Taxvalue=Function_Call.get_value(self,f'(//input[@name="est_catalog[tax_per][]"])[{row}]')
        Taxvalue=float(Taxvalue)
        print(Taxvalue)
        # Debug print all values
        print(f"Gwt={Gwt}, Nwt={Nwt}, PCS={PCS}, Stone={Stone}, "
            f"Wast_per={Wast_per}, Wast={Wast}, Mc={Mc}, Mc_type={Mc_type},"
            f"Taxable={Taxable_Amt}")
        
        
       
        Result=ESTIMATION_NonTag.calculation(self,Cal_current_value,gold_rate,Gwt,Nwt,Wast_per,Mc,Stone,other_Amt,Mc_type,Taxvalue)
        ceil_value,Cal_type,IGst,SGst =Result
        if ceil_value==Taxable_Amt:
            Test_Status= "Pass"
            Actual_Status =(f"✅ Calculation Value is correct {ceil_value}")
        else:
            Test_Status= "Fail"
            Actual_Status =(f"❌ Calculation Error in ={ceil_value} | Web Value={Taxable_Amt}")
        return ceil_value,Test_Status,Actual_Status
            
    def calculation(self,Cal_current_value,gold_rate,Gwt,Nwt,Wast_per,Mc,Stone,other_Amt,Mc_type,Taxvalue):
        wait = self.wait 
        data = {
            "0": "Mc & Wast On Gross",
            "1": "Mc & Wast On Net",
            "2": "Mc on Gross, Wast On Net",
            "3": "Fixed Rate",
            "4": "Fixed Rate based on Weight"
        }
        Cal_type = data[str(Cal_current_value)]   # convert int → str because keys are strings
        print(f"\n{'='*20}")
        print(f"📊 CALCULATION TYPE: {Cal_type}")
        print(f"{'='*20}\n")

        gross_weight=Gwt
        net_weight = Nwt  
        wastage_percentage = Wast_per 
        Making_cost_pergram = Mc 
        diamond_cost =Stone
        Charge_Amt = other_Amt
        Tax = float(Taxvalue)       
        # initialize
        ceil_value = "0.00"
        IGst = "0.00"
        SGst = "0.00"
        
        if Cal_type=="Mc on Gross, Wast On Net":
        # calculation making cost on gross Wastage% on Net  
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va = round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal = (total*gold_rate)+diamond_cost+mc+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(Cal)))
            print(ceil_value)
                       
        if  Cal_type=="Mc & Wast On Net":
            # calculation making cost & Wastage% on Net  
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*net_weight
                mc=float("{:.2f}".format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va= round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal2 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(Cal2)))
                        
        if  Cal_type == "Mc & Wast On Gross":
            # calculation making cost & Wastage% on Gross 81148.00
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                Mc= gross_weight*Making_cost_pergram
                mc=float("{:.2f}".format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
           
        if Cal_type== "Fixed Rate based on Weight":
            if Mc_type=='Piece':
                mc = Making_cost_pergram
            else:
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
        
        if Cal_type == "Fixed Rate":
            ceil_value
        
        if Tax:  
           salevalue=float(ceil_value)
           Find_Tax=(salevalue*Tax)/100
           Tol_Amt = salevalue+Find_Tax
           Gst=Find_Tax/2
           IGst=("{:.2f}".format(Gst))
           SGst=("{:.2f}".format(Gst))
           ceil_value=("{:.2f}".format(Tol_Amt))        
        print(ceil_value)
        print(type(ceil_value))
        return ceil_value,Cal_type,IGst,SGst



      # --- Helper to fetch field values safely ---
    def get_val(self,locator, cast=float, default=0):
        wait = self.wait
        el = wait.until(EC.presence_of_element_located((By.XPATH, locator)))
        val = el.get_attribute("value")
        if not val:  
            return default
        return cast(val)

    def update_excel_status(self,row_num, Test_Status, Actual_Status, function_name):
        print(function_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]  # or workbook["SheetName"]
        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, function_name)
        # Print and return status
        print(Status)
        return Status

    @staticmethod
    def _find_nontag_source(sec, product, design, sub_design):
        """
        Look for available inventory in NonTag_Detail or Purchase_NonTagDetail.
        """
        sheets_to_check = ["NonTag_Detail", "Purchase_NonTagDetail"]
        workbook = load_workbook(FILE_PATH)
        
        for sheet_name in sheets_to_check:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            # Mapping based on Lot.py updates: 3:Product, 4:Design, 5:SubDesign, 7:Pcs, 8:G.Wt
            # Inventory tracking in Col 15 (Taken Pcs) and Col 16 (Taken Weight)
            for r in range(2, sheet.max_row + 1):
                s_sec = str(sheet.cell(row=r, column=2).value or "").strip()
                s_prod = str(sheet.cell(row=r, column=3).value or "").strip()
                s_des = str(sheet.cell(row=r, column=4).value or "").strip()
                s_sub = str(sheet.cell(row=r, column=5).value or "").strip()
                
                if s_sec.lower() == sec.lower() and \
                   s_prod.lower() == product.lower() and \
                   s_des.lower() == design.lower() and \
                   s_sub.lower() == sub_design.lower():
                    
                    total_purity = float(sheet.cell(row=r, column=6).value or 0)
                    total_pcs = int(sheet.cell(row=r, column=7).value or 0)
                    total_gwt = float(sheet.cell(row=r, column=8).value or 0)
                    total_Mctype = str(sheet.cell(row=r, column=9).value or '')
                    total_McValue = float(sheet.cell(row=r, column=10).value or 0)
                    total_Wastage = float(sheet.cell(row=r, column=11).value or 0)
                    othercharge = str(sheet.cell(row=r, column=12).value or '')
                    taken_pcs = float(sheet.cell(row=r, column=14).value or 0)
                    
                    if total_pcs > 0 and taken_pcs < total_pcs:
                        return sheet_name, r, total_pcs, total_purity, total_gwt, total_Mctype, total_McValue, total_Wastage, othercharge
        return None, None, None, None, None, None, None, None, None

    @staticmethod
    def update_source_inventory(found_rows,estimation_no):
        """
        Increments the 'Taken' PCS and Weight in the source sheets.
        found_rows: list of (sheet_name, row_idx, taken_pcs, taken_weight)
        """
        from datetime import datetime
        try:
            workbook = load_workbook(FILE_PATH)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            entry = f"{estimation_no} - {timestamp}"
            
            # 1. Update Inventory Sheets
            for sheet_name, r, t_pcs, t_gwt in found_rows:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    curr_pcs = float(sheet.cell(row=r, column=14).value or 0)
                    curr_gwt = float(sheet.cell(row=r, column=15).value or 0)
                    
                    sheet.cell(row=r, column=14, value=curr_pcs + t_pcs)
                    sheet.cell(row=r, column=15, value=curr_gwt + t_gwt)
                    # Update column 13 status in inventory sheet to "Estimated"
                    sheet.cell(row=r, column=13, value="Estimated").font = Font(bold=True, color="00B050")
                    # Update column 16 with Estimation No and Time
                    sheet.cell(row=r, column=16, value=entry)
                    print(f"✅ Updated {sheet_name} row {r}: +{t_pcs} Pcs, +{t_gwt} Weight with Est No: {estimation_no}")

            workbook.save(FILE_PATH)
        except Exception as e:
            print(f"❌ Error updating source inventory or EST sheet: {e}")





