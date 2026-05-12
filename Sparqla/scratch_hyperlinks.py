import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

file_path = r'C:\Users\admin\Desktop\sqrqlas\Book1.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)

    if 'Master' in wb.sheetnames:
        master_sheet = wb['Master']
        # Start iterating from row 2 (assuming row 1 is header)
        for row in range(2, master_sheet.max_row + 1):
            cell = master_sheet.cell(row=row, column=1)
            if cell.value:
                sheet_name = str(cell.value)
                # Make sure the sheet exists before linking
                if sheet_name in wb.sheetnames:
                    # Add hyperlink to cell
                    cell.hyperlink = f"#'{sheet_name}'!A1"
                    cell.style = 'Hyperlink' # Use default Excel hyperlink styling

                    # Now add hyperlink back to Master from the target sheet
                    target_sheet = wb[sheet_name]
                    for t_row in range(1, target_sheet.max_row + 1):
                         t_cell = target_sheet.cell(row=t_row, column=1) # TestCaseId column
                         if t_row == 1:
                              # Only hyperlink the header back to Master
                              if t_cell.value:
                                   t_cell.hyperlink = "#'Master'!A1"
                                   t_cell.style = 'Hyperlink'
                         else:
                              # Remove hyperlinks from all other rows (TC001, etc.)
                              t_cell.hyperlink = None
                              t_cell.style = 'Normal'

    wb.save(file_path)
    print('Hyperlinks added successfully.')
except Exception as e:
    print(f"Error: {e}")
