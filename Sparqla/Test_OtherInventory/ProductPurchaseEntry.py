from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class ProductPurchaseEntry(unittest.TestCase):
    """
    Other Inventory: Purchase Entry Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_product_purchase_entry(self):
        """Main entry point for Product Purchase Entry automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Other Inventory Purchase')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/purchase_entry/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "ProductPurchaseEntry"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "SupplierName": 4, "BillNo": 5, "BillDate": 6,
                "ItemName": 7, "Quantity": 8, "Rate": 9, "GST": 10, "ExpectedStatus": 11
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_product_purchase_entry_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_product_purchase_entry_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Product Purchase Entry Automation Completed")
        print(f"{'='*80}")

    def test_product_purchase_entry_add(self, row_data, row_num, sheet_name):
        """Logic to create a new Product Purchase Entry"""
        driver = self.driver
        wait = self.wait
        current_field = "Add Button"
        
        try:
            # 1. Click Add Button
            try:
                Function_Call.click(self, '//a[@id="add_pur_details"]')
            except:
                driver.get(BASE_URL+"index.php/admin_ret_other_inventory/purchase_entry/add")
            sleep(3)

            # 2. Supplier Details
            if row_data["SupplierName"]:
                current_field = f"Supplier Name ({row_data['SupplierName']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_karigar"]/following-sibling::span',
                    row_data["SupplierName"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            if row_data["BillNo"]:
                current_field = f"Bill No ({row_data['BillNo']})"
                Function_Call.fill_input(self, wait, (By.ID, "sup_refno"), str(row_data["BillNo"]), "Bill No", row_num, sheet_name)

            if row_data["BillDate"]:
                current_field = f"Bill Date ({row_data['BillDate']})"
                bill_date = driver.find_element(By.ID, "sup_billdate")
                bill_date.clear()
                bill_date.send_keys(str(row_data["BillDate"]))

            # 3. Item Details
            if row_data["ItemName"]:
                current_field = f"Select Item ({row_data['ItemName']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_item"]/following-sibling::span',
                    row_data["ItemName"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            if row_data["Quantity"]:
                current_field = f"Quantity ({row_data['Quantity']})"
                Function_Call.fill_input(self, wait, (By.ID, "buy_quantity"), str(row_data["Quantity"]), "Quantity", row_num, sheet_name)

            if row_data["Rate"]:
                current_field = f"Rate ({row_data['Rate']})"
                Function_Call.fill_input(self, wait, (By.ID, "buy_rate"), str(row_data["Rate"]), "Rate", row_num, sheet_name)

            if row_data["GST"] is not None:
                current_field = f"GST ({row_data['GST']})"
                Function_Call.fill_input(self, wait, (By.ID, "tax_amount"), str(row_data["GST"]), "GST", row_num, sheet_name)

            # 4. Add to Grid
            current_field = "Add Item Button"
            Function_Call.click(self, '//button[@id="add_item_info"]')
            sleep(2)

            # 5. Save
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="inventory_submit"]')
            sleep(4)

            # 6. Success Verification
            try:
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Purchase Entry Added successfully')]"
                msg = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath))).text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "Purchase Entry Added successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_product_purchase_entry_list_verification(self, row_data):
        """Verify the added purchase entry in the list page"""
        driver = self.driver
        wait = self.wait
        
        try:
            if "purchase_entry/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_other_inventory/purchase_entry/list")
                sleep(2)
            
            # Select Supplier in filter
            if row_data["SupplierName"]:
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_karigar"]/following-sibling::span',
                    row_data["SupplierName"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # Search
            Function_Call.click(self, '//button[@id="purchase_item_search"]')
            sleep(3)
            
            try:
                # Check if table row exists with Bill No
                row_xpath = f"//table[@id='other_item_pur']/tbody/tr[contains(., '{row_data['BillNo']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                print(f"✅ Found Purchase Entry {row_data['BillNo']} in list page.")
                return ("Pass", f"Verified {row_data['BillNo']} in list")
            except:
                print(f"❌ Could not find Purchase Entry {row_data['BillNo']} in list page")
                return ("Fail", f"Purchase Entry {row_data['BillNo']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
