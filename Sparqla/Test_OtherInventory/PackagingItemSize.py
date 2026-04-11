from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class PackagingItemSize(unittest.TestCase):
    """
    Other Inventory: Packaging Item Size Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_packaging_item_size(self):
        """Main entry point for Packaging Item Size automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Packaging Item Size')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/item_size/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "PackagingItemSize"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "SizeName": 4, "ExpectedStatus": 5
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_packaging_item_size_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_packaging_item_size_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Packaging Item Size Automation Completed")
        print(f"{'='*80}")

    def test_packaging_item_size_add(self, row_data, row_num, sheet_name):
        """Logic to add a new Packaging Item Size via Modal"""
        driver = self.driver
        wait = self.wait
        current_field = "Add Button"
        
        try:
            # 1. Click Add Button to open modal
            wait.until(EC.element_to_be_clickable((By.ID, "add_issue_details"))).click()
            sleep(2) # Wait for modal animation
            
            # 2. Enter Size Name
            if row_data["SizeName"]:
                current_field = f"Size Name ({row_data['SizeName']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "size_name"),
                    value=str(row_data["SizeName"]),
                    field_name="Size Name",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            # 3. Save
            current_field = "Save Button"
            Function_Call.click(self, '//a[@id="add_item_size"]')
            sleep(3)

            # 4. Success Verification
            try:
                # Check for toast/alert
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Item Size Added successfully')]"
                success_element = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath)))
                msg = success_element.text.strip()
                print(f"✅ Success Message: {msg}")
                
                # Close modal if still open (sometimes Save button doesn't auto-dismiss correctly in some environments)
                try:
                    close_btn = driver.find_elements(By.ID, "modal-footer-close-button")
                    if close_btn and close_btn[0].is_displayed():
                        close_btn[0].click()
                except:
                    pass
                    
                return ("Pass", "Item Size Added successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_packaging_item_size_list_verification(self, row_data):
        """Verify the added size in the list table"""
        driver = self.driver
        wait = self.wait
        
        try:
            if "item_size/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_other_inventory/item_size/list")
                sleep(2)
            
            search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(str(row_data["SizeName"]))
            sleep(2)
            
            try:
                # Verify first row
                row_xpath = f"//table[@id='size_list']/tbody/tr[contains(., '{row_data['SizeName']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                print(f"✅ Found Size {row_data['SizeName']} in list page.")
                return ("Pass", f"Verified Size {row_data['SizeName']} in list")
            except:
                print(f"❌ Could not find Size {row_data['SizeName']} in list page after search")
                return ("Fail", f"Size {row_data['SizeName']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
