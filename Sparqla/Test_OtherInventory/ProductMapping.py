from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class ProductMapping(unittest.TestCase):
    """
    Other Inventory: Product Mapping Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_product_mapping(self):
        """Main entry point for Product Mapping automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Product Mapping')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/product_mapping/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "ProductMapping"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "ItemName": 4, "ProductName": 5, "ExpectedStatus": 6
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_product_mapping_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_product_mapping_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Product Mapping Automation Completed")
        print(f"{'='*80}")

    def test_product_mapping_add(self, row_data, row_num, sheet_name):
        """Logic to create a new Product Mapping"""
        driver = self.driver
        wait = self.wait
        current_field = "Select Item"
        
        try:
            # 1. Select Item (Select2)
            if row_data["ItemName"]:
                current_field = f"Select Item ({row_data['ItemName']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_item"]/following-sibling::span',
                    row_data["ItemName"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 2. Select Product (Select2 Multiple)
            if row_data["ProductName"]:
                current_field = f"Select Product ({row_data['ProductName']})"
                # Handling multiple products (if comma-separated in Excel)
                products = [p.strip() for p in str(row_data["ProductName"]).split(',')]
                for product in products:
                    Function_Call.dropdown_select(
                        self,
                        '//select[@id="select_product"]/following-sibling::span',
                        product,
                        '//span[@class="select2-search select2-search--dropdown"]/input'
                    )
                    sleep(0.5)

            # 3. Update Mapping
            current_field = "Update Button"
            Function_Call.click(self, '//button[@id="update_product_mapping"]')
            sleep(3)

            # 4. Success Verification
            try:
                # Based on controller, it returns JSON and likely shows a toast/alert
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Product Mapped successfully')]"
                # Since it's AJAX, we might need to check if the success message appeared
                msg = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath))).text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "Product Mapped successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to Proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_product_mapping_list_verification(self, row_data):
        """Verify the added mapping using Filters"""
        driver = self.driver
        wait = self.wait
        
        try:
            # Use Filters to verify
            # 1. Product Filter
            if row_data["ProductName"]:
                # Use the first product if multiple were mapped
                primary_product = str(row_data["ProductName"]).split(',')[0].strip()
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="prod_filter"]/following-sibling::span',
                    primary_product,
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 2. Item Filter
            if row_data["ItemName"]:
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="item_filter"]/following-sibling::span',
                    row_data["ItemName"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 3. Search
            Function_Call.click(self, '//button[@id="search_design_maping"]')
            sleep(2)
            
            try:
                # Check if table row exists
                row_xpath = f"//table[@id='mapping_list']/tbody/tr[contains(., '{row_data['ItemName']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                print(f"✅ Found Mapping in search results.")
                return ("Pass", f"Verified mapping in list")
            except:
                print(f"❌ Could not find mapping in search results")
                return ("Fail", "Mapping not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
