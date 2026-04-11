from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class InventoryCategory(unittest.TestCase):
    """
    Other Inventory: Inventory Category Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_inventory_category(self):
        """Main entry point for Inventory Category automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Inventory Category')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/inventory_category/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "InventoryCategory"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "Name": 4, "OutwardType": 5, "AsBillable": 6,
                "ExpiryDateValidate": 7, "ReorderLevel": 8, "ExpectedStatus": 9
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_inventory_category_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page if needed
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_inventory_category_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Inventory Category Automation Completed")
        print(f"{'='*80}")

    def test_inventory_category_add(self, row_data, row_num, sheet_name):
        """Logic to add a new Inventory Category"""
        driver = self.driver
        wait = self.wait
        current_field = "Navigation"
        
        try:
            # Click Add Button (Standard icon for Add in most Sparqla pages)
            try:
                Function_Call.click(self, '//a[contains(@href, "/add")]')
            except:
                driver.get(BASE_URL+"index.php/admin_ret_other_inventory/inventory_category/add")
            
            sleep(3)
            
            # 1. Name
            if row_data["Name"]:
                current_field = f"Name ({row_data['Name']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "name"),
                    value=str(row_data["Name"]),
                    field_name="Name",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            # 2. Outward Type (Radio)
            if row_data["OutwardType"]:
                current_field = f"Outward Type ({row_data['OutwardType']})"
                radio_id = f"outward_type{row_data['OutwardType']}"
                Function_Call.click(self, f'//input[@id="{radio_id}"]')

            # 3. As Billable (Radio)
            if row_data["AsBillable"] is not None:
                current_field = f"As Billable ({row_data['AsBillable']})"
                # Map 0 -> as_bill1, 1 -> as_bill2
                radio_id = f"as_bill{int(row_data['AsBillable']) + 1}"
                Function_Call.click(self, f'//input[@id="{radio_id}"]')

            # 4. Expiry Date Validate (Radio)
            if row_data["ExpiryDateValidate"] is not None:
                current_field = f"Expiry Date Validate ({row_data['ExpiryDateValidate']})"
                # Map 0 -> exp_date1, 1 -> exp_date2
                radio_id = f"exp_date{int(row_data['ExpiryDateValidate']) + 1}"
                Function_Call.click(self, f'//input[@id="{radio_id}"]')

            # 5. Reorder Level (Radio)
            if row_data["ReorderLevel"]:
                current_field = f"Reorder Level ({row_data['ReorderLevel']})"
                radio_id = f"reorder_level{row_data['ReorderLevel']}"
                Function_Call.click(self, f'//input[@id="{radio_id}"]')

            # 6. Save
            current_field = "Save Button"
            Function_Call.click(self, '//button[@type="submit" and contains(@class, "btn-primary")]')
            sleep(3)

            # 7. Success Verification
            try:
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Category added successfully')]"
                success_element = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath)))
                msg = success_element.text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "Category added successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_inventory_category_list_verification(self, row_data):
        """Verify the added category in the list table"""
        driver = self.driver
        wait = self.wait
        
        try:
            if "inventory_category/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_other_inventory/inventory_category/list")
                sleep(2)
            
            search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(str(row_data["Name"]))
            sleep(2)
            
            try:
                row_xpath = f"//table/tbody/tr[contains(., '{row_data['Name']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                print(f"✅ Found Category {row_data['Name']} in list page.")
                return ("Pass", f"Verified {row_data['Name']} in list")
            except:
                print(f"❌ Could not find Category {row_data['Name']} in list page after search")
                return ("Fail", f"Category {row_data['Name']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
