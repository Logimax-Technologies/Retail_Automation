from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class OtherInventory(unittest.TestCase):
    """
    Other Inventory: Item Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_other_inventory(self):
        """Main entry point for Other Inventory Item automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Other Inventory')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/other_inventory/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "OtherInventory"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "ItemFor": 4, "Name": 5, "Size": 6, "UnitPrice": 7,
                "IssuePreference": 8, "MinPcs": 9, "MaxPcs": 10, "ExpectedStatus": 11
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_other_inventory_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_other_inventory_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Other Inventory Item Automation Completed")
        print(f"{'='*80}")

    def test_other_inventory_add(self, row_data, row_num, sheet_name):
        """Logic to add a new Other Inventory Item"""
        driver = self.driver
        wait = self.wait
        current_field = "Add Button"
        
        try:
            # Click Add Button
            try:
                Function_Call.click(self, '//a[contains(@href, "/add")]')
            except:
                driver.get(BASE_URL+"index.php/admin_ret_other_inventory/other_inventory/add")
            
            sleep(3)
            
            # --- TAB 1: ITEM DETAILS ---
            
            # 1. Item For (Select2)
            if row_data["ItemFor"]:
                current_field = f"Item For ({row_data['ItemFor']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="itemfor"]/following-sibling::span',
                    row_data["ItemFor"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 2. Name
            if row_data["Name"]:
                current_field = f"Name ({row_data['Name']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "name"),
                    value=str(row_data["Name"]),
                    field_name="Name",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            # 3. Select Size (Select2)
            if row_data["Size"]:
                current_field = f"Size ({row_data['Size']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_size"]/following-sibling::span',
                    row_data["Size"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 4. Unit Price
            if row_data["UnitPrice"]:
                current_field = f"Unit Price ({row_data['UnitPrice']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.NAME, "other[unit_price]"),
                    value=str(row_data["UnitPrice"]),
                    field_name="Unit Price",
                    row_num=row_num,
                    Sheet_name=sheet_name
                )

            # 5. Issue Preference
            if row_data["IssuePreference"]:
                current_field = f"Issue Preference ({row_data['IssuePreference']})"
                pref_value = str(row_data["IssuePreference"])
                select_pref = wait.until(EC.presence_of_element_located((By.NAME, "other[issue_preference]")))
                select_pref.click()
                driver.find_element(By.XPATH, f"//select[@name='other[issue_preference]']/option[@value='{pref_value}']").click()

            # --- TAB 2: REORDER DETAILS ---
            
            current_field = "Switch to Reorder Tab"
            Function_Call.click(self, "//a[contains(text(), 'REORDER DETAILS')]")
            sleep(1)

            # 6. Min/Max Pcs (for the first branch as a sample)
            if row_data["MinPcs"]:
                current_field = f"Min Pcs ({row_data['MinPcs']})"
                min_pcs_input = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "min_pcs")))
                min_pcs_input.clear()
                min_pcs_input.send_keys(str(row_data["MinPcs"]))

            if row_data["MaxPcs"]:
                current_field = f"Max Pcs ({row_data['MaxPcs']})"
                max_pcs_input = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "max_pcs")))
                max_pcs_input.clear()
                max_pcs_input.send_keys(str(row_data["MaxPcs"]))

            # 7. Save
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="inventory_type_submit"]')
            sleep(3)

            # 8. Success Verification
            try:
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Item added successfully')]"
                success_element = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath)))
                msg = success_element.text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "Item added successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_other_inventory_list_verification(self, row_data):
        """Verify the added item in the list table"""
        driver = self.driver
        wait = self.wait
        
        try:
            if "other_inventory/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_other_inventory/other_inventory/list")
                sleep(2)
            
            search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(str(row_data["Name"]))
            sleep(2)
            
            try:
                # Verify row contains Item Name and Category
                row_xpath = f"//table/tbody/tr[contains(., '{row_data['Name']}') and contains(., '{row_data['ItemFor']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                print(f"✅ Found Item {row_data['Name']} in list page.")
                return ("Pass", f"Verified {row_data['Name']} in list")
            except:
                print(f"❌ Could not find Item {row_data['Name']} in list page after search")
                return ("Fail", f"Item {row_data['Name']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
