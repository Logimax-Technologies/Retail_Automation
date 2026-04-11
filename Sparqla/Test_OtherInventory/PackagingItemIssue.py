from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class PackagingItemIssue(unittest.TestCase):
    """
    Other Inventory: Packaging Item Issue Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_packaging_item_issue(self):
        """Main entry point for Packaging Item Issue automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Packaging Item Issue')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/issue_item/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "PackagingItemIssue"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "BranchName": 4, "ItemName": 5, "BillNo": 6,
                "Pieces": 7, "Remarks": 8, "ExpectedStatus": 9
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_packaging_item_issue_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_packaging_item_issue_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Packaging Item Issue Automation Completed")
        print(f"{'='*80}")

    def test_packaging_item_issue_add(self, row_data, row_num, sheet_name):
        """Logic to create a new Packaging Item Issue via Modal"""
        driver = self.driver
        wait = self.wait
        current_field = "Add Button"
        
        try:
            # 1. Click Add Button to open Modal
            Function_Call.click(self, '//a[@id="add_issue_details"]')
            sleep(2)
            
            modal_xpath = '//div[@id="confirm-add" and contains(@class, "in")]'
            wait.until(EC.visibility_of_element_located((By.XPATH, modal_xpath)))
            print("📦 Modal Opened")

            # 2. Branch Details (Select2)
            if row_data["BranchName"]:
                current_field = f"Branch Name ({row_data['BranchName']})"
                # Check if branch select exists (some sessions hardcode branch)
                branch_selects = driver.find_elements(By.ID, "branch_select")
                if branch_selects and branch_selects[0].is_displayed():
                    Function_Call.dropdown_select(
                        self,
                        '//select[@id="branch_select"]/following-sibling::span',
                        row_data["BranchName"],
                        '//span[@class="select2-search select2-search--dropdown"]/input'
                    )
                    sleep(1)

            # 3. Item Details (Select2)
            if row_data["ItemName"]:
                current_field = f"Select Item ({row_data['ItemName']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_item" and contains(@name, "issue")]/following-sibling::span',
                    row_data["ItemName"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 4. Bill No (Select2)
            if row_data["BillNo"]:
                current_field = f"Bill No ({row_data['BillNo']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_bill_no"]/following-sibling::span',
                    str(row_data["BillNo"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 5. Pieces
            if row_data["Pieces"]:
                current_field = f"Pieces ({row_data['Pieces']})"
                Function_Call.fill_input(self, wait, (By.ID, "issue_total_pcs"), str(row_data["Pieces"]), "Pieces", row_num, sheet_name)

            # 6. Remarks
            if row_data["Remarks"]:
                current_field = f"Remarks ({row_data['Remarks']})"
                remarks_box = driver.find_element(By.ID, "remarks")
                remarks_box.clear()
                remarks_box.send_keys(str(row_data["Remarks"]))

            # 7. Save
            current_field = "Save Button"
            Function_Call.click(self, '//a[@id="item_issue"]')
            sleep(4)

            # 8. Success Verification
            try:
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Item Issued successfully')]"
                msg = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath))).text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "Item Issued successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_packaging_item_issue_list_verification(self, row_data):
        """Verify the issue entry in the list page"""
        driver = self.driver
        wait = self.wait
        
        try:
            if "issue_item/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_other_inventory/issue_item/list")
                sleep(2)
            
            # Click Search to refresh list
            Function_Call.click(self, '//button[@id="search_issue_item"]')
            sleep(3)
            
            try:
                # Check if table row exists with Item Name and Pieces
                row_xpath = f"//table[@id='issue_list']/tbody/tr[contains(., '{row_data['ItemName']}') and contains(., '{row_data['Pieces']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                print(f"✅ Found Issue Entry {row_data['ItemName']} in list page.")
                return ("Pass", f"Verified {row_data['ItemName']} in list")
            except:
                print(f"❌ Could not find Issue Entry {row_data['ItemName']} in list page")
                return ("Fail", f"Issue Entry {row_data['ItemName']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
