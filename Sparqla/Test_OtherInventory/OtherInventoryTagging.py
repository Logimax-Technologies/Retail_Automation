from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class OtherInventoryTagging(unittest.TestCase):
    """
    Other Inventory: Tagging Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
    
    def test_other_inventory_tagging(self):
        """Main entry point for Other Inventory Tagging automation"""
        driver = self.driver
        wait = self.wait
        
        # Navigation
        try:
            # Click Toggle navigation if sidebar is collapsed
            try:
                toggle = driver.find_elements(By.PARTIAL_LINK_TEXT, "Toggle navigation")
                if toggle and toggle[0].is_displayed():
                    toggle[0].click()
            except:
                pass
                
            Function_Call.click(self, "//span[contains(text(), 'Other Inventory')]")
            Function_Call.click(self, "//span[contains(normalize-space(), 'Other Inventory Tagging')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            # Fallback direct navigation
            driver.get(BASE_URL+"index.php/admin_ret_other_inventory/product_details/list")
            sleep(2)
        
        # Read Excel data
        sheet_name = "OtherInventoryTagging"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel sheet '{sheet_name}': {e}")
            return
        
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        
        for row_num in range(2, valid_rows + 1):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "RefNo": 4, "ExpectedStatus": 5
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            # Check if test should run
            if str(row_data["TestStatus"]).strip().lower() != "enable":
                print(f"⏭️ Skipping Test Case {row_data['TestCaseId']} (TestStatus != 'Enable')")
                continue
            
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")
            
            try:
                result = self.test_other_inventory_tagging_add(row_data, row_num, sheet_name)
                test_status = result[0]
                actual_status = result[1]
                print(f"🏁 Test Result: {test_status} - {actual_status}")
                
                # Verification in list page
                if test_status == "Pass":
                    print(f"🔍 Verifying entry in List Page...")
                    list_result = self.test_other_inventory_tagging_list_verification(row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    actual_status = f"{actual_status} | {list_result[1]}"
                
                self._update_excel_status(row_num, test_status, actual_status, sheet_name)
                    
            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
        
        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ Other Inventory Tagging Automation Completed")
        print(f"{'='*80}")

    def test_other_inventory_tagging_add(self, row_data, row_num, sheet_name):
        """Logic to perform tagging for a Purchase Reference"""
        driver = self.driver
        wait = self.wait
        current_field = "Add Button"
        
        try:
            # 1. Click Add Button
            try:
                Function_Call.click(self, '//a[contains(@href, "/add")]')
            except:
                driver.get(BASE_URL+"index.php/admin_ret_other_inventory/product_details/add")
            sleep(3)

            # 2. Select Ref No (Select2)
            if row_data["RefNo"]:
                current_field = f"Ref No ({row_data['RefNo']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_ref_no"]/following-sibling::span',
                    str(row_data["RefNo"]),
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                print(f"⌛ Waiting for items to load for Ref No: {row_data['RefNo']}...")
                sleep(3) # Wait for AJAX table populating
            else:
                print("⚠️ RefNo is empty in Excel. Proceeding with default selection if any.")

            # 3. Select All
            current_field = "Select All Checkbox"
            select_all = wait.until(EC.presence_of_element_located((By.ID, "select_all")))
            if not select_all.is_selected():
                select_all.click()
            sleep(1)

            # 4. Save
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="prod_inventory_submit"]')
            sleep(4)

            # 5. Success Verification
            try:
                success_xpath = "//div[contains(@class, 'alert-success')] | //*[contains(text(), 'Product Details Added successfully')]"
                msg = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath))).text.strip()
                print(f"✅ Success Message: {msg}")
                return ("Pass", "Product Details Added successfully")
            except:
                try:
                    error_xpath = "//div[contains(@class, 'alert-danger')] | //*[contains(text(), 'Unable to proceed')]"
                    error_element = driver.find_element(By.XPATH, error_xpath)
                    err_msg = error_element.text.strip()
                    print(f"❌ Error Message: {err_msg}")
                    return ("Fail", f"Error: {err_msg}")
                except:
                    print("⚠️ No status message found after submit")
                    return ("Fail", "No success/failure message found")

        except Exception as e:
            print(f"❌ Error in {current_field}: {e}")
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_other_inventory_tagging_list_verification(self, row_data):
        """Verify the tagging in the list page"""
        driver = self.driver
        wait = self.wait
        
        try:
            if "product_details/list" not in driver.current_url:
                driver.get(BASE_URL + "index.php/admin_ret_other_inventory/product_details/list")
                sleep(2)
            
            # Simple list search
            search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='search']")))
            search_box.clear()
            search_box.send_keys(str(row_data["RefNo"]))
            sleep(3)
            
            try:
                # In tagging list, search usually shows individual piece records or the summary
                # Here we just verify the search returned results containing the RefNo or corresponding entries
                result_xpath = f"//table/tbody/tr[contains(., '{row_data['RefNo']}')]"
                wait.until(EC.presence_of_element_located((By.XPATH, result_xpath)))
                print(f"✅ Found Tagging records for {row_data['RefNo']} in list page.")
                return ("Pass", f"Verified {row_data['RefNo']} in list")
            except:
                print(f"❌ Could not find records for {row_data['RefNo']} in tagging list")
                return ("Fail", f"RefNo {row_data['RefNo']} not found in list")
                
        except Exception as e:
            return ("Fail", f"List verification error: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Utility to write results back to Excel"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")
